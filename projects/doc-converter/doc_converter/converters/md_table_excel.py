"""Markdown 表格 → Excel 提取转换器"""

import re
from pathlib import Path
from .base import BaseConverter, ConvertResult, register


def extract_md_tables(text: str) -> list[list[list[str]]]:
    """
    从 Markdown 文本中提取所有表格。
    返回: [table1, table2, ...], 每个 table 是 [[row], [row], ...]
    """
    tables = []
    lines = text.split("\n")
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        # 检测表格起始 (含 | 的行，下一行是分隔行)
        if "|" in line and i + 1 < len(lines):
            next_line = lines[i + 1].strip()
            if re.match(r"^\|[\s\-:|]+\|$", next_line):
                table = []
                while i < len(lines) and "|" in lines[i].strip():
                    row_text = lines[i].strip()
                    # 跳过分隔行
                    if not re.match(r"^\|[\s\-:|]+\|$", row_text):
                        cells = [c.strip() for c in row_text.strip("|").split("|")]
                        table.append(cells)
                    i += 1
                if table:
                    tables.append(table)
                continue
        i += 1
    return tables


@register
class MdTableToExcel(BaseConverter):
    name = "md-table-excel"
    source_formats = ["md", "markdown"]
    target_formats = ["xlsx"]
    description = "从 Markdown 中提取表格转 Excel"
    dependencies = ["openpyxl"]

    def convert(self, input_path: Path, output_path: Path, **options) -> ConvertResult:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        text = input_path.read_text(encoding="utf-8")

        # 只在 extract=table 或默认 md→xlsx 时提取表格
        tables = extract_md_tables(text)
        if not tables:
            return ConvertResult(False, message="未找到 Markdown 表格")

        wb = Workbook()
        wb.remove(wb.active)  # 删除默认sheet

        for t_idx, table in enumerate(tables):
            ws = wb.create_sheet(title=f"表格{t_idx + 1}")
            for r_idx, row in enumerate(table):
                for c_idx, cell_text in enumerate(row):
                    cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=cell_text)
                    if r_idx == 0:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E8E8E8", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")

            # 自动列宽
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        wb.save(str(output_path))
        return ConvertResult(
            True, output_path=output_path,
            message=f"已提取 {len(tables)} 个表格到 Excel"
        )
