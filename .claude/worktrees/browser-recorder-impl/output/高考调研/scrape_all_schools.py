#!/usr/bin/env python3
"""安徽高考历年录取数据抓取工具

功能：
  - 自定义筛选条件：年份、科类、院校所在地、层次、院校特性、指定院校、指定专业、排名
  - 多枚举参数循环抓取（年份、科类、层次、院校特性各值组合）
  - 院校所在地利用页面多选能力，不需循环
  - 即时写入 Excel，每条院校获取后立即保存
  - 断点续传：记录当前组合+页码+行号，中断后自动继续
  - 频率限制检测与自动重试
  - 自动检测浏览器状态，未打开则启动并等待登录

用法：
  python3 scrape_all_schools.py                # 使用默认配置
  python3 scrape_all_schools.py --test         # 仅抓取每个组合的第一页（测试）
  python3 scrape_all_schools.py --config config.json  # 使用自定义配置文件
"""

import json
import time
import os
import sys
import argparse
import traceback
import subprocess
import requests
import openpyxl
from openpyxl.styles import Font, PatternFill
from itertools import product

# ==================== 路径常量 ====================

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# 项目根目录 = ai-exploration/
PROJECT_ROOT = os.path.dirname(os.path.dirname(SCRIPT_DIR))
SESSION_HELPER = os.path.join(
    PROJECT_ROOT, "plugins/my-toolkit/skills/web-explorer/scripts/session_helper.sh"
)
SESSION_FILE = "/tmp/web-explorer-session.json"
TARGET_URL = "https://xgk.ahzsks.cn/lnsjcx/lnsjcx"
TARGET_URL_XGK = "https://xgk.ahzsks.cn/lnsjcx/xgk/lnsjcx"  # 新高考（2024+）

OUTPUT_DIR = os.path.join(PROJECT_ROOT, "output")
EXCEL_PATH = os.path.join(OUTPUT_DIR, "安徽高考录取数据.xlsx")
CHECKPOINT_PATH = os.path.join(OUTPUT_DIR, "scrape_checkpoint.json")

# ==================== 频率控制 ====================

RATE_LIMIT_SLEEP = 5
EXPAND_SLEEP = 2.0
COLLAPSE_SLEEP = 1.0
PAGE_SLEEP = 3.0
FILTER_SLEEP = 3.0  # 切换筛选条件后等待
MAX_RETRY = 3

# ==================== 筛选参数定义 ====================

# 枚举类型参数（需要循环的）
PARAM_YEAR = {
    "name": "年份",
    "type": "enum",
    "values": ["2021", "2022", "2023", "2024", "2025"],
    "default": ["2025"],
}
# 科类：2021-2023 用文史/理工，2024-2025 用历史科目组合/物理科目组合
PARAM_CATEGORY = {
    "name": "科类",
    "type": "enum",
    # 旧模式(2021-2023)
    "values_old": ["文史", "理工"],
    # 新模式(2024-2025)，首选科目与科类联动
    "values_new": ["物理科目组合", "历史科目组合"],
    "default": ["物理科目组合"],
}
# 再选科目：仅2024-2025年使用，必选两项，用组合循环
PARAM_RESELECT = {
    "name": "再选科目",
    "type": "combination",  # C(4,2)=6种组合
    "values": ["化学", "生物学", "思想政治", "地理"],
    "default": [["化学", "生物学"]],  # list[list[str]], 每项是一个两选组合
}
PARAM_LEVEL = {
    "name": "层次",
    "type": "enum",
    "values": ["全部", "本科", "高职（专科）"],
    "default": ["全部"],
}
PARAM_FEATURE = {
    "name": "院校特性",
    "type": "enum",
    "values": ["全部", "\"双一流\"建设高校"],
    "default": ["全部"],
}

# 多选参数（页面本身支持多选，不需循环）
PARAM_LOCATION = {
    "name": "院校所在地",
    "type": "multiselect",
    "values": [
        "全部", "北京", "天津", "河北", "山西", "内蒙古", "辽宁", "吉林",
        "黑龙江", "上海", "江苏", "浙江", "安徽", "福建", "江西", "山东",
        "河南", "湖北", "湖南", "广东", "广西", "海南", "重庆", "四川",
        "贵州", "云南", "西藏", "陕西", "甘肃", "青海", "宁夏", "新疆", "香港",
    ],
    "default": [],  # 空=不筛选（全部）
}

# 字符串参数
PARAM_SCHOOL = {
    "name": "指定院校",
    "type": "str",
    "default": "",
}
PARAM_MAJOR = {
    "name": "指定专业",
    "type": "str",
    "default": "",
}

# 排名参数（整数范围）
PARAM_RANK_FROM = {
    "name": "排名从",
    "type": "int",
    "default": None,
}
PARAM_RANK_TO = {
    "name": "排名至",
    "type": "int",
    "default": None,
}

# 默认配置
DEFAULT_CONFIG = {
    "year": PARAM_YEAR["default"],              # list[str]: 要抓取的年份
    "category": PARAM_CATEGORY["default"],      # list[str]: 要抓取的科类
    "reselect": PARAM_RESELECT["default"],      # list[list[str]]: 再选科目组合（仅2024+）
    "location": PARAM_LOCATION["default"],      # list[str]: 院校所在地（多选）
    "level": PARAM_LEVEL["default"],            # list[str]: 层次
    "feature": PARAM_FEATURE["default"],        # list[str]: 院校特性
    "school": PARAM_SCHOOL["default"],          # str: 指定院校（模糊）
    "major": PARAM_MAJOR["default"],            # str: 指定专业（模糊）
    "rank_from": PARAM_RANK_FROM["default"],    # int|None: 排名起始
    "rank_to": PARAM_RANK_TO["default"],        # int|None: 排名截止
}

def is_new_gaokao(year):
    """判断是否为新高考模式（2024+）"""
    return int(year) >= 2024

BASE_URL = None  # 动态获取


# ==================== 浏览器管理 ====================

def get_session_port():
    """从会话文件获取端口"""
    if os.path.exists(SESSION_FILE):
        try:
            with open(SESSION_FILE) as f:
                info = json.load(f)
            return info.get("port")
        except:
            pass
    return None


def is_browser_running():
    """检测浏览器服务是否运行"""
    port = get_session_port()
    if port is None:
        return False
    try:
        r = requests.get(f"http://localhost:{port}/session/status", timeout=3)
        data = r.json()
        return data.get("success") and data.get("data", {}).get("status") == "running"
    except:
        return False


def is_browser_healthy():
    """深度检测浏览器实例是否真正可用（非僵尸状态）

    /session/status 可能返回 running 但底层浏览器已断开，
    这里通过调用 /url 验证浏览器实例的实际响应能力。
    """
    port = get_session_port()
    if port is None:
        return False
    try:
        r = requests.get(f"http://localhost:{port}/url", timeout=5)
        if r.status_code == 200:
            data = r.json()
            return data.get("success", False)
        return False
    except:
        return False


def kill_stale_session():
    """清理僵尸会话：终止旧进程并删除会话文件"""
    if os.path.exists(SESSION_FILE):
        try:
            with open(SESSION_FILE) as f:
                info = json.load(f)
            pid = info.get("pid")
            if pid:
                print(f"[浏览器] 终止僵尸进程 PID={pid}...")
                try:
                    os.kill(pid, 9)
                    time.sleep(1)
                except ProcessLookupError:
                    pass
        except:
            pass
        os.remove(SESSION_FILE)
        print("[浏览器] 已清理旧会话文件")


def start_browser_and_wait_login():
    """启动浏览器并等待用户登录"""
    print("[浏览器] 启动中...")
    result = subprocess.run(
        ["bash", SESSION_HELPER, "start"],
        capture_output=True, text=True, timeout=30,
        cwd=PROJECT_ROOT  # session_helper.sh 内部使用 .venv 相对路径
    )
    if result.returncode != 0:
        print(f"[ERROR] 浏览器启动失败: {result.stderr.strip()}")
        sys.exit(1)
    print(f"[浏览器] {result.stdout.strip()}")

    port = get_session_port()
    if port is None:
        print("[ERROR] 无法获取浏览器端口")
        sys.exit(1)

    global BASE_URL
    BASE_URL = f"http://localhost:{port}"

    # 导航到目标页面
    print(f"[浏览器] 导航到 {TARGET_URL}")
    requests.post(
        f"{BASE_URL}/navigate",
        json={"url": TARGET_URL},
        timeout=30
    )
    time.sleep(3)

    # 检查是否在登录页
    r = requests.get(f"{BASE_URL}/url", timeout=10)
    current_url = r.json().get("data", {}).get("url", "")

    if "menu" in current_url or "login" in current_url.lower():
        print("\n" + "=" * 50)
        print("请在浏览器中完成登录")
        print("登录后页面会自动跳转到数据查询页面")
        print("=" * 50)
        input("登录完成后按 Enter 继续...")

        # 确保在正确的页面
        r = requests.get(f"{BASE_URL}/url", timeout=10)
        current_url = r.json().get("data", {}).get("url", "")
        if "lnsjcx" not in current_url:
            print(f"[浏览器] 当前页面: {current_url}，导航到数据查询页...")
            requests.post(f"{BASE_URL}/navigate", json={"url": TARGET_URL}, timeout=30)
            time.sleep(3)

    print("[浏览器] 就绪\n")


def ensure_browser():
    """确保浏览器已打开并可用

    检测逻辑：
    1. 服务进程在运行 + 浏览器实例健康 → 直接复用
    2. 服务进程在运行但实例不健康（僵尸） → 清理后重启
    3. 服务未运行 → 直接启动
    """
    global BASE_URL

    if is_browser_running():
        port = get_session_port()
        BASE_URL = f"http://localhost:{port}"

        if is_browser_healthy():
            # 浏览器确实可用，检查是否在目标页面
            try:
                r = requests.get(f"{BASE_URL}/url", timeout=10)
                current_url = r.json().get("data", {}).get("url", "")
                if "lnsjcx" in current_url:
                    print(f"[浏览器] 已在运行，端口 {port}")
                    return
            except:
                pass
        else:
            # 进程在但浏览器实例已失效，清理僵尸
            print("[浏览器] 检测到僵尸会话（服务在运行但浏览器无响应），正在清理...")
            kill_stale_session()

    start_browser_and_wait_login()


# ==================== JS 操作封装 ====================

def run_js(script):
    """执行页面JS"""
    try:
        r = requests.post(
            f"{BASE_URL}/run_js",
            json={"script": script, "as_expr": True, "timeout": 15000},
            timeout=30
        )
        return r.json()
    except Exception as e:
        print(f"    [WARN] 请求异常: {e}")
        return {"success": False, "error": str(e), "data": None}


def check_rate_limit():
    """检测频率限制提示"""
    js = """(function(){
    var found = false;
    document.querySelectorAll('.el-message').forEach(function(m){
        if(m.textContent.indexOf('频繁') !== -1 || m.textContent.indexOf('稍后') !== -1) found = true;
    });
    document.querySelectorAll('.el-message-box__message').forEach(function(d){
        if(d.textContent.indexOf('频繁') !== -1 || d.textContent.indexOf('稍后') !== -1) found = true;
    });
    document.querySelectorAll('.el-notification').forEach(function(n){
        if(n.textContent.indexOf('频繁') !== -1 || n.textContent.indexOf('稍后') !== -1) found = true;
    });
    return found ? 'rate_limited' : 'ok';
})()"""
    resp = run_js(js)
    if resp is None:
        return False
    return resp.get("success") and (resp.get("data") or {}).get("result") == "rate_limited"


def wait_for_rate_limit():
    """检测并等待频率限制"""
    if check_rate_limit():
        print(f"\n    [RATE LIMIT] 等待 {RATE_LIMIT_SLEEP}s...", end="", flush=True)
        time.sleep(RATE_LIMIT_SLEEP)
        run_js("""(function(){ var btn = document.querySelector('.el-message-box__btns .el-button--primary'); if(btn) btn.click(); })()""")
        time.sleep(1)
        return True
    return False


# ==================== 筛选操作 ====================

def click_radio(label_text, value):
    """点击 radio 按钮（年份、科类、层次、院校特性）"""
    # 用 label 找到对应 form-item，再找到 radio
    js = f"""(function(){{
    var items = document.querySelectorAll('.el-form-item');
    for(var i=0; i<items.length; i++){{
        var label = items[i].querySelector('.el-form-item__label');
        if(label && label.textContent.indexOf('{label_text}') !== -1){{
            var radios = items[i].querySelectorAll('.el-radio');
            for(var j=0; j<radios.length; j++){{
                var rl = radios[j].querySelector('.el-radio__label');
                if(rl && rl.textContent.trim() === '{value}'){{
                    radios[j].click();
                    return 'clicked';
                }}
            }}
            return 'value_not_found';
        }}
    }}
    return 'label_not_found';
}})()"""
    resp = run_js(js)
    result = resp.get("data", {}).get("result", "")
    if result != "clicked":
        print(f"    [WARN] 点击 {label_text}={value} 失败: {result}")
    return result == "clicked"


def set_location(locations):
    """设置院校所在地（多选 checkbox）"""
    if not locations:
        return True

    # 先清除已选
    js_clear = """(function(){
    var items = document.querySelectorAll('.el-form-item');
    for(var i=0; i<items.length; i++){
        var label = items[i].querySelector('.el-form-item__label');
        if(label && label.textContent.indexOf('院校所在地') !== -1){
            var checks = items[i].querySelectorAll('.el-checkbox.is-checked');
            checks.forEach(function(c){ c.click(); });
            return 'cleared:' + checks.length;
        }
    }
    return 'not_found';
})()"""
    run_js(js_clear)
    time.sleep(0.5)

    # 点选目标地区
    for loc in locations:
        js_check = f"""(function(){{
    var items = document.querySelectorAll('.el-form-item');
    for(var i=0; i<items.length; i++){{
        var label = items[i].querySelector('.el-form-item__label');
        if(label && label.textContent.indexOf('院校所在地') !== -1){{
            var checks = items[i].querySelectorAll('.el-checkbox');
            for(var j=0; j<checks.length; j++){{
                var cl = checks[j].querySelector('.el-checkbox__label');
                if(cl && cl.textContent.trim() === '{loc}'){{
                    checks[j].click();
                    return 'clicked';
                }}
            }}
            return 'value_not_found';
        }}
    }}
    return 'label_not_found';
}})()"""
        resp = run_js(js_check)
        result = resp.get("data", {}).get("result", "")
        if result != "clicked":
            print(f"    [WARN] 选择地区 {loc} 失败: {result}")
        time.sleep(0.2)
    return True


def set_text_input(label_text, value):
    """设置文本输入框（指定院校、指定专业）"""
    if not value:
        return True
    js = f"""(function(){{
    var items = document.querySelectorAll('.el-form-item');
    for(var i=0; i<items.length; i++){{
        var label = items[i].querySelector('.el-form-item__label');
        if(label && label.textContent.indexOf('{label_text}') !== -1){{
            var input = items[i].querySelector('.el-form-item__content input.el-input__inner');
            if(input){{
                var nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
                nativeSetter.call(input, '{value}');
                input.dispatchEvent(new Event('input', {{bubbles: true}}));
                input.dispatchEvent(new Event('change', {{bubbles: true}}));
                return 'set';
            }}
            return 'no_input';
        }}
    }}
    return 'label_not_found';
}})()"""
    resp = run_js(js)
    return resp.get("data", {}).get("result") == "set"


def set_rank(rank_from, rank_to):
    """设置排名范围"""
    js = f"""(function(){{
    var items = document.querySelectorAll('.el-form-item');
    for(var i=0; i<items.length; i++){{
        var label = items[i].querySelector('.el-form-item__label');
        if(label && label.textContent.indexOf('排') !== -1 && label.textContent.indexOf('名') !== -1){{
            var inputs = items[i].querySelectorAll('.el-form-item__content input.el-input__inner');
            var nativeSetter = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            if(inputs.length >= 2){{
                if('{rank_from}' !== 'None'){{
                    nativeSetter.call(inputs[0], '{rank_from}');
                    inputs[0].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
                if('{rank_to}' !== 'None'){{
                    nativeSetter.call(inputs[1], '{rank_to}');
                    inputs[1].dispatchEvent(new Event('input', {{bubbles: true}}));
                }}
                return 'set';
            }}
            return 'no_inputs';
        }}
    }}
    return 'label_not_found';
}})()"""
    run_js(js)


def click_query_button():
    """点击查询按钮"""
    js = """(function(){
    var btns = document.querySelectorAll('.el-button--primary');
    for(var i=0; i<btns.length; i++){
        if(btns[i].textContent.trim() === '查询'){
            btns[i].click();
            return 'clicked';
        }
    }
    return 'not_found';
})()"""
    resp = run_js(js)
    return resp.get("data", {}).get("result") == "clicked"


def set_reselect_subjects(subjects):
    """设置再选科目（el-checkbox-button，必选两项）

    Args:
        subjects: list[str], 如 ["化学", "生物学"]
    """
    if not subjects:
        return True

    # 先取消所有已选
    js_clear = """(function(){
    var items = document.querySelectorAll('.el-form-item');
    for(var i=0; i<items.length; i++){
        var label = items[i].querySelector('.el-form-item__label');
        if(label && label.textContent.indexOf('再选科目') !== -1){
            var checks = items[i].querySelectorAll('.el-checkbox-button.is-checked');
            checks.forEach(function(c){ c.click(); });
            return 'cleared:' + checks.length;
        }
    }
    return 'not_found';
})()"""
    run_js(js_clear)
    time.sleep(0.3)

    # 选中目标科目
    for subj in subjects:
        js_check = f"""(function(){{
    var items = document.querySelectorAll('.el-form-item');
    for(var i=0; i<items.length; i++){{
        var label = items[i].querySelector('.el-form-item__label');
        if(label && label.textContent.indexOf('再选科目') !== -1){{
            var checks = items[i].querySelectorAll('.el-checkbox-button');
            for(var j=0; j<checks.length; j++){{
                var inner = checks[j].querySelector('.el-checkbox-button__inner');
                if(inner && inner.textContent.trim() === '{subj}'){{
                    checks[j].click();
                    return 'clicked';
                }}
            }}
            return 'value_not_found';
        }}
    }}
    return 'label_not_found';
}})()"""
        resp = run_js(js_check)
        result = resp.get("data", {}).get("result", "")
        if result != "clicked":
            print(f"    [WARN] 选择再选科目 {subj} 失败: {result}")
        time.sleep(0.2)
    return True


def apply_filters(year, category, reselect, location, level, feature, school, major, rank_from, rank_to):
    """设置所有筛选条件并点击查询

    Args:
        year: str, 年份
        category: str, 科类
        reselect: list[str], 再选科目（仅新高考2024+使用）
        其余参数同旧版
    """
    reselect_str = f" 再选={'+'.join(reselect)}" if reselect else ""
    print(f"  [筛选] 年份={year} 科类={category}{reselect_str} 层次={level} 院校特性={feature}")

    # 新旧高考使用不同页面，需要导航到对应 URL
    target = TARGET_URL_XGK if is_new_gaokao(year) else TARGET_URL
    r = requests.get(f"{BASE_URL}/url", timeout=10)
    current_url = r.json().get("data", {}).get("url", "")
    if is_new_gaokao(year) and "/xgk/" not in current_url:
        print(f"  [导航] 切换到新高考页面")
        requests.post(f"{BASE_URL}/navigate", json={"url": target}, timeout=30)
        time.sleep(3)
    elif not is_new_gaokao(year) and "/xgk/" in current_url:
        print(f"  [导航] 切换到旧高考页面")
        requests.post(f"{BASE_URL}/navigate", json={"url": target}, timeout=30)
        time.sleep(3)

    # 先切换年份（可能触发表单结构变化）
    click_radio("选择年份", year)
    time.sleep(0.5)

    # 科类
    click_radio("科", category)
    time.sleep(0.3)

    # 新高考模式下设置再选科目
    if is_new_gaokao(year) and reselect:
        set_reselect_subjects(reselect)
        time.sleep(0.3)

    click_radio("层", level)
    time.sleep(0.3)
    click_radio("院校特性", feature)
    time.sleep(0.3)

    if location:
        set_location(location)
        time.sleep(0.3)
    if school:
        set_text_input("指定院校", school)
        time.sleep(0.3)
    if major:
        set_text_input("指定专业", major)
        time.sleep(0.3)
    if rank_from is not None or rank_to is not None:
        set_rank(rank_from, rank_to)
        time.sleep(0.3)

    click_query_button()
    time.sleep(FILTER_SLEEP)
    wait_for_rate_limit()


# ==================== 数据提取 ====================

def get_total_pages():
    """获取总页数，带重试"""
    for attempt in range(MAX_RETRY):
        js = """(function(){
    var pages = document.querySelectorAll('.el-pagination .el-pager li');
    if(pages.length > 0) return pages[pages.length - 1].textContent.trim();
    return '0';
})()"""
        resp = run_js(js)
        if resp and resp.get("success") and resp.get("data"):
            try:
                return int(resp["data"]["result"])
            except:
                pass
        time.sleep(2)
    return 0


def get_current_page_number():
    """获取当前页码"""
    js = """(function(){
    var active = document.querySelector('.el-pagination .el-pager li.active');
    if(active) return active.textContent.trim();
    return '1';
})()"""
    resp = run_js(js)
    if resp and resp.get("success") and resp.get("data"):
        try:
            return int(resp["data"]["result"])
        except:
            pass
    return 1


def get_current_page_schools():
    """获取当前页的院校列表"""
    js = """(function(){
    var tbody = document.querySelector('.el-table__body tbody');
    if(tbody === null) return JSON.stringify([]);
    var rows = tbody.querySelectorAll(':scope > tr');
    var results = [];
    for(var i=0; i<rows.length; i++){
        var row = rows[i];
        if(row.classList.contains('el-table__row') && row.querySelector('.el-table__expanded-cell') === null){
            var tds = row.querySelectorAll('td');
            if(tds.length >= 7){
                results.push({
                    seq: tds[1].textContent.trim(),
                    name: tds[2].textContent.trim(),
                    batch: tds[3].textContent.trim(),
                    subject_req: tds[4].textContent.trim(),
                    max_score: tds[5].textContent.trim(),
                    min_score: tds[6].textContent.trim(),
                    remark: tds.length >= 8 ? tds[7].textContent.trim() : ''
                });
            } else if(tds.length >= 6){
                results.push({
                    seq: tds[1].textContent.trim(),
                    name: tds[2].textContent.trim(),
                    batch: tds[3].textContent.trim(),
                    subject_req: '',
                    max_score: tds[4].textContent.trim(),
                    min_score: tds[5].textContent.trim(),
                    remark: ''
                });
            }
        }
    }
    return JSON.stringify(results);
})()"""
    resp = run_js(js)
    if resp.get("success") and resp.get("data", {}).get("result"):
        try:
            return json.loads(resp["data"]["result"])
        except:
            pass
    return []


def expand_and_get_majors(row_index, retry=0):
    """展开某行获取专业数据"""
    js_click = f"""(function(){{
    var icons = document.querySelectorAll('.el-table__body tbody > tr.el-table__row .el-table__expand-icon');
    if(icons.length > {row_index}){{ icons[{row_index}].click(); return 'clicked'; }}
    return 'no_icon';
}})()"""

    resp = run_js(js_click)
    if not (resp.get("success") and resp.get("data", {}).get("result") == "clicked"):
        if retry < MAX_RETRY:
            time.sleep(RATE_LIMIT_SLEEP)
            return expand_and_get_majors(row_index, retry + 1)
        return []

    time.sleep(EXPAND_SLEEP)

    if wait_for_rate_limit():
        run_js(js_click)
        time.sleep(COLLAPSE_SLEEP)
        if retry < MAX_RETRY:
            return expand_and_get_majors(row_index, retry + 1)
        return []

    js_get = """(function(){
    var expanded = document.querySelectorAll('.el-table__expanded-cell');
    if(expanded.length === 0) return JSON.stringify([]);
    var cell = expanded[expanded.length - 1];
    var trs = cell.querySelectorAll('.el-table__body tbody tr');
    var results = [];
    for(var i=0; i<trs.length; i++){
        var tds = trs[i].querySelectorAll('td');
        if(tds.length >= 7){
            results.push({
                year: tds[0].textContent.trim(),
                major: tds[1].textContent.trim(),
                subject_req: tds[2].textContent.trim(),
                max_score: tds[3].textContent.trim(),
                avg_score: tds[4].textContent.trim(),
                min_score: tds[5].textContent.trim(),
                remark: tds[6].textContent.trim()
            });
        } else if(tds.length >= 6){
            results.push({
                year: tds[0].textContent.trim(),
                major: tds[1].textContent.trim(),
                subject_req: tds[2].textContent.trim(),
                max_score: tds[3].textContent.trim(),
                avg_score: tds[4].textContent.trim(),
                min_score: tds[5].textContent.trim(),
                remark: ''
            });
        }
    }
    return JSON.stringify(results);
})()"""

    resp = run_js(js_get)
    majors = []
    if resp.get("success") and resp.get("data", {}).get("result"):
        try:
            majors = json.loads(resp["data"]["result"])
        except:
            pass

    run_js(js_click)
    time.sleep(COLLAPSE_SLEEP)
    wait_for_rate_limit()

    return majors


def go_to_next_page(retry=0):
    """点击下一页，带重试"""
    js = """(function(){
    var btn = document.querySelector('.el-pagination .btn-next');
    if(btn && btn.disabled === false){ btn.click(); return 'next'; }
    return 'end';
})()"""
    resp = run_js(js)
    success = resp.get("success") and resp.get("data", {}).get("result") == "next"

    if success:
        time.sleep(PAGE_SLEEP)
        if wait_for_rate_limit():
            time.sleep(2)
    elif retry < MAX_RETRY:
        print(f"    [翻页重试 {retry+1}/{MAX_RETRY}] 等待 {RATE_LIMIT_SLEEP}s...")
        time.sleep(RATE_LIMIT_SLEEP)
        wait_for_rate_limit()
        return go_to_next_page(retry + 1)

    return success


# ==================== Excel 管理 ====================

def init_excel(path):
    """初始化 Excel"""
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "院校录取分数"
    headers1 = ["年份", "科类", "再选科目", "层次", "院校特性", "序号", "院校名称",
                "批次", "选考科目要求", "录取最高分/排名", "录取最低分/排名", "说明"]
    ws1.append(headers1)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")

    ws2 = wb.create_sheet("专业录取明细")
    headers2 = ["筛选年份", "筛选科类", "再选科目", "筛选层次", "筛选院校特性",
                "院校名称", "批次", "年份", "专业名称", "选考科目要求",
                "录取最高分/排名", "录取平均分", "录取最低分/排名", "说明"]
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4472C4")

    wb.save(path)
    return wb


def append_to_excel(wb, path, combo, school, majors):
    """即时追加数据"""
    year, category, reselect, level, feature = combo
    reselect_str = "+".join(reselect) if reselect else ""

    ws1 = wb["院校录取分数"]
    ws1.append([year, category, reselect_str, level, feature,
                school["seq"], school["name"], school["batch"],
                school.get("subject_req", ""),
                school["max_score"], school["min_score"],
                school.get("remark", "")])

    ws2 = wb["专业录取明细"]
    for m in majors:
        ws2.append([year, category, reselect_str, level, feature,
                    school["name"], school["batch"],
                    m.get("year", ""), m.get("major", ""),
                    m.get("subject_req", ""),
                    m.get("max_score", ""), m.get("avg_score", ""),
                    m.get("min_score", ""), m.get("remark", "")])

    wb.save(path)


# ==================== 断点管理 ====================

def save_checkpoint(combo_index, page, row_index):
    """保存断点：当前组合索引 + 页码 + 行号"""
    checkpoint = {
        "combo_index": combo_index,
        "page": page,
        "row_index": row_index,
    }
    with open(CHECKPOINT_PATH, 'w') as f:
        json.dump(checkpoint, f)


def load_checkpoint():
    """加载断点"""
    if os.path.exists(CHECKPOINT_PATH):
        try:
            with open(CHECKPOINT_PATH, 'r') as f:
                return json.load(f)
        except:
            pass
    return None


def clear_checkpoint():
    """清除断点"""
    if os.path.exists(CHECKPOINT_PATH):
        os.remove(CHECKPOINT_PATH)


# ==================== 主流程 ====================

def load_config(config_path):
    """加载配置文件"""
    if config_path and os.path.exists(config_path):
        with open(config_path) as f:
            user_config = json.load(f)
        config = dict(DEFAULT_CONFIG)
        config.update(user_config)
        return config
    return dict(DEFAULT_CONFIG)


def build_combos(config):
    """构建筛选组合列表

    新高考(2024+): (year, category, reselect, level, feature)
    旧高考(2021-2023): (year, category, [], level, feature)

    枚举参数做笛卡尔积，再选科目仅在新高考年份参与组合
    """
    years = config["year"]
    categories = config["category"]
    levels = config["level"]
    features = config["feature"]
    reselects = config.get("reselect", [[]])

    combos = []
    for year in years:
        if is_new_gaokao(year):
            # 新高考：科类+再选科目+层次+特性
            for cat in categories:
                for resel in reselects:
                    for level in levels:
                        for feature in features:
                            combos.append((year, cat, resel, level, feature))
        else:
            # 旧高考：科类+层次+特性，无再选科目
            # 旧模式科类映射
            old_categories = []
            for cat in categories:
                if cat in ("物理科目组合", "理工"):
                    old_categories.append("理工")
                elif cat in ("历史科目组合", "文史"):
                    old_categories.append("文史")
                else:
                    old_categories.append(cat)
            old_categories = list(dict.fromkeys(old_categories))  # 去重
            for cat in old_categories:
                for level in levels:
                    for feature in features:
                        combos.append((year, cat, [], level, feature))

    return combos


def scrape_combo(wb, combo, combo_index, config, test_mode=False, start_page=1, start_row=0):
    """抓取一个筛选组合的所有数据"""
    year, category, reselect, level, feature = combo

    # 设置筛选条件并查询
    apply_filters(
        year=year, category=category, reselect=reselect,
        location=config["location"], level=level, feature=feature,
        school=config["school"], major=config["major"],
        rank_from=config["rank_from"], rank_to=config["rank_to"],
    )

    # 收起所有展开行
    run_js("""(function(){ var icons = document.querySelectorAll('.el-table__expand-icon--expanded'); icons.forEach(function(icon){ icon.click(); }); })()""")
    time.sleep(1)

    total_pages = get_total_pages()
    if total_pages == 0:
        print("  无数据")
        return 0, 0

    print(f"  共 {total_pages} 页")

    # 如果需要从断点页开始，逐页翻到目标页
    if start_page > 1:
        current = get_current_page_number()
        while current < start_page:
            if not go_to_next_page():
                print(f"    翻到第{start_page}页失败")
                break
            current = get_current_page_number()

    total_schools = 0
    total_majors = 0
    end_page = 2 if test_mode else total_pages + 1

    for page in range(start_page, end_page):
        # 动态获取当前页码
        current_page = get_current_page_number()
        print(f"  --- 第 {current_page}/{total_pages} 页 ---")

        schools = get_current_page_schools()
        print(f"    院校数: {len(schools)}")

        row_start = start_row if page == start_page else 0

        for i in range(row_start, len(schools)):
            school = schools[i]
            print(f"    [{i+1}/{len(schools)}] {school['name']} - {school['batch']}", end="", flush=True)

            majors = expand_and_get_majors(i)
            print(f" => {len(majors)} 个专业")

            append_to_excel(wb, EXCEL_PATH, combo, school, majors)
            total_schools += 1
            total_majors += len(majors)

            # 断点：combo_index + 当前实际页码 + 下一行
            save_checkpoint(combo_index, current_page, i + 1)

        # 翻页
        if page < end_page - 1:
            if not go_to_next_page():
                print("    翻页失败，停止当前组合")
                save_checkpoint(combo_index, current_page + 1, 0)
                break

        # 翻页后重置 start_row
        start_row = 0

    return total_schools, total_majors


def main():
    parser = argparse.ArgumentParser(description="安徽高考录取数据抓取")
    parser.add_argument("--test", action="store_true", help="测试模式：每个组合只抓第一页")
    parser.add_argument("--config", type=str, default=None, help="配置文件路径(JSON)")
    parser.add_argument("--clean", action="store_true", help="清除断点和已有数据，重新开始")
    args = parser.parse_args()

    config = load_config(args.config)
    combos = build_combos(config)

    print("=" * 60)
    print("安徽高考录取数据抓取工具")
    print(f"模式: {'测试（仅第一页）' if args.test else '全量'}")
    print(f"筛选组合数: {len(combos)}")
    for i, c in enumerate(combos):
        year, cat, resel, level, feat = c
        resel_str = f" 再选={'+'.join(resel)}" if resel else ""
        print(f"  [{i+1}] 年份={year} 科类={cat}{resel_str} 层次={level} 院校特性={feat}")
    if config["location"]:
        print(f"院校所在地: {', '.join(config['location'])}")
    if config["school"]:
        print(f"指定院校: {config['school']}")
    if config["major"]:
        print(f"指定专业: {config['major']}")
    if config["rank_from"] or config["rank_to"]:
        print(f"排名范围: {config['rank_from']} - {config['rank_to']}")
    print(f"输出: {EXCEL_PATH}")
    print("=" * 60)

    # 清除模式
    if args.clean:
        clear_checkpoint()
        if os.path.exists(EXCEL_PATH):
            os.remove(EXCEL_PATH)
        print("[已清除断点和数据文件]")

    # 确保浏览器可用
    ensure_browser()

    # 检查断点
    checkpoint = load_checkpoint()
    start_combo = 0
    start_page = 1
    start_row = 0

    if checkpoint and not args.clean:
        start_combo = checkpoint["combo_index"]
        start_page = checkpoint["page"]
        start_row = checkpoint["row_index"]
        print(f"\n[断点续传] 从组合{start_combo+1} 第{start_page}页 第{start_row+1}行继续")
        wb = openpyxl.load_workbook(EXCEL_PATH)
    else:
        wb = init_excel(EXCEL_PATH)

    grand_total_schools = 0
    grand_total_majors = 0

    try:
        for ci in range(start_combo, len(combos)):
            combo = combos[ci]
            print(f"\n{'='*40}")
            print(f"组合 [{ci+1}/{len(combos)}]: {combo}")
            print(f"{'='*40}")

            sp = start_page if ci == start_combo else 1
            sr = start_row if ci == start_combo and sp == start_page else 0

            s, m = scrape_combo(wb, combo, ci, config, test_mode=args.test,
                                start_page=sp, start_row=sr)
            grand_total_schools += s
            grand_total_majors += m

        clear_checkpoint()
        print(f"\n{'='*60}")
        print(f"全部完成! 共 {grand_total_schools} 条院校, {grand_total_majors} 条专业")
        print(f"已保存: {EXCEL_PATH}")

    except KeyboardInterrupt:
        print(f"\n\n[手动中断] 断点已保存，再次运行（不带 --clean）将自动续传")
        print(f"当前进度: {grand_total_schools} 条院校, {grand_total_majors} 条专业")

    except Exception as e:
        print(f"\n\n[ERROR] 异常中断: {e}")
        traceback.print_exc()
        print(f"断点已保存，重新运行将自动续传")
        print(f"当前进度: {grand_total_schools} 条院校, {grand_total_majors} 条专业")


if __name__ == "__main__":
    main()
