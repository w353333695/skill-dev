#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成告警规则模板 Excel 文件，可导入 ALERT_RULE_TEMPLATE 模型。

职责划分：
- 脚本负责：读取插件指标定义 → 生成告警条件骨架 → 填充 Excel 框架
- LLM 负责：分析采集脚本语义 → 确定每个指标的告警类型和阈值

两种使用方式：
1. 骨架模式（默认）：
   python generate_alert_rule_template.py <插件目录>
   → 生成 conditions.json（含指标列表但无阈值）+ Excel 框架

2. 填充模式（LLM 填完阈值后）：
   python generate_alert_rule_template.py <插件目录> --conditions-json <已填阈值的JSON>
   → 用 LLM 配置好的条件重新生成完整 Excel
"""

import argparse
import json
import os
import sys

import openpyxl
import yaml

# Excel 列定义 (1-indexed)
# 告警条件策略子字段: E(5) ~ U(21)
COL_CONDITION_NAME = 5       # 条件策略名称格式
COL_ALERT_COUNT = 6          # 告警命中个数
COL_ALERT_INTERVAL = 7       # 告警间隔
COL_ALERT_TIMEOUT = 8        # 告警超时自动恢复（单位：秒）
COL_SUGGESTION = 9           # 告警治理建议
COL_CONDITIONS_OPERATOR = 10 # 复合条件关系
COL_DETECT_WINDOW = 11       # 告警判断窗口
COL_GRANULARITY = 12         # 粒度
COL_CONDITION_DESC = 13      # 描述
COL_AND_CONDITIONS = 14      # AND 告警多条件
COL_BLOCK_TIME = 15          # 屏蔽时间配置
COL_EFFECTIVE_TIME = 16      # 生效时间
COL_EVENT_FILTER = 17        # 事件过滤器
COL_MULTI_CONDITIONS = 18    # 复合告警条件
COL_TEMPLATES = 19           # 模板配置
COL_UPGRADE_CONDITION = 20   # 告警升级条件
COL_RECOVER_COUNT = 21       # 告警恢复个数
# 其他主字段
COL_ALERT_RULE_TYPE = 22     # 告警规则类别
COL_TARGET_STRATEGY = 23     # 监控目标策略 > 目标策略名称格式
COL_RULE_DESC = 24           # 规则描述
# 告警通知策略子字段: AX(50) ~ BD(56)
COL_NOTIFY_NAME = 50         # 通知策略名称格式
COL_NOTIFY_TEMPLATE = 51     # 模版信息
COL_NOTIFY_EVENT_FILTER = 52 # 事件过滤器
COL_NOTIFY_EFFECTIVE_TIME = 53  # 规则生效时间
COL_NOTIFY_FILTER_TYPE = 54  # 事件过滤类型
COL_NOTIFY_REPEAT = 55       # 告警重复时间(分钟)
COL_NOTIFY_DESC = 56         # 描述


def read_json(filepath):
    """读取 JSON 文件"""
    with open(filepath, 'r', encoding='utf-8') as f:
        return json.load(f)


def find_plugin_config(plugin_dir):
    """
    查找并读取插件配置文件。

    Returns:
        tuple: (plugin_config, origin_metrics, alias_metrics, model_name)
    """
    plugin_dir = os.path.abspath(plugin_dir)

    # 读取 plugin.yaml
    yaml_path = os.path.join(plugin_dir, 'plugin.yaml')
    if not os.path.exists(yaml_path):
        print("错误: 找不到 plugin.yaml: {}".format(yaml_path))
        sys.exit(1)
    with open(yaml_path, 'r', encoding='utf-8') as f:
        plugin_config = yaml.safe_load(f)

    # 读取 origin_metric.json
    metric_path = os.path.join(plugin_dir, 'origin_metric.json')
    if not os.path.exists(metric_path):
        print("错误: 找不到 origin_metric.json: {}".format(metric_path))
        sys.exit(1)
    origin_metrics = read_json(metric_path)

    # 读取 alias_metric.json (可选)
    alias_metrics = {}
    alias_path = os.path.join(plugin_dir, 'alias_metric.json')
    if os.path.exists(alias_path):
        for m in read_json(alias_path):
            alias_metrics[m.get('name', '')] = m.get('displayName', '')

    # 查找模型 JSON 获取模型名称
    model_name = ''
    for f in os.listdir(plugin_dir):
        if f.endswith('.json') and f not in (
            'origin_metric.json', 'alias_metric.json', 'metric_set.json'
        ):
            model_data = read_json(os.path.join(plugin_dir, f))
            if isinstance(model_data, list) and model_data:
                model_name = model_data[0].get('name', '')
            elif isinstance(model_data, dict):
                model_name = model_data.get('name', '')
            break

    return plugin_config, origin_metrics, alias_metrics, model_name


def should_skip_metric(key):
    """
    判断是否应跳过该指标。

    纯信息性文本字段（如 error_msg、status_code）不适合生成告警条件。
    其余指标是否适合告警由 LLM 根据 metricType 和语义判断，脚本不硬编码。
    """
    skip_keywords = [
        'error_msg', 'error_message', 'status_code',
        'msg', 'message', 'desc', 'description',
    ]
    return key.lower() in skip_keywords


def build_metric_condition_skeleton(metric_key, data_type, display_name='', metric_type=''):
    """
    构建单个指标的告警条件骨架（不含阈值）。

    仅根据数据类型确定比较方式，阈值由 LLM 分析采集脚本后填写：
    - string 类型或 status 类指标 → unequal（值不等于期望值时告警）
    - double 类型 → bigger_than（默认，LLM 可根据语义改为 smaller_than 等）

    携带 metricType 辅助字段，供 LLM 判断该指标是否适合生成告警。
    counter 类型指标（累加型）通常不适合静态阈值告警，LLM 可据此决定跳过。
    """
    base = {
        "alertCount": 3,
        "alertDims": None,
        "comparators": [],
        "detectWindow": 3,
        "metricName": metric_key,
        "metricDisplayName": display_name or metric_key,
        "metricDataType": data_type,
        "metricType": metric_type,  # 辅助字段：gauge/counter，供 LLM 判断
        "mode": "static",
        "recoverCount": 5,
    }

    if data_type == 'string' or metric_key == 'status' or metric_key == 'metric_status':
        base["comparators"] = [
            {"level": "info", "tolerance": 0, "type": "unequal"},
            {"level": "warning", "tolerance": 0, "type": "unequal"},
            {"level": "critical", "tolerance": 0, "type": "unequal"}
        ]
        base["type"] = "unequal"
    else:
        base["comparators"] = [
            {"level": "info", "tolerance": 0, "type": "bigger_than"},
            {"level": "warning", "tolerance": 0, "type": "bigger_than"},
            {"level": "critical", "tolerance": 0, "type": "bigger_than"}
        ]
        base["recoverCount"] = 0
        base["type"] = "bigger_than"

    return base


def export_conditions_json(conditions, output_path):
    """
    导出条件骨架 JSON 文件，供 LLM 填写阈值后回传。

    JSON 中包含辅助字段（metricDisplayName、metricDataType），
    导入 Excel 时会自动剔除这些非标准字段。
    """
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(conditions, f, ensure_ascii=False, indent=2)
    print("条件骨架已导出: {}".format(output_path))


def import_conditions_json(json_path):
    """
    导入 LLM 已填写阈值的 conditions JSON。

    自动剔除辅助字段（metricDisplayName、metricDataType），
    确保写入 Excel 的 JSON 符合平台格式要求。
    """
    conditions = read_json(json_path)
    # 清理辅助字段
    for c in conditions:
        c.pop('metricDisplayName', None)
        c.pop('metricDataType', None)
        c.pop('metricType', None)
    return conditions


def create_workbook():
    """
    创建带标准表头的 Excel 工作簿。

    表头结构与 ALERT_RULE_TEMPLATE 模型导入格式一致。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet1"

    # Row 1: 主表头
    row1 = {
        1: '模板名称', 2: '资源模型', 3: '规则名称格式', 4: '优先级',
        5: '告警条件策略',       # E-U (17 列合并)
        22: '告警规则类别',
        23: '监控目标策略',
        24: '规则描述',
        25: '告警分组策略',      # Y-AI (11 列合并)
        36: '告警丰富策略',      # AJ-AW (14 列合并)
        50: '告警通知策略',      # AX-BD (7 列合并)
        57: '告警规则',
        58: '_创建者', 59: '_创建时间', 60: '_修改者', 61: '_修改时间',
    }
    for col, val in row1.items():
        ws.cell(row=1, column=col, value=val)

    # Row 2: 子表头
    row2 = {
        # 告警条件策略 (E-U)
        5: '条件策略名称格式',
        6: '告警命中个数',
        7: '告警间隔',
        8: '告警超时自动恢复（单位：秒）',
        9: '告警治理建议',
        10: '复合条件关系',
        11: '告警判断窗口',
        12: '粒度',
        13: '描述',
        14: 'AND 告警多条件',
        15: '屏蔽时间配置',
        16: '生效时间',
        17: '事件过滤器',
        18: '复合告警条件',
        19: '模板配置',
        20: '告警升级条件',
        21: '告警恢复个数',
        # 监控目标策略 (W)
        23: '目标策略名称格式',
        # 告警分组策略 (Y-AI)
        25: '告警分组策略名称',
        26: '模型ID',
        27: '分组类型',
        28: '事件过滤类型',
        29: '分组间隔',
        30: '分组等待',
        31: '屏蔽源事件通知',
        32: '描述',
        33: '分组字段',
        34: '事件过滤器',
        35: '告警升级条件',
        # 告警丰富策略 (AJ-AW)
        36: '告警丰富策略名称',
        37: '模型ID',
        38: '事件过滤类型',
        39: '丰富信息类型',
        40: '描述',
        41: '关联资源定义Id',
        42: '规则类型',
        43: '数据条数限制',
        44: '事件过滤器',
        45: '字段列表',
        46: '匹配规则',
        47: '按维度聚合',
        48: '指标列表',
        49: '排序定义',
        # 告警通知策略 (AX-BD)
        50: '通知策略名称格式',
        51: '模版信息',
        52: '事件过滤器',
        53: '规则生效时间',
        54: '事件过滤类型',
        55: '告警重复时间(分钟)',
        56: '描述',
        # 告警规则
        57: '名称',
    }
    for col, val in row2.items():
        ws.cell(row=2, column=col, value=val)

    # 合并单元格
    for merge_range in [
        'A1:A2', 'B1:B2', 'C1:C2', 'D1:D2',
        'E1:U1',          # 告警条件策略
        'V1:V2',          # 告警规则类别
        'X1:X2',          # 规则描述
        'Y1:AI1',         # 告警分组策略
        'AJ1:AW1',        # 告警丰富策略
        'AX1:BD1',        # 告警通知策略
        'BF1:BF2', 'BG1:BG2', 'BH1:BH2', 'BI1:BI2',
    ]:
        ws.merge_cells(merge_range)

    return wb, ws


def write_data_row(ws, row_num, template_name, model_id, conditions):
    """写入一条告警规则模板数据行。"""

    effective_time = json.dumps(
        {"static": {"dayEndTime": "23:59:59",
                    "dayOfWeek": [1, 2, 3, 4, 5, 6, 0],
                    "dayStartTime": "00:00:00"}, "type": "static"},
        ensure_ascii=False
    )
    event_filter = json.dumps(
        {"filter": {"conditionGroup": [{"conditionList": []}]}, "type": "filter"},
        ensure_ascii=False
    )
    upgrade_condition = json.dumps({"enable": False}, ensure_ascii=False)

    notify_template = json.dumps({
        "contentTemplate": (
            "{{time|ts2str:'%H:%M'}}发生{{levelName}}告警 \n"
            "告警资源：{{target}} \n"
            "所属系统：{{alertDims|mvalue:'appSystemNames'|join:','}} \n"
            "告警信息：{{originContent}} \n"
            "首次发生时间：{{startTime|ts2str:'%Y-%m-%d %H:%M'}} \n"
            "持续时长：{{duration|duration_format:'zh'}} \n"
            "事件详情：http://YOUR-EASYOPS-HOST/next/events/{{eventId}}/detail \n"
            "规则详情：http://YOUR-EASYOPS-HOST/next/events/alert-rule/detail/{{ruleId}}"
        ),
        "recoveryContentTemplate": (
            "{{time|ts2str:'%H:%M'}}{{levelName}}告警已解除 \n"
            "告警资源：{{target}} \n"
            "所属系统：{{alertDims|mvalue:'appSystemNames'|join:','}} \n"
            "告警信息：{{originContent}} \n"
            "首次发生时间：{{startTime|ts2str:'%Y-%m-%d %H:%M'}} \n"
            "持续时长：{{duration|duration_format:'zh'}} \n"
            "事件详情：http://YOUR-EASYOPS-HOST/next/events/{{eventId}}/detail \n"
            "规则详情：http://YOUR-EASYOPS-HOST/next/events/alert-rule/detail/{{ruleId}}"
        ),
    }, ensure_ascii=False)

    notify_effective_time = json.dumps(
        {"calendar": {"dayOfCalendar": "working", "period": "onWork"},
         "static": {"dayEndTime": "23:59:59", "dayOfWeek": [0, 1, 2, 3, 4, 5, 6],
                    "dayStartTime": "00:00:00"},
         "type": "static"},
        ensure_ascii=False
    )

    short_name = template_name.replace("初始化模板", "")

    data = {
        # 基本信息区 (A-D)
        1: template_name,
        2: model_id,
        3: '{objectName}{appSystemName}{serviceSetName}通用告警规则',
        4: 50,
        # 告警条件策略 (E-U)
        COL_CONDITION_NAME: '{objectName}通用阈值策略',
        COL_ALERT_COUNT: 0,
        COL_ALERT_INTERVAL: 60,
        COL_ALERT_TIMEOUT: '',
        COL_SUGGESTION: '',
        COL_CONDITIONS_OPERATOR: 'or',
        COL_DETECT_WINDOW: 0,
        COL_GRANULARITY: '',
        COL_CONDITION_DESC: short_name + '通用告警规则',
        COL_AND_CONDITIONS: '[]',
        COL_BLOCK_TIME: '',
        COL_EFFECTIVE_TIME: effective_time,
        COL_EVENT_FILTER: event_filter,
        COL_MULTI_CONDITIONS: json.dumps(conditions, ensure_ascii=False),
        COL_TEMPLATES: '',
        COL_UPGRADE_CONDITION: upgrade_condition,
        COL_RECOVER_COUNT: 0,
        # 告警规则类别 / 监控目标策略 / 规则描述 (V-X)
        COL_ALERT_RULE_TYPE: '',
        COL_TARGET_STRATEGY: '{appSystemName}{serviceSetName}目标资源',
        COL_RULE_DESC: '',
        # 告警通知策略 (AX-BD)
        COL_NOTIFY_NAME: '{objectName}通用通知策略',
        COL_NOTIFY_TEMPLATE: notify_template,
        COL_NOTIFY_EVENT_FILTER: event_filter,
        COL_NOTIFY_EFFECTIVE_TIME: notify_effective_time,
        COL_NOTIFY_FILTER_TYPE: 'all',
        COL_NOTIFY_REPEAT: 0,
        COL_NOTIFY_DESC: '',
    }

    for col, val in data.items():
        ws.cell(row=row_num, column=col, value=val)


def generate(plugin_dir, output_path=None, conditions_json_path=None):
    """
    主生成函数。

    Args:
        plugin_dir: 监控插件目录路径
        output_path: 输出 Excel 文件路径 (默认在插件目录下生成)
        conditions_json_path: LLM 已填写阈值的 conditions JSON 路径 (可选)
    """
    plugin_config, origin_metrics, alias_metrics, model_name = find_plugin_config(plugin_dir)

    model_id = plugin_config.get('relateObjectId', '')
    plugin_name = plugin_config.get('name', '')

    if not model_id:
        print("错误: plugin.yaml 中缺少 relateObjectId 字段")
        sys.exit(1)

    template_name = "{}初始化模板".format(model_name or plugin_name)

    if not output_path:
        output_path = os.path.join(
            os.path.dirname(os.path.abspath(plugin_dir)),
            '{}_告警规则模板.xlsx'.format(plugin_name or model_name)
        )

    conditions_json_output = output_path.replace('.xlsx', '_conditions.json')

    if conditions_json_path:
        # ========== 填充模式：使用 LLM 已配置的条件 ==========
        conditions = import_conditions_json(conditions_json_path)
        print("已加载 LLM 配置的条件: {} ({} 个)".format(
            conditions_json_path, len(conditions)))
    else:
        # ========== 骨架模式：生成无阈值的条件骨架 ==========
        conditions = []
        for metric in origin_metrics:
            key = metric.get('key', '')
            data_type = metric.get('dataType', 'double')
            display_name = alias_metrics.get(key, '')

            if should_skip_metric(key):
                continue

            conditions.append(
                build_metric_condition_skeleton(key, data_type, display_name, metric.get('metricType', ''))
            )

        if not conditions:
            print("警告: 没有可用于生成告警规则的指标")
            sys.exit(1)

        # 导出条件骨架 JSON 供 LLM 填写
        export_conditions_json(conditions, conditions_json_output)

        print()
        print("===== 条件骨架已生成，需 LLM 填写阈值 =====")
        print("1. 请将 {} 交给 LLM".format(conditions_json_output))
        print("2. LLM 分析采集脚本后填写每个指标的 threshold/displayThreshold")
        print("3. 使用以下命令重新生成完整 Excel:")
        print("   python generate_alert_rule_template.py {} --conditions-json {}".format(
            plugin_dir, conditions_json_output))

    # 创建 Excel
    wb, ws = create_workbook()
    write_data_row(ws, 3, template_name, model_id, conditions)

    # 保存
    wb.save(output_path)

    # 输出结果摘要
    print()
    print("告警规则模板已生成: {}".format(output_path))
    print("模板名称: {}".format(template_name))
    print("资源模型: {}".format(model_id))
    print("包含指标: {} 个".format(len(conditions)))
    for c in conditions:
        name = c.get('metricDisplayName', c['metricName'])
        has_threshold = any(
            'threshold' in comp or 'displayThreshold' in comp
            for comp in c['comparators']
        )
        marker = "有阈值" if has_threshold else "待配置"
        print("  - {} ({}) [{}]".format(name, c['type'], marker))


def main():
    parser = argparse.ArgumentParser(
        description='生成告警规则模板 Excel (ALERT_RULE_TEMPLATE 模型)'
    )
    parser.add_argument('plugin_dir', help='监控插件目录路径')
    parser.add_argument('-o', '--output', help='输出 Excel 文件路径 (默认在插件目录下生成)')
    parser.add_argument(
        '--conditions-json',
        help='LLM 已填写阈值的 conditions JSON 路径 (不指定则生成骨架)'
    )
    args = parser.parse_args()

    generate(args.plugin_dir, args.output, args.conditions_json)


if __name__ == '__main__':
    main()
