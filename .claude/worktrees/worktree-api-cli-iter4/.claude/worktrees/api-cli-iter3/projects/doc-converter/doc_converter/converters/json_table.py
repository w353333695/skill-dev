"""JSON/YAML → Excel/CSV 转换器"""

import csv
import json
from pathlib import Path
from .base import BaseConverter, ConvertResult, register


def flatten_dict(d: dict, prefix: str = "") -> dict:
    """扁平化嵌套字典: {"a": {"b": 1}} → {"a.b": 1}"""
    items = {}
    for k, v in d.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            items.update(flatten_dict(v, key))
        else:
            items[key] = v
    return items


def load_as_records(input_path: Path) -> list[dict]:
    """加载 JSON/YAML 为记录列表"""
    text = input_path.read_text(encoding="utf-8")
    suffix = input_path.suffix.lower()

    if suffix in (".yaml", ".yml"):
        import yaml
        data = yaml.safe_load(text)
    else:
        data = json.loads(text)

    # 统一为列表
    if isinstance(data, dict):
        # 尝试找到列表字段
        for v in data.values():
            if isinstance(v, list) and v and isinstance(v[0], dict):
                data = v
                break
        else:
            data = [data]
    elif not isinstance(data, list):
        data = [{"value": data}]

    # 扁平化
    return [flatten_dict(item) if isinstance(item, dict) else {"value": item} for item in data]


@register
class JsonToExcel(BaseConverter):
    name = "json-excel"
    source_formats = ["json", "yaml", "yml"]
    target_formats = ["xlsx"]
    description = "JSON/YAML 转 Excel (自动扁平化)"
    dependencies = ["openpyxl"]

    def convert(self, input_path: Path, output_path: Path, **options) -> ConvertResult:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        records = load_as_records(input_path)
        if not records:
            return ConvertResult(False, message="数据为空")

        # 收集所有列
        headers = list(dict.fromkeys(k for r in records for k in r.keys()))

        wb = Workbook()
        ws = wb.active
        ws.title = options.get("sheet_name", input_path.stem)

        # 写表头
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E8E8E8", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # 写数据
        for row_idx, record in enumerate(records, 2):
            for col_idx, h in enumerate(headers, 1):
                val = record.get(h, "")
                if isinstance(val, (list, dict)):
                    val = json.dumps(val, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=val)

        # 自动列宽
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        wb.save(str(output_path))
        return ConvertResult(
            True, output_path=output_path,
            message=f"已转换 {len(records)} 条记录, {len(headers)} 列"
        )


@register
class JsonToCsv(BaseConverter):
    name = "json-csv"
    source_formats = ["json", "yaml", "yml"]
    target_formats = ["csv"]
    description = "JSON/YAML 转 CSV (自动扁平化)"
    dependencies = []

    def convert(self, input_path: Path, output_path: Path, **options) -> ConvertResult:
        records = load_as_records(input_path)
        if not records:
            return ConvertResult(False, message="数据为空")

        headers = list(dict.fromkeys(k for r in records for k in r.keys()))
        encoding = options.get("encoding", "utf-8")

        with open(output_path, "w", encoding=encoding, newline="") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            for record in records:
                row = {}
                for h in headers:
                    val = record.get(h, "")
                    if isinstance(val, (list, dict)):
                        val = json.dumps(val, ensure_ascii=False)
                    row[h] = val
                writer.writerow(row)

        return ConvertResult(
            True, output_path=output_path,
            message=f"已转换 {len(records)} 条记录, {len(headers)} 列"
        )
