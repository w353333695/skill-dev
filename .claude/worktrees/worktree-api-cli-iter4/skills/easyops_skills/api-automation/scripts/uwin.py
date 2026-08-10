#! /usr/bin/env python3

import os
import sys
import requests
import platform
import yaml
from pprint import pp
from typing import *
import hmac
import hashlib
import time
import json
import logging
import traceback
from functools import wraps
import re
from pathlib import Path

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def retry(times=3, delay=1, backoff=2):
    """
    重试装饰器

    Args:
        times: 最大重试次数
        delay: 初始延迟时间（秒）
        backoff: 延迟时间的倍数增长因子
    """
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            attempt = 0
            current_delay = delay

            while attempt < times:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    attempt += 1
                    if attempt >= times:
                        tb = traceback.format_exc()
                        raise Exception(f"重试{times}次后仍然失败,堆栈信息:\n{tb}")
                    else:
                        logger.warning(f"第{attempt}次尝试失败,{current_delay}秒后重试: {e}")
                        time.sleep(current_delay)
                    current_delay *= backoff
        return wrapper
    return decorator


class EnvManager:
    """环境配置管理器"""

    CONFIG_NAME = 'environments.yaml'

    def __init__(self):
        self.config_path = self._find_config_path()
        self.config = self._load_config()

    def _find_config_path(self) -> Path:
        """查找配置文件路径，按优先级：当前目录 > 脚本目录 > 用户目录"""
        candidates = [
            Path.cwd() / self.CONFIG_NAME,
            Path(__file__).parent / self.CONFIG_NAME,
            Path.home() / '.easyops' / self.CONFIG_NAME,
        ]
        for p in candidates:
            if p.exists():
                return p
        # 默认使用脚本目录
        return Path(__file__).parent / self.CONFIG_NAME

    def _load_config(self) -> dict:
        """加载配置文件"""
        if self.config_path.exists():
            with open(self.config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f) or {}
        return {'current': None, 'environments': {}}

    def _save_config(self):
        """保存配置文件"""
        self.config_path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.config_path, 'w', encoding='utf-8') as f:
            yaml.dump(self.config, f, allow_unicode=True, default_flow_style=False)

    def get_current_env(self) -> Optional[dict]:
        """获取当前环境配置"""
        current_id = self.config.get('current')
        if not current_id:
            return None
        return self.config.get('environments', {}).get(current_id)

    def get_current_env_id(self) -> Optional[str]:
        """获取当前环境ID"""
        return self.config.get('current')

    def list_envs(self):
        """列出所有环境"""
        envs = self.config.get('environments', {})
        current = self.config.get('current')
        if not envs:
            print("暂无配置环境，使用 --env-add <id> 添加")
            return
        print(f"配置文件: {self.config_path}\n")
        for env_id, env in envs.items():
            marker = " *" if env_id == current else "  "
            name = env.get('name', '')
            host = env.get('host', '')
            org = env.get('org', '')
            print(f"{marker} {env_id}: {name} ({host}, org={org})")

    def show_env(self, env_id: str = None):
        """显示环境详情"""
        if not env_id:
            env_id = self.config.get('current')
        if not env_id:
            print("未指定环境ID，且未设置当前环境")
            return
        env = self.config.get('environments', {}).get(env_id)
        if not env:
            print(f"环境 '{env_id}' 不存在")
            return
        print(f"环境ID: {env_id}")
        for k, v in env.items():
            if k in ('ak', 'sk') and v:
                v = v[:4] + '****'  # 隐藏敏感信息
            print(f"  {k}: {v}")

    def add_env(self, env_id: str):
        """交互式添加环境"""
        if env_id in self.config.get('environments', {}):
            print(f"环境 '{env_id}' 已存在，使用 --env-edit 修改")
            return
        print(f"添加环境: {env_id}")
        env = {
            'name': input("  名称: ").strip() or env_id,
            'host': input("  主机地址: ").strip(),
            'org': input("  组织ID: ").strip(),
            'user': input("  用户名 (默认 defaultUser): ").strip() or 'defaultUser',
            'ak': input("  Access Key (可选): ").strip(),
            'sk': input("  Secret Key (可选): ").strip(),
        }
        if 'environments' not in self.config:
            self.config['environments'] = {}
        self.config['environments'][env_id] = env
        if not self.config.get('current'):
            self.config['current'] = env_id
        self._save_config()
        print(f"环境 '{env_id}' 已添加")

    def edit_env(self, env_id: str):
        """交互式编辑环境"""
        env = self.config.get('environments', {}).get(env_id)
        if not env:
            print(f"环境 '{env_id}' 不存在")
            return
        print(f"编辑环境: {env_id} (直接回车保留原值)")
        env['name'] = input(f"  名称 [{env.get('name', '')}]: ").strip() or env.get('name', '')
        env['host'] = input(f"  主机地址 [{env.get('host', '')}]: ").strip() or env.get('host', '')
        env['org'] = input(f"  组织ID [{env.get('org', '')}]: ").strip() or env.get('org', '')
        env['user'] = input(f"  用户名 [{env.get('user', 'defaultUser')}]: ").strip() or env.get('user', 'defaultUser')
        env['ak'] = input(f"  Access Key (输入空格清除): ").strip() or env.get('ak', '')
        env['sk'] = input(f"  Secret Key (输入空格清除): ").strip() or env.get('sk', '')
        self._save_config()
        print(f"环境 '{env_id}' 已更新")

    def use_env(self, env_id: str):
        """切换当前环境"""
        if env_id not in self.config.get('environments', {}):
            print(f"环境 '{env_id}' 不存在")
            return
        self.config['current'] = env_id
        self._save_config()
        env = self.config['environments'][env_id]
        print(f"已切换到环境: {env_id} ({env.get('name', '')})")

    def del_env(self, env_id: str):
        """删除环境"""
        if env_id not in self.config.get('environments', {}):
            print(f"环境 '{env_id}' 不存在")
            return
        confirm = input(f"确认删除环境 '{env_id}'? (y/N): ").strip().lower()
        if confirm != 'y':
            print("已取消")
            return
        del self.config['environments'][env_id]
        if self.config.get('current') == env_id:
            self.config['current'] = None
        self._save_config()
        print(f"环境 '{env_id}' 已删除")


class LocalDataStore:
    """本地 JSON 数据离线查询引擎，兼容 CMDB 查询语法"""

    def __init__(self, data_dir: str):
        """
        初始化离线数据存储

        :param data_dir: 包含 {model_id}.json 文件的目录路径
        """
        self.data_dir = Path(data_dir)
        self._cache = {}

    def has_model(self, model_id: str) -> bool:
        """检查本地是否有指定模型的数据文件"""
        return self._get_file_path(model_id).exists()

    def search(self, model_id: str, query: dict) -> list:
        """
        在本地数据中执行查询

        支持的查询语法：
        - 精确匹配: {'field': 'value'}
        - 模糊匹配: {'field': {'$like': '%pattern%'}}
        - 逻辑与: {'$and': [cond1, cond2, ...]}
        - 存在判断: {'field': {'$exists': True/False}}
        - 嵌套字段: 'a.b.c' 访问 obj['a']['b']['c']

        :param model_id: 模型 ID（对应文件名）
        :param query: CMDB 风格查询条件
        :return: 匹配的实例列表
        """
        data = self._load_data(model_id)
        if not data:
            return []
        if not query:
            return data
        return [item for item in data if self._match(item, query)]

    def _get_file_path(self, model_id: str) -> Path:
        """获取模型对应的 JSON 文件路径"""
        return self.data_dir / f"{model_id}.json"

    def _load_data(self, model_id: str) -> list:
        """加载并缓存 JSON 数据"""
        if model_id in self._cache:
            return self._cache[model_id]
        file_path = self._get_file_path(model_id)
        if not file_path.exists():
            return []
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, list):
            data = [data]
        self._cache[model_id] = data
        logger.info(f"[离线模式] 已加载 {file_path.name}，共 {len(data)} 条记录")
        return data

    def _match(self, item: dict, query: dict) -> bool:
        """递归匹配单条记录"""
        for key, condition in query.items():
            if key == '$and':
                if not all(self._match(item, sub) for sub in condition):
                    return False
            elif key == '$or':
                if not any(self._match(item, sub) for sub in condition):
                    return False
            else:
                value = self._get_nested(item, key)
                if not self._match_condition(value, condition):
                    return False
        return True

    def _get_nested(self, obj: dict, key: str):
        """获取嵌套字段值，支持 'a.b.c' 点号路径"""
        parts = key.split('.')
        current = obj
        for part in parts:
            if isinstance(current, dict):
                current = current.get(part)
            else:
                return None
            if current is None:
                return None
        return current

    def _match_condition(self, value, condition) -> bool:
        """匹配单个字段条件"""
        if isinstance(condition, dict):
            for op, operand in condition.items():
                if op == '$like':
                    if value is None:
                        return False
                    return self._like_match(str(value), operand)
                elif op == '$exists':
                    return (value is not None) == operand
                elif op == '$in':
                    return value in operand
                elif op == '$ne':
                    return value != operand
                elif op == '$gt':
                    return value is not None and value > operand
                elif op == '$lt':
                    return value is not None and value < operand
                elif op == '$gte':
                    return value is not None and value >= operand
                elif op == '$lte':
                    return value is not None and value <= operand
            return False
        else:
            # 精确匹配
            return value == condition

    @staticmethod
    def _like_match(value: str, pattern: str) -> bool:
        """SQL LIKE 风格匹配，% 作为通配符"""
        # 转换为正则：% -> .*, _ -> .
        regex_pattern = '^'
        for ch in pattern:
            if ch == '%':
                regex_pattern += '.*'
            elif ch == '_':
                regex_pattern += '.'
            else:
                regex_pattern += re.escape(ch)
        regex_pattern += '$'
        return bool(re.match(regex_pattern, value, re.IGNORECASE))


def print_env_help(current_env_id: str = None, host: str = None):
    """打印环境相关的帮助信息"""
    print("\n" + "=" * 50)
    if current_env_id:
        print(f"当前环境: {current_env_id} ({host})")
    print("\n可能的解决方案:")
    print("  1. 检查网络连接: ping <host>")
    print("  2. 查看可用环境: python uwin.py --env-list")
    print("  3. 切换环境: python uwin.py --env-use <环境ID>")
    print("  4. 添加新环境: python uwin.py --env-add <环境ID>")
    print("=" * 50)


class EasyopsAPI:
    def __init__(self, host: Optional[str]=None, org: Optional[str]=None, **kwargs):
        """
        初始化EasyopsAPI客户端

        :param host: EasyOps服务器主机地址,如果不提供则从环境配置或agent配置中读取
        :param org: 组织ID,如果不提供则从环境配置或agent配置中读取
        :param user: 用户名,默认为'defaultUser'
        :param ak: 访问密钥(Access Key),用于OpenAPI认证
        :param sk: 密钥(Secret Key),用于OpenAPI签名
        :param dry_run: 是否只打印curl命令而不发送请求,默认False
        :param local_data: 离线数据目录路径,包含 {model_id}.json 文件;
                           指定后优先从本地查询,无本地数据时回退在线
        """
        self._env_id = None
        self.dry_run = kwargs.get('dry_run', False)
        self._local_store = None
        local_data = kwargs.get('local_data')
        if local_data:
            self._local_store = LocalDataStore(local_data)
        user = kwargs.get('user')
        ak = kwargs.get('ak')
        sk = kwargs.get('sk')

        if not host:
            # 优先从环境配置读取
            env_mgr = EnvManager()
            env = env_mgr.get_current_env()
            if env:
                self._env_id = env_mgr.get_current_env_id()
                host = env.get('host')
                org = env.get('org')
                user = user or env.get('user', 'defaultUser')
                ak = ak or env.get('ak')
                sk = sk or env.get('sk')
            elif self._local_store:
                # 纯离线模式：无需网络配置
                logger.info("[离线模式] 未配置在线环境，仅使用本地数据")
            else:
                # 回退到 agent 配置
                host, org = self.__get_host_and_org()

        self.host = host
        self.org = org
        self.headers = {
            "user": user or "defaultUser",
            "org": self.org,
            'Content-Type': 'application/json'
        }
        if ak and sk:
            self.is_openapi = True
            self.ak = ak
            self.sk = sk
            self.headers['Host'] = 'openapi.easyops-only.com'
            openapi_yaml_path = os.path.join(os.path.dirname(__file__), 'openapi.yaml')
            openapi_conf = yaml.safe_load(open(openapi_yaml_path))
            self.port2app_name = {i['port']: i['app_name'] for i in openapi_conf['app_route']}
        else:
            self.is_openapi = False

    def __get_host_and_org(self) -> tuple[str, int]:
        """
        从agent配置文件中获取host和org信息
        
        :return: 包含host(IP地址)和org(组织ID)的元组
        :rtype: tuple[str, int]
        """
        if platform.system().lower() == "windows":  # 修正low()为lower()
            conf_path = "C:\\easyOps\\agent\\conf\\conf.yaml"
        else:
            conf_path = "/usr/local/easyops/agent/conf/conf.yaml"
        with open(conf_path,'r') as f:
            dic = yaml.load(f, Loader=yaml.FullLoader)
        org = dic['base']['client_id']
        host = dic['command']['server_groups'][0]['hosts'][0]['ip']
        return host, str(org)
    
    @retry()
    def _request(self, method: str, path: str, port: int, timeout: int = 3, **kwargs) -> requests.Response:
        """
        发送HTTP请求到EasyOps服务器

        :param method: HTTP请求方法,如'GET', 'POST', 'PUT', 'DELETE'
        :param path: 请求路径,如'/v3/object/instance/_search'
        :param port: 服务端口号
        :param timeout: 请求超时时间,默认3秒
        :param data: 请求体数据,会被自动转换为JSON字符串
        :param params: URL查询参数
        :param files: 文件上传参数,格式为 {'field_name': (filename, file_obj, content_type)}
        :param form_data: 表单数据,用于文件上传时的附加字段（不会被JSON序列化）
        :return: HTTP响应对象
        :rtype: requests.Response
        """
        from urllib.parse import urlencode

        # 文件上传模式：使用 files 和 form_data，不序列化 data
        if kwargs.get('files'):
            form_data = kwargs.pop('form_data', None)
            if form_data:
                kwargs['data'] = form_data
        elif kwargs.get('data'):
            kwargs['data'] = json.dumps(kwargs['data'])

        headers = self.headers.copy()
        if self.is_openapi:
            app_name = self.port2app_name.get(port)
            uri = f"/{app_name}/{path.lstrip('/')}"
            url = f"http://{self.host}{uri}"
            params = self.__signature(method, uri, kwargs.get('params', {}), kwargs.get('data', '{}'))
            url = url + "?" + urlencode(params)
            if kwargs.get('params'):
                del kwargs['params']
            # OpenAPI 模式下 GET/DELETE 不发 Content-Type
            if method.upper() in ('GET', 'DELETE'):
                headers.pop('Content-Type', None)
            headers.pop('org', None)
        else:
            url = f"http://{self.host}:{port}/{path.lstrip('/')}"
        logger.debug(f"Method: {method}\nUrl: {url}\nHeaders:{headers}\nKwargs:{kwargs}")

        # dry_run 模式：打印 curl 命令，不发送请求
        if self.dry_run:
            curl_cmd = self._build_curl_command(method, url, kwargs)
            print(curl_cmd)
            class DryRunResponse:
                status_code = 200
                text = '{}'
                headers = {'Content-Type': 'application/json'}
                content = b''
                def json(self): return {}
            return DryRunResponse()

        # 文件上传时使用不含 Content-Type 的 headers（让 requests 自动设置 multipart boundary）
        if kwargs.get('files'):
            upload_headers = {k: v for k, v in headers.items() if k.lower() != 'content-type'}
            response = requests.request(method, url, headers=upload_headers, timeout=timeout, verify=False, **kwargs)
        else:
            response = requests.request(method, url, headers=headers, timeout=timeout, verify=False, **kwargs)
        logger.debug(f"StatusCode:{response.status_code}\nResponse: {response.text}")

        # 处理非200状态码
        if response.status_code != 200:
            logger.error(
                f"Request failed with status code {response.status_code} ({response.reason}). "
                f"Response body: {response.text}"
            )
            response.raise_for_status()
        return response

    def _build_curl_command(self, method: str, url: str, kwargs: dict) -> str:
        """
        构建 curl 命令字符串

        :param method: HTTP请求方法
        :param url: 请求URL
        :param kwargs: 请求参数(data, params等)
        :return: curl 命令字符串
        """
        parts = ['curl', '-X', method.upper()]

        # 添加 headers
        for k, v in self.headers.items():
            parts.append(f"-H '{k}: {v}'")

        # 添加 query params (如果未拼接到url)
        if kwargs.get('params') and '?' not in url:
            from urllib.parse import urlencode
            url = url + '?' + urlencode(kwargs['params'])

        # 添加 data
        if kwargs.get('data'):
            data = kwargs['data']
            # 转义单引号
            data_escaped = data.replace("'", "'\\''")
            parts.append(f"-d '{data_escaped}'")

        parts.append(f"'{url}'")
        return ' \\\n  '.join(parts)

    def __signature(self, method: str,url: str,params={},data='{}') -> dict:
        """
        为OpenAPI请求生成签名

        :param method: HTTP请求方法
        :param url: 完整的请求URL
        :param params: URL查询参数
        :param data: 请求体数据
        :return: 包含签名和其他认证参数的字典
        :rtype: dict
        """
        request_time = str(int(time.time()))
        method = method.upper()
        content_type = 'application/json' if method in ['POST', 'PUT'] else ''
        url_param = ''.join(['%s%s' % (k, params[k]) for k in sorted(params.keys())])
        content_md5 = ''
        if method in ['POST','PUT'] and data:
            md5 = hashlib.md5()
            if isinstance(data, str):
                data_bytes = data.encode('utf-8')
            else:
                data_bytes = data
            md5.update(data_bytes)
            content_md5 = md5.hexdigest()
        string_to_signaure = "\n".join([
            method,
            url,
            url_param,
            content_type,
            content_md5,
            request_time,
            self.ak
        ]).encode()
        s = self.sk.encode()
        params.update({
            'accesskey': self.ak,
            'signature': hmac.new(s, string_to_signaure, hashlib.sha1).hexdigest(),
            'expires': request_time
        })
        return params
    
    def search_instance(self, model_id:str,**kwargs) -> list:
        """
        搜索指定对象类型的实例

        EasyOps API: PostSearchV3WithAdmin
        服务: logic.cmdb.service

        支持离线模式：当 local_data 目录中存在 {model_id}.json 时，
        直接从本地文件查询，无需网络连接。

        :param model_id: 对象ID,指定要搜索的模型类型
        :param fields: 要返回的字段列表,可填写属性、关系ID，支持多层嵌套，如：app表示关联应用，app.system表示关联应用.关联的系统，app.system.name表示关联应用.关联的系统.名称,默认为['*']返回所有属性，默认不返回关系
        :param query: 查询条件,如{'name': {'$like': '%value%'}}
        :param page_size: 每页大小,默认1000
        :return: 包含所有匹配实例的列表
        :rtype: list，格式：list[dict{'id': str, 'relation_id':list[dict]}]
        """
        query = kwargs.get('query', {})

        # 离线模式：优先从本地数据查询
        if self._local_store and self._local_store.has_model(model_id):
            logger.info(f"[离线模式] 从本地数据查询 {model_id}")
            insts = self._local_store.search(model_id, query)
            if insts:
                logger.info(f"[离线模式] 找到 {len(insts)} 条 {model_id} 记录")
            else:
                logger.warning(f"[离线模式] 未找到匹配的 {model_id} 记录, query: {query}")
            return insts

        # 在线模式
        port = 8079
        url = f"v3/object/{model_id}/instance/_search"
        data = {
            'fields': kwargs.get('fields',['*']),
            'query': query,
            'page':1,
            'page_size': kwargs.get('page_size',1000)
        }
        insts = []
        for page in range(1,10000):
            data['page'] = page
            response = self._request('POST', url, port=port, data=data)
            tmp = response.json()['data']['list']
            insts.extend(tmp)
            if len(tmp) < data['page_size']:
                break
        if insts:
            logger.info(f"Found {len(insts)} instances of {model_id}.")
        else:
            logger.warning(f"No instances found of {model_id} with the given data: {data}.")
        return insts

    def get_api_desc(self,project_name: str = 'Easyops', namespaceId: Optional[str] = None,**kwargs) -> dict:
        """
        根据api名称,获取api描述

        :param project_name: 所属api项目的名称,默认Easyops
        :param namespaceId: 命名空间,默认为None,如：easyops.api.artifact.
        :param api_name: api名称,默认为None,如：PostSearchV3WithAdmin
        :param description: api描述包含关键字,默认为None
        :param detail: api详情包含关键字,默认None
        :param serviceName: 服务名包含关键字,默认None
        :return: 模型配置信息,示例：{
            "description": "搜索实例V3", # api描述
            "serviceName": "logic.cmdb.service", # 服务名称,可使用get_service_port获取服务端口
            "endpoint": {
                "ext_fields": [], # 模型字段
                "method": "POST", # 请求方法
                "uri": "/v3/object/:objectId/instance/_search" # 请求路径
            },
            "name": "PostSearchV3WithAdmin", # api名称
            "request": {
                "default": {},
                "description": "",
                "fields": [
                    {
                        "deprecated": false,
                        "description": "页大小",
                        "enum": [],
                        "fields": [],
                        "name": "page_size",
                        "ref": "",
                        "type": "page_size",
                        "validate": null
                    },
                    {
                        "deprecated": false,
                        "description": "按字段排序, 留空默认按照实例ID降序排序(1表示升序, -1表示降序, 2表示自然升序, -2表示自然降序) e.g.: [{ key: instanceId, order: 1}]",
                        "enum": [],
                        "fields": [
                            {
                                "deprecated": false,
                                "description": "属性id",
                                "enum": [],
                                "fields": [],
                                "name": "key",
                                "ref": "",
                                "type": "string",
                                "validate": null
                            },
                            {
                                "deprecated": false,
                                "description": "1表示升序, -1表示降序, 2表示自然升序, -2表示自然降序",
                                "enum": [
                                    -1,
                                    1,
                                    -2,
                                    2
                                ],
                                "fields": [],
                                "name": "order",
                                "ref": "",
                                "type": "int",
                                "validate": null
                            }
                        ],
                        "name": "sort",
                        "ref": "",
                        "type": "object[]",
                        "validate": null
                    }
                ],
                "required": [],
                "type": "object"
            },
            "response": {
                "default": {},
                "description": "",
                "fields": [
                    {
                        "deprecated": false,
                        "description": "instance list",
                        "enum": [],
                        "fields": [],
                        "name": "list",
                        "ref": "",
                        "type": "map[]",
                        "validate": null
                    },
                    {
                        "deprecated": false,
                        "description": "实例总数",
                        "enum": [],
                        "fields": [],
                        "name": "total",
                        "ref": "",
                        "type": "int",
                        "validate": null
                    }
                ],
                "required": [],
                "type": "object",
                "wrapper": true
            }
        }
        :rtype: list[dict]
        """
        is_offline = (self._local_store and
                      self._local_store.has_model('FLOW_BUILDER_API_CONTRACT@EASYOPS'))

        query = {}
        # namespace.apiProject.name 是 CMDB 关系字段，离线模式无此字段，跳过
        if not is_offline:
            query['namespace.apiProject.name'] = project_name

        if namespaceId:
            query['namespaceId'] = {'$like':f'%{namespaceId}%'}

        if kwargs.get('fullContractName'):
            query['fullContractName'] = kwargs.get('fullContractName')
            del kwargs['fullContractName']
        if kwargs.get('detail'):
            detail_val = kwargs.get('detail')
            query['detail'] = {'$like': f'%{detail_val}%'}
            del kwargs['detail']
        if kwargs.get('description'):
            desc_val = kwargs.get('description')
            query['description'] = {'$like': f'%{desc_val}%'}
            del kwargs['description']
        # api_name 映射为本地数据的 name 字段
        if kwargs.get('api_name'):
            query['name'] = kwargs.pop('api_name')
        query.update(kwargs)
        apis = self.search_instance('FLOW_BUILDER_API_CONTRACT@EASYOPS',query=query)
        if apis:
            return apis

    def get_api_desc_by_url(self, url: str, **kwargs) -> dict:
        """
        通过浏览器中的 gateway URL 获取 API 描述

        支持 URL 中包含变量的情况，如模型ID（HOST@EASYOPS）、实例ID（64d481bf0a74d）等，
        会自动将变量部分替换为通配符进行模糊匹配。

        :param url: 浏览器中复制的完整 URL，如:
            http://172.30.0.90/next/api/gateway/pipeline.docker_worker.List/api/pipeline/v1/docker_workers?page=1&page_size=1000
            http://172.30.0.90/next/api/gateway/cmdb.instance.PostSearchV3/v3/object/FLOW_BUILDER_API_CONTRACT@EASYOPS/instance/_search
            https://11.66.19.194/next/api/gateway/cmdb.instance.GetDetail/object/HOST@EASYOPS/instance/64d481bf0a74d?fields=_object_id
        :return: API 描述信息（同 get_api_desc 返回值）
        """
        result = re.findall(r'([a-z]+\..*?)(/.*?)(?:\?|$)', url)
        if not result:
            logger.error('Unsupported url: %s', url)
            return None
        query = {'$and':[]}
        tag,uri = result[0]
        tag_split = tag.split('.')
        if tag_split[-1][0].islower():
            query['$and'].append({'serviceName':{'$like':f'%{tag}%'}})
        else:
            fullContractName = '.'.join(tag_split[:-1])+'@'+tag_split[-1]
            query['$and'].append({'fullContractName':{'$like':f'%{fullContractName}%'}})
        # 识别 URI 中的变量部分（模型ID、实例ID、纯数字ID），按变量位置切分为静态片段
        # 每个片段作为独立的 $like 条件，用 % 前后缀拼接
        var_pattern = re.compile(
            r'/[A-Z][A-Z0-9_]*@[A-Z][A-Z0-9_]*'  # 模型ID: HOST@EASYOPS
            r'|/[a-z0-9]{13}(?=/|$)'               # 实例ID: 64d481bf0a74d
            r'|/\d+(?=/|$)'                         # 纯数字ID
        )
        if not var_pattern.search(uri):
            # 无变量，精确匹配
            query['$and'].append({'endpoint.uri':uri})
        else:
            # 有变量，按变量切分后每个静态片段加 % 通配
            parts = [p for p in var_pattern.split(uri) if p]
            for i, part in enumerate(parts):
                prefix = '%' if i > 0 else ''
                suffix = '%' if i < len(parts) - 1 else ''
                # 末尾片段：如果原始 URI 以变量结尾，也需要加 % 后缀
                if i == len(parts) - 1 and not uri.endswith(part):
                    suffix = '%'
                query['$and'].append({'endpoint.uri':{'$like':f'{prefix}{part}{suffix}'}})
        apis = self.search_instance('FLOW_BUILDER_API_CONTRACT@EASYOPS',query=query)
        return apis

    def get_service_port(self,serviceName: str) -> int:
        """
        根据服务名称,获取服务端口

        EasyOps API: GetRoutingTableApi
        服务: ens.routing.service

        :param serviceName: 服务名称,如：logic.cmdb.service
        :return: 服务端口
        :rtype: str
        """
        self.headers['org'] = '2'
        res = self.search_instance('ENS_ROUTING_TABLE',query={"contract":serviceName})
        self.headers['org'] = self.org
        return res[0]['port']
    
    def get_model_desc(self,modle_id: str) -> dict:
        '''
        获取模型配置信息

        EasyOps API: GetObjectApi
        服务: logic.cmdb.service

        :param modle_id: 模型ID
        :return: 模型详情,示例：{
            "objectId": "TESTWWH",
            "name": "测试模型",
            "icon": "",
            "category": "部署",
            "memo": "",
            "protected": false,
            "system": "",
            "notifyDenied": false,
            "view": {
                "attr_authorizers": null,
                "attr_category_order": [
                    "基本信息"
                ],
                "attr_order": [],
                "hide_columns": null,
                "icon": {
                    "category": "second-menu",
                    "icon": "placeholder-second-menu",
                    "lib": "easyops"
                },
                "image": "",
                "inherit_attr_category_map": null,
                "relation_default_attr": null,
                "relation_group_order": null,
                "relation_order": null,
                "show_key": [
                    "instanceId"
                ],
                "trans_hier_relation_list": null,
                "visible": true
            },
            "attrList": [
                {
                    "id": "ip",
                    "name": "IP示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "ip",
                        "regex": "((^\\s*((([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5])\\.){3}([0-9]|[1-9][0-9]|1[0-9]{2}|2[0-4][0-9]|25[0-5]))\\s*$)|(^\\s*((([0-9A-Fa-f]{1,4}:){7}([0-9A-Fa-f]{1,4}|:))|(([0-9A-Fa-f]{1,4}:){6}(:[0-9A-Fa-f]{1,4}|((25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)(\\.(25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)){3})|:))|(([0-9A-Fa-f]{1,4}:){5}(((:[0-9A-Fa-f]{1,4}){1,2})|:((25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)(\\.(25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)){3})|:))|(([0-9A-Fa-f]{1,4}:){4}(((:[0-9A-Fa-f]{1,4}){1,3})|((:[0-9A-Fa-f]{1,4})?:((25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)(\\.(25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)){3}))|:))|(([0-9A-Fa-f]{1,4}:){3}(((:[0-9A-Fa-f]{1,4}){1,4})|((:[0-9A-Fa-f]{1,4}){0,2}:((25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)(\\.(25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)){3}))|:))|(([0-9A-Fa-f]{1,4}:){2}(((:[0-9A-Fa-f]{1,4}){1,5})|((:[0-9A-Fa-f]{1,4}){0,3}:((25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)(\\.(25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)){3}))|:))|(([0-9A-Fa-f]{1,4}:){1}(((:[0-9A-Fa-f]{1,4}){1,6})|((:[0-9A-Fa-f]{1,4}){0,4}:((25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)(\\.(25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)){3}))|:))|(:(((:[0-9A-Fa-f]{1,4}){1,7})|((:[0-9A-Fa-f]{1,4}){0,5}:((25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)(\\.(25[0-5]|2[0-4]\\d|1\\d\\d|[1-9]?\\d)){3}))|:)))(%.+)?\\s*$))",
                        "default_type": "",
                        "default": null,
                        "struct_define": [],
                        "mode": "",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "str",
                    "name": "字符示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "str",
                        "regex": null,
                        "default_type": "value",
                        "default": null,
                        "struct_define": [],
                        "mode": "default",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "bool",
                    "name": "布尔示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "bool",
                        "regex": null,
                        "default_type": "",
                        "default": null,
                        "struct_define": [],
                        "mode": "",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "date",
                    "name": "日期示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "date",
                        "regex": null,
                        "default_type": "",
                        "default": null,
                        "struct_define": [],
                        "mode": "",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "enum",
                    "name": "单选枚举",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "enum",
                        "regex": [
                            "枚举1",
                            "枚举2"
                        ],
                        "default_type": "",
                        "default": null,
                        "struct_define": [],
                        "mode": "default",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "float",
                    "name": "浮点型示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "float",
                        "regex": null,
                        "default_type": "",
                        "default": null,
                        "struct_define": [],
                        "mode": "",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "int",
                    "name": "整型示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "int",
                        "regex": null,
                        "default_type": "value",
                        "default": null,
                        "struct_define": [],
                        "mode": "",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "time",
                    "name": "时间示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "datetime",
                        "regex": null,
                        "default_type": "",
                        "default": null,
                        "struct_define": [],
                        "mode": "",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "enums",
                    "name": "多枚举实例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "enums",
                        "regex": [
                            "枚举1",
                            "枚举2",
                            "枚举3"
                        ],
                        "default_type": "",
                        "default": null,
                        "struct_define": [],
                        "mode": "default",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "file",
                    "name": "附件示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "structs",
                        "regex": null,
                        "default_type": "",
                        "default": null,
                        "struct_define": [
                            {
                                "id": "name",
                                "name": "name",
                                "type": "str",
                                "regex": "",
                                "protected": false,
                                "mode": "default"
                            },
                            {
                                "id": "type",
                                "name": "type",
                                "type": "str",
                                "regex": "",
                                "protected": false,
                                "mode": "default"
                            },
                            {
                                "id": "url",
                                "name": "url",
                                "type": "str",
                                "regex": "",
                                "protected": false,
                                "mode": "default"
                            },
                            {
                                "id": "size",
                                "name": "size",
                                "type": "float",
                                "regex": null,
                                "protected": false,
                                "mode": ""
                            }
                        ],
                        "mode": "attachment",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "list",
                    "name": "数组示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "arr",
                        "regex": null,
                        "default_type": "",
                        "default": null,
                        "struct_define": [],
                        "mode": "default",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "stract",
                    "name": "结构体示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "struct",
                        "regex": null,
                        "default_type": "",
                        "default": null,
                        "struct_define": [
                            {
                                "id": "attr1",
                                "name": "属性1",
                                "type": "str",
                                "regex": "",
                                "protected": false,
                                "mode": "default"
                            },
                            {
                                "id": "attr2",
                                "name": "属性2",
                                "type": "int",
                                "regex": "",
                                "protected": false,
                                "mode": ""
                            }
                        ],
                        "mode": "",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                },
                {
                    "id": "stracts",
                    "name": "结构体数组示例",
                    "protected": false,
                    "custom": "true",
                    "unique": "false",
                    "readonly": "false",
                    "required": "false",
                    "tag": [
                        "基本信息"
                    ],
                    "description": "",
                    "tips": "",
                    "value": {
                        "type": "structs",
                        "regex": null,
                        "default_type": "",
                        "default": null,
                        "struct_define": [
                            {
                                "id": "attr1",
                                "name": "属性1",
                                "type": "str",
                                "regex": "",
                                "protected": false,
                                "mode": "password"
                            },
                            {
                                "id": "attr2",
                                "name": "属性2",
                                "type": "date",
                                "regex": null,
                                "protected": false,
                                "mode": ""
                            }
                        ],
                        "mode": "",
                        "prefix": "",
                        "start_value": 0,
                        "series_number_length": 0
                    },
                    "wordIndexDenied": false,
                    "isInherit": false,
                    "notifyDenied": false,
                    "inheritObjectId": "",
                    "isMetadata": false
                }
            ],
            "relation_groups": [],
            "relation_list": [
                {
                    "relation_id": "TESTWWH_HOST_TEST_HOST",
                    "name": "",
                    "protected": false,
                    "notifyDenied": false,
                    "isInherit": false,
                    "left_object_id": "TESTWWH",
                    "leftInheritObjectId": "",
                    "left_id": "HOST",
                    "left_description": "关联测试模型实例",
                    "left_remark": "",
                    "left_name": "关联主机",
                    "left_min": 0,
                    "left_max": -1,
                    "left_groups": [],
                    "left_tags": [],
                    "left_required": false,
                    "right_object_id": "HOST",
                    "rightInheritObjectId": "",
                    "right_id": "TEST",
                    "right_description": "关联主机",
                    "right_remark": "",
                    "right_name": "关联测试模型实例",
                    "right_min": 0,
                    "right_max": -1,
                    "right_groups": [],
                    "right_tags": [],
                    "right_required": false,
                    "_version": 1,
                    "attrList": [],
                    "indexList": []
                }
            ],
            "indexList": [],
            "updateAuthorizers": [],
            "deleteAuthorizers": [],
            "readAuthorizers": [],
            "wordIndexDenied": false,
            "_version": 56,
            "creator": "easyops",
            "modifier": "easyops",
            "isAbstract": false,
            "parentObjectId": "",
            "permissionDenied": false,
            "parentObjectIds": []
        }
        :rtype: dict
        '''
        return self._request('get', f'object/{modle_id}', port=8079).json()['data']
    
    @staticmethod
    def _batch_process(data_list:list, batch_size:int):
        """
        按批次处理列表数据
        
        :param data_list: 待处理的列表
        :param batch_size: 每批次的大小
        :yield: 当前批次的数据
        """
        for i in range(0, len(data_list), batch_size):
            yield data_list[i:i + batch_size]
    
    def get_import_instance_excel(self, type: str = None, save_path: str = None) -> Union[str, list]:
        """
        获取数据录入模板文件

        EasyOps API: 资源管理数据录入模板导出
        服务: logic.resource_manage

        :param type: 模板类型，如 system。默认 None 时获取所有非隐藏模型并逐个下载模板
        :param save_path: 保存路径，默认为当前目录下的 data_entry_template.xlsx
        :return: 保存的文件路径，或文件路径列表（type=None时）
        """
        port = 8073
        path = 'api/v1/resource_manage/data_entry/templates/export'

        if type is None:
            # 获取所有非隐藏模型
            objects = self.list_object_basic(visible='visible')
            # 构建 objectId -> name/category 映射
            obj_info = {}
            for obj in objects:
                obj_info[obj['objectId']] = {
                    'name': obj.get('name', obj['objectId']),
                    'category': obj.get('category', ''),
                }
            object_ids = [obj['objectId'] for obj in objects]
            object_ids.insert(0, 'system')
            obj_info['system'] = {'name': '系统', 'category': ''}
            logger.info(f'Found {len(object_ids)} visible models')

            # 逐个下载模板，按 category 分目录存储
            saved_files = []
            save_dir = save_path or 'templates'

            for obj_id in object_ids:
                try:
                    response = self._request(method='get', path=path, port=port, params={'type': obj_id})
                    info = obj_info.get(obj_id, {})
                    model_name = info.get('name', obj_id)
                    category = info.get('category', '')
                    # category 如 '平台资源.数据库' -> '平台资源/数据库'
                    sub_dir = category.replace('.', os.sep) if category else ''
                    target_dir = os.path.join(save_dir, sub_dir) if sub_dir else save_dir
                    os.makedirs(target_dir, exist_ok=True)
                    file_path = os.path.join(target_dir, f'{model_name}_template.xlsx')
                    with open(file_path, 'wb') as f:
                        f.write(response.content)
                    saved_files.append(file_path)
                    logger.info(f'Saved template: {file_path}')
                except Exception as e:
                    model_name = obj_info.get(obj_id, {}).get('name', obj_id)
                    logger.warning(f'Failed to download template for {model_name}({obj_id}): {e}')

            logger.info(f'Downloaded {len(saved_files)} templates to {save_dir}')
            return saved_files

        # 指定 type 时直接下载
        response = self._request(method='get', path=path, port=port, params={'type': type})
        if save_path is None:
            save_path = f'data_entry_template_{type}.xlsx'
        with open(save_path, 'wb') as f:
            f.write(response.content)
        logger.info(f'Excel template saved to: {save_path}')
        return save_path

    def import_instance(self, obj_id:str, data_list:list, key:Union[str,list], batch_size:int=1000):
        """
        批量导入实例数据

        EasyOps API: PostImportInstanceApi
        服务: logic.cmdb.service

        :param obj_id: 实例所属模型ID
        :param data_list: 实例数据列表
        :param key: 实例数据列表中的键名
        :param batch_size: 每批次处理的数据量
        """
        port = 8079
        path = 'object/{}/instance/_import'.format(obj_id)
        insert_count = update_count = failed_count = 0
        for batch in self._batch_process(data_list, batch_size):
            data = {
                'keys': [key] if isinstance(key, str) else key,
                'datas': batch,
                'importMetadata':True
            }
            res = self._request(method='post', path=path, port=port, data=data).json()
            insert_count += res['data']['insert_count']
            update_count += res['data']['update_count']
            failed_count += res['data']['failed_count']
        logger.info(
            f'Imported {insert_count} instances, updated {update_count} instances, failed to import {failed_count} instances.'
        )

    def import_instance_excel(self, object_id: str, file_path: str, keys: List[str] = None, relation_operation: str = None, mark_failures: bool = True) -> dict:
        """
        使用Excel文件导入实例

        EasyOps API: ImportInstanceWithExcel
        服务: logic.cmdb.service

        :param object_id: 实例所属的模型ID
        :param file_path: Excel文件路径
        :param keys: 导入实例的字段,作为更新/创建依据（可选，不传则自动使用唯一且必填的属性）
        :param relation_operation: 关系操作方式，可选值: 'set', 'append'（可选）
        :param mark_failures: 是否在原Excel中标记失败行（默认True）
        :return: 导入结果，包含 insert_count, update_count, failed_count, data, marked_file(如有失败)
        :rtype: dict
        """
        # 如果未指定 keys，自动获取唯一且必填的属性
        if not keys:
            keys = self.get_unique_required_attrs(object_id)
            if not keys:
                raise ValueError(f"模型 {object_id} 没有唯一且必填的属性，无法进行导入。请手动指定 keys 参数或为模型添加唯一必填属性。")
            logger.info(f"自动使用唯一必填属性作为导入键: {keys}")

        port = 8079
        path = f'import/object/{object_id}/instance/excel'

        files = {'attachment': (os.path.basename(file_path), open(file_path, 'rb'), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')}
        form_data = {}
        # keys 需要用数组格式传递: keys[0], keys[1], ...
        if isinstance(keys, list):
            for i, key in enumerate(keys):
                form_data[f'keys[{i}]'] = key
        elif keys:
            form_data['keys[0]'] = keys
        if relation_operation:
            form_data['relation_operation'] = relation_operation

        try:
            response = self._request('POST', path, port=port, timeout=120, files=files, form_data=form_data if form_data else None)
            result = response.json()
            data = result.get('data', result)
            logger.info(f"Excel导入完成 [{object_id}]: 插入 {data.get('insert_count', 0)}, 更新 {data.get('update_count', 0)}, 失败 {data.get('failed_count', 0)}")
            if data.get('failed_count', 0) > 0:
                logger.warning(f"失败详情: {data.get('data', [])}")
                if mark_failures:
                    marked_file = self._mark_failed_rows_in_excel(file_path, data.get('data', []))
                    if marked_file:
                        data['marked_file'] = marked_file
                        logger.info(f"失败行已标记，文件保存至: {marked_file}")
            return data
        finally:
            files['attachment'][1].close()

    def import_instance_json(self, object_id: str, file_path: str, keys: List[str] = None, create_if_id_not_exists: bool = True) -> dict:
        """
        使用JSON文件导入实例（适用于Dashboard等模型）

        EasyOps API: ImportInstanceWithJson
        服务: cmdb.instance.ImportInstanceWithJson

        :param object_id: 实例所属的模型ID，如 '_DASHBOARD'
        :param file_path: JSON文件路径
        :param keys: 导入实例的字段,作为更新/创建依据（默认 ['instanceId']）
        :param create_if_id_not_exists: 当instanceId不存在时是否创建新实例（默认True）
        :return: 导入结果
        :rtype: dict
        """
        if not keys:
            keys = ['instanceId']

        port = 8079
        path = f'import/object/{object_id}/instance/json'

        files = {'attachment': (os.path.basename(file_path), open(file_path, 'rb'), 'application/json')}
        form_data = {}
        for i, key in enumerate(keys):
            form_data[f'keys[{i}]'] = key
        if create_if_id_not_exists:
            form_data['createIfIdNotExists'] = 'true'

        try:
            response = self._request('POST', path, port=port, timeout=120, files=files, form_data=form_data)
            result = response.json()
            data = result.get('data', result)
            if isinstance(data, dict):
                logger.info(f"JSON导入完成 [{object_id}]: 插入 {data.get('insert_count', 0)}, 更新 {data.get('update_count', 0)}, 失败 {data.get('failed_count', 0)}")
                if data.get('failed_count', 0) > 0:
                    logger.warning(f"失败详情: {data.get('data', [])}")
            else:
                logger.info(f"JSON导入完成 [{object_id}]: 返回数据条数 {len(data) if isinstance(data, list) else 'N/A'}")
            return data
        finally:
            files['attachment'][1].close()

    def import_dashboard(self, data: Union[dict, list, str], object_id: str = '_DASHBOARD', keys: List[str] = None, create_if_id_not_exists: bool = True) -> dict:
        """
        导入仪表盘配置（支持直接传入JSON数据）

        EasyOps API: ImportInstanceWithJson
        服务: logic.cmdb.service

        :param data: 仪表盘配置数据，支持以下格式：
                     - dict/list: Python字典或列表，会自动转为JSON
                     - str: JSON文件路径或JSON字符串
        :param object_id: 模型ID，默认 '_DASHBOARD'
        :param keys: 导入实例的字段,作为更新/创建依据（默认 ['name']）
        :param create_if_id_not_exists: 当instanceId不存在时是否创建新实例（默认True）
        :return: 导入结果
        :rtype: dict
        """
        import tempfile
        import io

        if not keys:
            keys = ['name']

        # 处理不同类型的输入
        if isinstance(data, (dict, list)):
            # 直接传入的字典或列表
            json_content = json.dumps(data if isinstance(data, list) else [data], ensure_ascii=False)
            file_obj = io.BytesIO(json_content.encode('utf-8'))
            filename = 'dashboard.json'
        elif isinstance(data, str):
            if os.path.isfile(data):
                # 文件路径
                file_obj = open(data, 'rb')
                filename = os.path.basename(data)
            else:
                # JSON字符串
                file_obj = io.BytesIO(data.encode('utf-8'))
                filename = 'dashboard.json'
        else:
            raise ValueError(f"不支持的数据类型: {type(data)}")

        port = 8079
        path = f'import/object/{object_id}/instance/json'

        files = {'attachment': (filename, file_obj, 'application/json')}
        form_data = {}
        for i, key in enumerate(keys):
            form_data[f'keys[{i}]'] = key
        if create_if_id_not_exists:
            form_data['createIfIdNotExists'] = 'true'

        try:
            response = self._request('POST', path, port=port, timeout=120, files=files, form_data=form_data)
            result = response.json()
            data_result = result.get('data', result)
            if isinstance(data_result, list):
                success = sum(1 for item in data_result if item.get('code') == 0)
                failed = len(data_result) - success
                logger.info(f"仪表盘导入完成 [{object_id}]: 成功 {success}, 失败 {failed}")
                if failed > 0:
                    failures = [item for item in data_result if item.get('code') != 0]
                    logger.warning(f"失败详情: {failures}")
            return data_result
        finally:
            file_obj.close()

    def get_unique_required_attrs(self, object_id: str) -> List[str]:
        """
        获取模型中唯一且必填的属性ID列表

        :param object_id: 模型ID
        :return: 唯一且必填的属性ID列表
        """
        try:
            model_info = self.get_model_desc(object_id)
            attrs = model_info.get('attrList', [])
            unique_required = []
            for attr in attrs:
                # unique 和 required 可能是布尔值或字符串
                is_unique = attr.get('unique') in [True, 'true']
                is_required = attr.get('required') in [True, 'true']
                if is_unique and is_required:
                    unique_required.append(attr['id'])
            return unique_required
        except Exception as e:
            logger.error(f"获取模型 {object_id} 属性信息失败: {e}")
            return []

    def _mark_failed_rows_in_excel(self, file_path: str, failed_data: List[dict]) -> Optional[str]:
        """
        在Excel文件中标记失败的行（红色背景）并添加错误信息列

        :param file_path: 原Excel文件路径
        :param failed_data: 失败详情列表，格式为 [{'line': 行号, 'error': '错误信息'}, ...]
        :return: 标记后的文件路径，失败返回None
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import PatternFill, Font, Alignment
        except ImportError:
            logger.error("openpyxl 未安装，无法标记失败行。请运行: pip install openpyxl")
            return None

        if not failed_data:
            return None

        try:
            wb = load_workbook(file_path)
            ws = wb.active

            # 红色背景填充
            red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
            red_font = Font(color="CC0000")

            # 找到最后一列，添加"错误信息"列
            max_col = ws.max_column
            error_col = max_col + 1

            # 在表头行添加"错误信息"列标题（假设第3行是表头，与API返回的line对应）
            header_row = 3
            ws.cell(row=header_row, column=error_col, value="导入错误信息")
            ws.cell(row=header_row, column=error_col).font = Font(bold=True, color="CC0000")

            # 构建行号到错误信息的映射
            failed_lines = {item['line']: item['error'] for item in failed_data}

            # 标记失败行
            for line_num, error_msg in failed_lines.items():
                # 将整行标记为红色背景
                for col in range(1, error_col + 1):
                    cell = ws.cell(row=line_num, column=col)
                    cell.fill = red_fill
                # 在错误信息列写入错误原因
                error_cell = ws.cell(row=line_num, column=error_col)
                error_cell.value = error_msg
                error_cell.font = red_font
                error_cell.alignment = Alignment(wrap_text=True)

            # 调整错误信息列宽度
            ws.column_dimensions[ws.cell(row=1, column=error_col).column_letter].width = 50

            # 生成新文件名
            base, ext = os.path.splitext(file_path)
            marked_file = f"{base}_marked{ext}"
            wb.save(marked_file)
            wb.close()
            return marked_file

        except Exception as e:
            logger.error(f"标记失败行时出错: {e}")
            return None

    def init_model(
        self,
        objectId: str,
        name: str,
        data_sample: Union[List[Dict[str, Any]], Dict[str, Any]],
        category: Optional[str] = None,
        memo: str = '',
        show_key: List[str] = ["instanceId"],
        attr_order: List[str] = []
    ) -> None:
        """
        初始化模型并生成模型描述信息。

        该方法会根据传入的数据样本自动推断字段类型,并构建模型的 `attrList` 描述信息

        :param objectId: 模型的唯一标识符（ID）。
        :param name: 模型的名称。
        :param data_sample: 数据样本,可以是字典或包含多个字典的列表。
                            如果是列表,则会对所有字典的键进行合并处理。
        :param category: 模型分类（可选）。
        :param memo: 模型备注信息（可选）。
        :param show_key: 模型展示的关键字段列表,默认为 ["instanceId"]。
        :param attr_order: 模型字段的显示顺序（可选）。
        :return: 无返回值。
        :rtype: None
        """

        # 如果 data_sample 是列表,则合并所有字典的键值对
        if isinstance(data_sample, list):
            keys = set(k for i in data_sample for k, v in i.items())
            tmp = {}
            for k in keys:
                vs = [i[k] for i in data_sample if i.get(k)]
                if not vs:
                    continue
                tmp[k] = vs[0]
            data_sample = tmp

        # 类型映射字典
        type_dic = {
            'str': 'str',
            'unicode': 'str',
            'bool': 'bool',
            'int': 'int',
            'float': 'float',
        }

        # 构建模型基本信息
        model_info = {
            "objectId": objectId,
            "name": name,
            "category": category,
            "memo": memo,
            "protected": False,
            'isAbstract': False,
            'system': '',
            'wordIndexDenied': False,
            "notifyDenied": False,
            "view": {
                "attr_authorizers": None,
                "attr_category_order": ["基本信息"],
                "attr_order": attr_order,
                "hide_columns": None,
                "icon": {
                    "category": "second-menu",
                    "icon": "placeholder-second-menu",
                    "lib": "easyops"
                },
                "image": "",
                "inherit_attr_category_map": None,
                "relation_default_attr": None,
                "relation_group_order": None,
                "relation_order": None,
                "show_key": show_key,
                "visible": True
            },
            "attrList": [],
        }

        # 尝试导入翻译工具
        try:
            from tools import VariableTranslator
            translator = VariableTranslator()
        except ImportError:
            translator = None

        # 遍历数据样本,生成字段描述
        for k, v in data_sample.items():
            attr = {
                "id": k,
                "name": k if not translator else translator.translate_variable_name(k),
                'description': k,
                "tag": [],
                "required": "false",
                "readonly": "false",
                "unique": "false"
            }

            # 推断字段类型
            v_type = re.findall(r"<.*?'(.*)'>", str(type(v)))[0]
            if v_type in type_dic:
                attr["value"] = {
                    "default": "",
                    "mode": "default",
                    "default_type": "value",
                    "type": type_dic[v_type]
                }
            elif v_type == 'dict':
                attr["value"] = {
                    "default": "",
                    "struct_define": [],
                    "type": "struct"
                }
                for k_, v_ in v.items():
                    v_type_ = str(type(v_))[7:-2]
                    if v_type_ not in type_dic:
                        continue
                    attr['value']['struct_define'].append({
                        "id": k_,
                        "name": k_ if not translator else translator.translate_variable_name(k_),
                        "type": type_dic[v_type_],
                        "regex": "",
                    })
            elif v_type == 'list':
                if all(isinstance(i, str) for i in v):
                    attr["value"] = {
                        "default": "",
                        "mode": "default",
                        "default_type": "value",
                        "type": 'arr'
                    }
                elif all(isinstance(i, dict) for i in v):
                    attr["value"] = {
                        "default": "",
                        "struct_define": [],
                        "type": "structs"
                    }
                    for k_, v_ in v[0].items():
                        v_type_ = str(type(v_))[7:-2]
                        if v_type_ not in type_dic:
                            continue
                        attr['value']['struct_define'].append({
                            "id": k_,
                            "name": k_ if not translator else translator.translate_variable_name(k_),
                            "type": type_dic[v_type_],
                            "regex": "",
                        })
                else:
                    attr["value"] = {
                        "default": [],
                        "mode": "default",
                        "type": "arr"
                    }

            model_info['attrList'].append(attr)

        # 设置字段显示顺序
        if not model_info['view']['attr_order']:
            model_info['view']['attr_order'] = sorted([i['id'] for i in model_info['attrList']])

        # 导入模型
        self.import_model(json.dumps([model_info]))

    def list_object_basic(self, q: str = None, page_size: int = 3000, visible: str = "visible", **kwargs) -> list:
        """
        获取模型基本信息列表

        EasyOps API: ListObjectBasic
        服务: logic.cmdb.service

        :param q: 按模型Id模糊匹配
        :param page_size: 页大小,默认3000
        :param object_ids: 模型Id, 使用逗号分隔
        :param category: 分类 (xx:精确匹配，xx%:匹配xx或xx.*，xx.%:匹配xx\\.*)
        :param visible: visible/invisible/all，默认all
        :return: 模型基本信息列表
        :rtype: list
        """
        port = 8079
        params = {'page': 1, 'page_size': page_size}
        if q:
            params['q'] = q
        if visible:
            params['visible'] = visible
        for key in ['objectIds', 'category', 'isAbstract', 'fields']:
            if kwargs.get(key):
                params[key] = kwargs[key]

        objects = []
        for page in range(1, 10000):
            params['page'] = page
            response = self._request('GET', 'object_basic', port=port, params=params)
            data = response.json().get('data', {})
            tmp = data.get('list', [])
            objects.extend(tmp)
            if len(tmp) < page_size:
                break
        logger.info(f"Found {len(objects)} objects.")
        return objects

    def export_model_doc(self, model_ids: List[str] = None, category: str = None,
                         output_path: str = None, sort_by: str = 'id',
                         include_hidden: bool = False) -> str:
        """
        导出模型说明文档到Excel

        生成包含模型列表和各模型详情的Excel文件：
        - 第一个sheet：模型列表（按分类展示，可跳转到详情sheet）
        - 后续sheet：每个模型的属性和关系详情（可跳转回列表）

        :param model_ids: 要导出的模型ID列表，为空则导出全部
        :param category: 按分类筛选模型（支持模糊匹配）
        :param output_path: 输出文件路径，默认 ./output/model_doc_{timestamp}.xlsx
        :param sort_by: 属性排序字段，可选 'id'(默认), 'name', 'tag'
        :param include_hidden: 是否包含隐藏模型，默认False只导出非隐藏模型
        :return: 生成的Excel文件路径
        :rtype: str
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.comments import Comment
        except ImportError:
            raise ImportError("请先安装 openpyxl: pip install openpyxl")

        # 获取模型列表，根据 include_hidden 参数决定是否过滤隐藏模型
        visible_param = 'all' if include_hidden else 'visible'
        models_basic = self.list_object_basic(category=category, visible=visible_param) if category else self.list_object_basic(visible=visible_param)
        if model_ids:
            models_basic = [m for m in models_basic if m['objectId'] in model_ids]

        if not models_basic:
            logger.warning("没有找到符合条件的模型")
            return None

        # 按分类分组
        category_models = {}
        for m in models_basic:
            cat = m.get('category', '未分类')
            if cat not in category_models:
                category_models[cat] = []
            category_models[cat].append(m)

        # 创建工作簿
        wb = Workbook()
        ws_list = wb.active
        ws_list.title = "模型列表"

        # 样式定义
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        link_font = Font(color="0563C1", underline="single")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        cat_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

        # 写入模型列表表头
        list_headers = ["分类", "模型ID", "模型名称", "描述", "是否抽象", "详情"]
        for col, header in enumerate(list_headers, 1):
            cell = ws_list.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # 写入模型列表数据
        row = 2
        model_sheet_map = {}  # 记录模型ID到sheet名的映射
        model_row_map = {}    # 记录模型ID到列表行号的映射

        for cat in sorted(category_models.keys()):
            for m in category_models[cat]:
                obj_id = m['objectId']
                model_name = m.get('name', obj_id)
                # sheet名用模型名称，最多31字符，去除特殊字符，处理重名
                base_sheet_name = re.sub(r'[\\/*?:\[\]]', '_', model_name)[:28]
                sheet_name = base_sheet_name
                suffix = 1
                while sheet_name in model_sheet_map.values():
                    sheet_name = f"{base_sheet_name}_{suffix}"[:31]
                    suffix += 1
                model_sheet_map[obj_id] = sheet_name
                model_row_map[obj_id] = row

                ws_list.cell(row=row, column=1, value=cat).border = thin_border
                ws_list.cell(row=row, column=2, value=obj_id).border = thin_border
                ws_list.cell(row=row, column=3, value=model_name).border = thin_border
                ws_list.cell(row=row, column=4, value=m.get('memo', '')).border = thin_border
                ws_list.cell(row=row, column=5, value='是' if m.get('isAbstract') else '否').border = thin_border

                # 详情链接 - 使用HYPERLINK函数跳转到对应sheet的A1
                link_cell = ws_list.cell(row=row, column=6)
                link_cell.value = f'=HYPERLINK("#\'{sheet_name}\'!A1","查看详情")'
                link_cell.font = link_font
                link_cell.border = thin_border
                row += 1

        # 设置列宽
        ws_list.column_dimensions['A'].width = 15
        ws_list.column_dimensions['B'].width = 25
        ws_list.column_dimensions['C'].width = 20
        ws_list.column_dimensions['D'].width = 40
        ws_list.column_dimensions['E'].width = 10
        ws_list.column_dimensions['F'].width = 12

        # 收集所有struct属性，最后统一创建子sheet
        all_struct_attrs = []

        # 为每个模型创建详情sheet
        for m in models_basic:
            obj_id = m['objectId']
            sheet_name = model_sheet_map[obj_id]

            # 获取模型详情
            try:
                model_detail = self.get_model_desc(obj_id)
            except Exception as e:
                logger.warning(f"获取模型 {obj_id} 详情失败: {e}")
                continue

            ws = wb.create_sheet(title=sheet_name)

            # 模型基本信息
            ws.cell(row=1, column=1, value="模型ID:").font = Font(bold=True)
            ws.cell(row=1, column=2, value=obj_id)
            ws.cell(row=1, column=3, value="模型名称:").font = Font(bold=True)
            ws.cell(row=1, column=4, value=model_detail.get('name', ''))

            # 返回列表链接 - 使用HYPERLINK函数跳转到模型列表中对应的行
            back_cell = ws.cell(row=1, column=6)
            list_row = model_row_map.get(obj_id, 1)
            back_cell.value = f'=HYPERLINK("#\'模型列表\'!A{list_row}","返回列表")'
            back_cell.font = link_font

            ws.cell(row=2, column=1, value="分类:").font = Font(bold=True)
            ws.cell(row=2, column=2, value=model_detail.get('category', ''))
            ws.cell(row=2, column=3, value="描述:").font = Font(bold=True)
            ws.cell(row=2, column=4, value=model_detail.get('memo', ''))

            # 属性列表标题
            ws.cell(row=4, column=1, value="属性列表").font = Font(bold=True, size=12)
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=10)

            # 属性表头及提示
            attr_headers = [
                ("属性ID", "只读，不可修改"),
                ("属性名称", "可修改，显示名称"),
                ("类型", "只读，不可修改"),
                ("必填", "下拉选择：是/否"),
                ("唯一", "下拉选择：是/否"),
                ("只读", "下拉选择：是/否"),
                ("分组", "多个分组用逗号分隔"),
                ("描述", "可修改，属性说明"),
                ("默认值", "只读，不可修改"),
                ("枚举/结构", "只读，不可修改")
            ]
            for col, (header, tip) in enumerate(attr_headers, 1):
                cell = ws.cell(row=5, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.comment = Comment(tip, "系统提示")

            # 创建是/否下拉验证
            yes_no_validation = DataValidation(type="list", formula1='"是,否"', allow_blank=True)
            yes_no_validation.error = "请选择 是 或 否"
            yes_no_validation.errorTitle = "输入错误"
            yes_no_validation.prompt = "请选择"
            yes_no_validation.promptTitle = "选择值"
            ws.add_data_validation(yes_no_validation)

            # 获取属性列表并排序
            attr_list = model_detail.get('attrList', [])
            if sort_by == 'name':
                attr_list = sorted(attr_list, key=lambda x: x.get('name', ''))
            elif sort_by == 'tag':
                attr_list = sorted(attr_list, key=lambda x: (x.get('tag', [''])[0] if x.get('tag') else '', x.get('id', '')))
            else:
                attr_list = sorted(attr_list, key=lambda x: x.get('id', ''))

            # 写入属性数据
            attr_row = 6
            readonly_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            for attr in attr_list:
                value_info = attr.get('value', {})
                attr_type = value_info.get('type', 'str')

                # 枚举/结构信息
                extra_info = ''
                if attr_type in ['enum', 'enums'] and value_info.get('regex'):
                    regex = value_info.get('regex')
                    if isinstance(regex, list):
                        extra_info = ', '.join(str(r) for r in regex[:10])
                        if len(regex) > 10:
                            extra_info += f'... (共{len(regex)}项)'
                elif attr_type in ['struct', 'structs'] and value_info.get('struct_define'):
                    struct_define = value_info.get('struct_define', [])
                    all_struct_attrs.append({
                        'model_id': obj_id,
                        'model_name': model_detail.get('name', obj_id),
                        'model_sheet_name': sheet_name,
                        'attr_id': attr.get('id'),
                        'attr_name': attr.get('name'),
                        'struct_define': struct_define
                    })
                    extra_info = f"[查看结构定义] 共{len(struct_define)}个字段"

                # 只读列（灰色背景）：属性ID、类型、默认值、枚举/结构
                id_cell = ws.cell(row=attr_row, column=1, value=attr.get('id', ''))
                id_cell.border = thin_border
                id_cell.fill = readonly_fill

                ws.cell(row=attr_row, column=2, value=attr.get('name', '')).border = thin_border

                type_cell = ws.cell(row=attr_row, column=3, value=attr_type)
                type_cell.border = thin_border
                type_cell.fill = readonly_fill

                # 必填、唯一、只读 - 添加下拉验证
                req_cell = ws.cell(row=attr_row, column=4, value='是' if attr.get('required') == 'true' else '否')
                req_cell.border = thin_border
                yes_no_validation.add(req_cell)

                uniq_cell = ws.cell(row=attr_row, column=5, value='是' if attr.get('unique') == 'true' else '否')
                uniq_cell.border = thin_border
                yes_no_validation.add(uniq_cell)

                ro_cell = ws.cell(row=attr_row, column=6, value='是' if attr.get('readonly') == 'true' else '否')
                ro_cell.border = thin_border
                yes_no_validation.add(ro_cell)

                ws.cell(row=attr_row, column=7, value=', '.join(attr.get('tag', []))).border = thin_border
                ws.cell(row=attr_row, column=8, value=attr.get('description', '')).border = thin_border

                default_cell = ws.cell(row=attr_row, column=9, value=str(value_info.get('default', '')))
                default_cell.border = thin_border
                default_cell.fill = readonly_fill

                # 枚举/结构列 - struct类型添加跳转链接
                extra_cell = ws.cell(row=attr_row, column=10, value=extra_info)
                extra_cell.border = thin_border
                extra_cell.fill = readonly_fill
                if attr_type in ['struct', 'structs'] and value_info.get('struct_define'):
                    struct_sheet_name = f"S_{attr.get('id', '')}"[:31]
                    extra_cell.value = f'=HYPERLINK("#\'{struct_sheet_name}\'!A1","查看结构定义")'
                    extra_cell.font = link_font

                attr_row += 1

            # 关系列表
            relation_row = attr_row + 2
            ws.cell(row=relation_row, column=1, value="关系列表").font = Font(bold=True, size=12)
            ws.merge_cells(start_row=relation_row, start_column=1, end_row=relation_row, end_column=10)

            rel_headers = [
                ("关系ID", "只读，不可修改"),
                ("关系名称", "可修改"),
                ("本端ID", "只读，不可修改"),
                ("本端名称", "可修改"),
                ("对端模型", "新增时必填，已有关系不可修改"),
                ("对端ID", "只读，不可修改"),
                ("对端名称", "可修改"),
                ("本端基数", "1或多，新增时可设置"),
                ("对端基数", "1或多，新增时可设置"),
                ("描述", "可修改")
            ]
            relation_row += 1
            for col, (header, tip) in enumerate(rel_headers, 1):
                cell = ws.cell(row=relation_row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.comment = Comment(tip, "系统提示")

            # 写入关系数据（只导出当前模型自己定义的关系，排除继承的）
            relation_row += 1
            relation_list = model_detail.get('relation_list', [])
            for rel in relation_list:
                # 跳过继承的关系
                if rel.get('isInherit'):
                    continue

                # 判断当前模型是左侧还是右侧
                left_obj = rel.get('left_object_id', '')
                right_obj = rel.get('right_object_id', '')
                if left_obj == obj_id:
                    local_id = rel.get('left_id', '')
                    local_name = rel.get('left_name', '')
                    local_max = rel.get('left_max', 0)
                    remote_obj = rel.get('right_object_id', '')
                    remote_id = rel.get('right_id', '')
                    remote_name = rel.get('right_name', '')
                    remote_max = rel.get('right_max', 0)
                else:
                    local_id = rel.get('right_id', '')
                    local_name = rel.get('right_name', '')
                    local_max = rel.get('right_max', 0)
                    remote_obj = rel.get('left_object_id', '')
                    remote_id = rel.get('left_id', '')
                    remote_name = rel.get('left_name', '')
                    remote_max = rel.get('left_max', 0)

                def format_cardinality(val):
                    return '多' if val == -1 else str(val)

                ws.cell(row=relation_row, column=1, value=rel.get('relation_id', '')).border = thin_border
                ws.cell(row=relation_row, column=2, value=rel.get('name', '')).border = thin_border
                ws.cell(row=relation_row, column=3, value=local_id).border = thin_border
                ws.cell(row=relation_row, column=4, value=local_name).border = thin_border

                # 对端模型添加跳转链接 - 使用HYPERLINK函数
                remote_cell = ws.cell(row=relation_row, column=5)
                remote_cell.border = thin_border
                if remote_obj in model_sheet_map:
                    remote_cell.value = f'=HYPERLINK("#\'{model_sheet_map[remote_obj]}\'!A1","{remote_obj}")'
                    remote_cell.font = link_font
                else:
                    remote_cell.value = remote_obj

                ws.cell(row=relation_row, column=6, value=remote_id).border = thin_border
                ws.cell(row=relation_row, column=7, value=remote_name).border = thin_border
                ws.cell(row=relation_row, column=8, value=format_cardinality(local_max)).border = thin_border
                ws.cell(row=relation_row, column=9, value=format_cardinality(remote_max)).border = thin_border
                ws.cell(row=relation_row, column=10, value=rel.get('left_description', '')).border = thin_border
                relation_row += 1

            # 设置列宽
            col_widths = [20, 15, 12, 15, 20, 12, 15, 10, 10, 30]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

        # 所有模型sheet创建完成后，统一创建struct子sheet
        for struct_info in all_struct_attrs:
            struct_sheet_name = f"S_{struct_info['attr_id']}"[:31]
            ws_struct = wb.create_sheet(title=struct_sheet_name)

            # 标题信息
            ws_struct.cell(row=1, column=1, value="模型ID:").font = Font(bold=True)
            ws_struct.cell(row=1, column=2, value=struct_info['model_id'])
            ws_struct.cell(row=1, column=3, value="属性ID:").font = Font(bold=True)
            ws_struct.cell(row=1, column=4, value=struct_info['attr_id'])

            # 返回模型sheet链接 - 使用HYPERLINK函数
            back_cell = ws_struct.cell(row=1, column=6)
            back_cell.value = f'=HYPERLINK("#\'{struct_info["model_sheet_name"]}\'!A1","返回模型")'
            back_cell.font = link_font

            ws_struct.cell(row=2, column=1, value="模型名称:").font = Font(bold=True)
            ws_struct.cell(row=2, column=2, value=struct_info['model_name'])
            ws_struct.cell(row=2, column=3, value="属性名称:").font = Font(bold=True)
            ws_struct.cell(row=2, column=4, value=struct_info['attr_name'])

            # 结构体字段表头
            ws_struct.cell(row=4, column=1, value="结构体字段定义").font = Font(bold=True, size=12)
            struct_headers = [
                ("字段ID", "必填，字段唯一标识"),
                ("字段名称", "必填，显示名称"),
                ("字段类型", "str/int/float/bool/date/datetime/ip/json"),
                ("必填", "是/否"),
                ("描述", "字段说明")
            ]
            for col, (header, tip) in enumerate(struct_headers, 1):
                cell = ws_struct.cell(row=5, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.comment = Comment(tip, "系统提示")

            # 写入结构体字段数据
            struct_row = 6
            for field in struct_info['struct_define']:
                ws_struct.cell(row=struct_row, column=1, value=field.get('id', '')).border = thin_border
                ws_struct.cell(row=struct_row, column=2, value=field.get('name', '')).border = thin_border
                ws_struct.cell(row=struct_row, column=3, value=field.get('type', 'str')).border = thin_border
                ws_struct.cell(row=struct_row, column=4, value='是' if field.get('required') else '否').border = thin_border
                ws_struct.cell(row=struct_row, column=5, value=field.get('description', '')).border = thin_border
                struct_row += 1

            # 设置列宽
            ws_struct.column_dimensions['A'].width = 20
            ws_struct.column_dimensions['B'].width = 20
            ws_struct.column_dimensions['C'].width = 15
            ws_struct.column_dimensions['D'].width = 10
            ws_struct.column_dimensions['E'].width = 30

        # 保存文件
        if not output_path:
            os.makedirs('./output', exist_ok=True)
            output_path = f'./output/model_doc_{int(time.time())}.xlsx'

        wb.save(output_path)
        logger.info(f"模型文档已导出到: {output_path}")
        return output_path

    def import_model_doc(self, excel_path: str, dry_run: bool = True) -> dict:
        """
        从Excel文档解析并优化模型定义

        读取export_model_doc导出的Excel，解析用户修改后的属性/关系信息，
        生成模型更新操作。支持：
        - 属性：新增、修改、删除
        - 关系：新增、修改、删除

        新增属性：在Excel属性列表末尾添加新行，填写属性ID（必填）、名称、类型等
        删除属性：删除Excel中的属性行（protected属性不可删除）
        新增关系：在Excel关系列表末尾添加新行，填写关系ID、对端模型等
        删除关系：删除Excel中的关系行

        :param excel_path: Excel文件路径
        :param dry_run: 试运行模式，默认True只输出变更不执行
        :return: 变更摘要 {model_id: {attr_added, attr_modified, attr_deleted, rel_added, rel_modified, rel_deleted}}
        :rtype: dict
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("请先安装 openpyxl: pip install openpyxl")

        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"文件不存在: {excel_path}")

        wb = load_workbook(excel_path)
        changes = {}

        # 跳过第一个sheet（模型列表）和struct子sheet（以S_开头），处理模型详情sheet
        for sheet_name in wb.sheetnames[1:]:
            # 跳过struct子sheet
            if sheet_name.startswith('S_'):
                continue

            ws = wb[sheet_name]

            # 获取模型ID（第一行第二列）
            model_id = ws.cell(row=1, column=2).value
            if not model_id:
                logger.warning(f"Sheet '{sheet_name}' 未找到模型ID，跳过")
                continue

            # 获取当前模型详情
            try:
                current_model = self.get_model_desc(model_id)
            except Exception as e:
                logger.warning(f"获取模型 {model_id} 失败: {e}，跳过")
                continue

            model_changes = {
                'attr_added': [], 'attr_modified': [], 'attr_deleted': [],
                'rel_added': [], 'rel_modified': [], 'rel_deleted': []
            }

            # ========== 解析属性列表（从第6行开始） ==========
            excel_attrs = {}
            row = 6
            relation_start_row = None
            while row <= ws.max_row:
                attr_id = ws.cell(row=row, column=1).value
                # 找到关系列表标记
                if attr_id == '关系列表':
                    relation_start_row = row + 2  # 跳过标题行
                    break
                # 跳过空行
                if not attr_id or str(attr_id).strip() == '':
                    row += 1
                    continue
                attr_type = ws.cell(row=row, column=3).value or 'str'
                attr_data = {
                    'id': attr_id,
                    'name': ws.cell(row=row, column=2).value or attr_id,
                    'type': attr_type,
                    'required': 'true' if ws.cell(row=row, column=4).value == '是' else 'false',
                    'unique': 'true' if ws.cell(row=row, column=5).value == '是' else 'false',
                    'readonly': 'true' if ws.cell(row=row, column=6).value == '是' else 'false',
                    'tag': [t.strip() for t in (ws.cell(row=row, column=7).value or '').split(',') if t.strip()],
                    'description': ws.cell(row=row, column=8).value or '',
                    'value': {'type': attr_type}
                }

                # 如果是struct/structs类型，尝试从子sheet读取struct_define
                if attr_type in ['struct', 'structs']:
                    struct_sheet_name = f"S_{attr_id}"[:31]
                    if struct_sheet_name in wb.sheetnames:
                        struct_define = self._parse_struct_sheet(wb[struct_sheet_name])
                        if struct_define:
                            attr_data['value']['struct_define'] = struct_define

                excel_attrs[attr_id] = attr_data
                row += 1

            # 比较属性变更
            current_attrs = {a['id']: a for a in current_model.get('attrList', [])}

            # 检查新增和修改的属性
            for attr_id, excel_attr in excel_attrs.items():
                if attr_id not in current_attrs:
                    # 新增属性
                    model_changes['attr_added'].append({
                        'id': attr_id,
                        'data': excel_attr
                    })
                else:
                    # 检查修改
                    current_attr = current_attrs[attr_id]
                    changes_detail = []

                    if excel_attr['name'] != current_attr.get('name', ''):
                        changes_detail.append(f"名称: {current_attr.get('name')} -> {excel_attr['name']}")
                    if excel_attr['description'] != current_attr.get('description', ''):
                        changes_detail.append(f"描述: {current_attr.get('description', '')} -> {excel_attr['description']}")
                    if excel_attr['required'] != current_attr.get('required', 'false'):
                        changes_detail.append(f"必填: {current_attr.get('required')} -> {excel_attr['required']}")
                    if excel_attr['unique'] != current_attr.get('unique', 'false'):
                        changes_detail.append(f"唯一: {current_attr.get('unique')} -> {excel_attr['unique']}")
                    if excel_attr['readonly'] != current_attr.get('readonly', 'false'):
                        changes_detail.append(f"只读: {current_attr.get('readonly')} -> {excel_attr['readonly']}")

                    current_tags = current_attr.get('tag', [])
                    if set(excel_attr['tag']) != set(current_tags):
                        changes_detail.append(f"分组: {current_tags} -> {excel_attr['tag']}")

                    if changes_detail:
                        model_changes['attr_modified'].append({
                            'id': attr_id,
                            'changes': changes_detail,
                            'new_values': excel_attr
                        })

            # 检查删除的属性
            for attr_id, attr in current_attrs.items():
                if attr_id not in excel_attrs and not attr.get('protected', False):
                    model_changes['attr_deleted'].append(attr_id)

            # ========== 解析关系列表 ==========
            excel_rels = {}
            # 先获取当前模型的关系，用于判断当前模型是左端还是右端
            current_rels_raw = {r['relation_id']: r for r in current_model.get('relation_list', [])}

            if relation_start_row:
                row = relation_start_row
                while row <= ws.max_row:
                    rel_id = ws.cell(row=row, column=1).value
                    if not rel_id or str(rel_id).strip() == '':
                        row += 1
                        continue

                    # 处理 HYPERLINK 函数提取实际值
                    def extract_value(cell_value):
                        if cell_value and str(cell_value).startswith('=HYPERLINK'):
                            # 从 =HYPERLINK("#'sheet'!A1","显示文本") 提取显示文本
                            import re
                            match = re.search(r'"([^"]+)"\s*\)$', str(cell_value))
                            if match:
                                return match.group(1)
                        return cell_value or ''

                    remote_obj = extract_value(ws.cell(row=row, column=5).value)
                    local_id = ws.cell(row=row, column=3).value or ''
                    local_name = ws.cell(row=row, column=4).value or ''
                    remote_id = ws.cell(row=row, column=6).value or ''
                    remote_name = ws.cell(row=row, column=7).value or ''
                    local_max_str = ws.cell(row=row, column=8).value or '1'
                    remote_max_str = ws.cell(row=row, column=9).value or '1'

                    def parse_cardinality(val):
                        if val == '多' or val == '-1':
                            return -1
                        try:
                            return int(val)
                        except:
                            return 1

                    # 判断当前模型是左端还是右端（根据API返回的关系数据）
                    current_rel = current_rels_raw.get(rel_id, {})
                    is_left = current_rel.get('left_object_id') == model_id

                    if is_left:
                        # 当前模型是左端，Excel的本端=API的left，对端=API的right
                        excel_rels[rel_id] = {
                            'relation_id': rel_id,
                            'name': ws.cell(row=row, column=2).value or '',
                            'left_object_id': model_id,
                            'left_id': local_id,
                            'left_name': local_name,
                            'left_max': parse_cardinality(local_max_str),
                            'right_object_id': remote_obj,
                            'right_id': remote_id,
                            'right_name': remote_name,
                            'right_max': parse_cardinality(remote_max_str),
                            'left_description': ws.cell(row=row, column=10).value or ''
                        }
                    else:
                        # 当前模型是右端，Excel的本端=API的right，对端=API的left
                        excel_rels[rel_id] = {
                            'relation_id': rel_id,
                            'name': ws.cell(row=row, column=2).value or '',
                            'left_object_id': remote_obj,
                            'left_id': remote_id,
                            'left_name': remote_name,
                            'left_max': parse_cardinality(remote_max_str),
                            'right_object_id': model_id,
                            'right_id': local_id,
                            'right_name': local_name,
                            'right_max': parse_cardinality(local_max_str),
                            'left_description': ws.cell(row=row, column=10).value or ''
                        }
                    row += 1

            # 比较关系变更
            current_rels = current_rels_raw

            # 检查新增和修改的关系
            for rel_id, excel_rel in excel_rels.items():
                if rel_id not in current_rels:
                    # 新增关系
                    if excel_rel['right_object_id']:  # 必须有对端模型
                        model_changes['rel_added'].append({
                            'id': rel_id,
                            'data': excel_rel
                        })
                else:
                    # 检查修改
                    current_rel = current_rels[rel_id]
                    changes_detail = []

                    if excel_rel['name'] != current_rel.get('name', ''):
                        changes_detail.append(f"名称: {current_rel.get('name')} -> {excel_rel['name']}")
                    if excel_rel['left_name'] != current_rel.get('left_name', ''):
                        changes_detail.append(f"本端名称: {current_rel.get('left_name')} -> {excel_rel['left_name']}")
                    if excel_rel['right_name'] != current_rel.get('right_name', ''):
                        changes_detail.append(f"对端名称: {current_rel.get('right_name')} -> {excel_rel['right_name']}")

                    if changes_detail:
                        model_changes['rel_modified'].append({
                            'id': rel_id,
                            'changes': changes_detail,
                            'new_values': excel_rel
                        })

            # 检查删除的关系（只检查当前模型自己的关系，排除继承的）
            for rel_id, rel in current_rels.items():
                # 跳过继承的关系
                if rel.get('isInherit'):
                    continue
                # 跳过内置关系（protected）
                is_protected = rel.get('protected', False)
                if rel_id not in excel_rels and not is_protected:
                    model_changes['rel_deleted'].append(rel_id)

            # 记录有变更的模型
            has_changes = any([
                model_changes['attr_added'], model_changes['attr_modified'], model_changes['attr_deleted'],
                model_changes['rel_added'], model_changes['rel_modified'], model_changes['rel_deleted']
            ])
            if has_changes:
                changes[model_id] = model_changes

        # 输出变更摘要
        if not changes:
            logger.info("未检测到任何变更")
            return {}

        logger.info(f"检测到 {len(changes)} 个模型有变更:")
        for model_id, model_changes in changes.items():
            logger.info(f"\n模型 {model_id}:")
            if model_changes['attr_added']:
                logger.info(f"  属性新增 ({len(model_changes['attr_added'])} 个):")
                for attr in model_changes['attr_added']:
                    logger.info(f"    + {attr['id']}: {attr['data'].get('name')} ({attr['data'].get('type', 'str')})")
            if model_changes['attr_modified']:
                logger.info(f"  属性修改 ({len(model_changes['attr_modified'])} 个):")
                for attr in model_changes['attr_modified']:
                    logger.info(f"    ~ {attr['id']}: {', '.join(attr['changes'])}")
            if model_changes['attr_deleted']:
                logger.info(f"  属性删除 ({len(model_changes['attr_deleted'])} 个): {model_changes['attr_deleted']}")
            if model_changes['rel_added']:
                logger.info(f"  关系新增 ({len(model_changes['rel_added'])} 个):")
                for rel in model_changes['rel_added']:
                    logger.info(f"    + {rel['id']}: {model_id} -> {rel['data'].get('right_object_id')}")
            if model_changes['rel_modified']:
                logger.info(f"  关系修改 ({len(model_changes['rel_modified'])} 个):")
                for rel in model_changes['rel_modified']:
                    logger.info(f"    ~ {rel['id']}: {', '.join(rel['changes'])}")
            if model_changes['rel_deleted']:
                logger.info(f"  关系删除 ({len(model_changes['rel_deleted'])} 个): {model_changes['rel_deleted']}")

        # 执行变更
        if not dry_run:
            logger.info("\n开始执行变更...")
            for model_id, model_changes in changes.items():
                # 新增属性
                for attr_add in model_changes['attr_added']:
                    try:
                        self.create_model_attr(model_id, attr_add['data'])
                        logger.info(f"  新增属性 {model_id}.{attr_add['id']} 成功")
                    except Exception as e:
                        logger.error(f"  新增属性 {model_id}.{attr_add['id']} 失败: {e}")

                # 更新属性
                for attr_change in model_changes['attr_modified']:
                    try:
                        self.update_model_attr(model_id, attr_change['id'], attr_change['new_values'])
                        logger.info(f"  更新属性 {model_id}.{attr_change['id']} 成功")
                    except Exception as e:
                        logger.error(f"  更新属性 {model_id}.{attr_change['id']} 失败: {e}")

                # 删除属性
                for attr_id in model_changes['attr_deleted']:
                    try:
                        self.delete_model_attr(model_id, attr_id)
                        logger.info(f"  删除属性 {model_id}.{attr_id} 成功")
                    except Exception as e:
                        logger.error(f"  删除属性 {model_id}.{attr_id} 失败: {e}")

                # 新增关系
                for rel_add in model_changes['rel_added']:
                    try:
                        self.create_relation(rel_add['data'])
                        logger.info(f"  新增关系 {rel_add['id']} 成功")
                    except Exception as e:
                        logger.error(f"  新增关系 {rel_add['id']} 失败: {e}")

                # 更新关系
                for rel_change in model_changes['rel_modified']:
                    try:
                        self.update_relation(rel_change['id'], rel_change['new_values'])
                        logger.info(f"  更新关系 {rel_change['id']} 成功")
                    except Exception as e:
                        logger.error(f"  更新关系 {rel_change['id']} 失败: {e}")

                # 删除关系
                for rel_id in model_changes['rel_deleted']:
                    try:
                        self.delete_relation(rel_id)
                        logger.info(f"  删除关系 {rel_id} 成功")
                    except Exception as e:
                        logger.error(f"  删除关系 {rel_id} 失败: {e}")
        else:
            logger.info("\n[试运行模式] 如需执行变更，请设置 dry_run=False")

        return changes

    def _parse_struct_sheet(self, ws) -> List[dict]:
        """解析struct子sheet，返回struct_define列表"""
        struct_define = []
        row = 6  # 数据从第6行开始
        while True:
            field_id = ws.cell(row=row, column=1).value
            if not field_id:
                break
            if str(field_id).strip() == '':
                row += 1
                continue
            struct_define.append({
                'id': field_id,
                'name': ws.cell(row=row, column=2).value or field_id,
                'type': ws.cell(row=row, column=3).value or 'str',
                'required': ws.cell(row=row, column=4).value == '是',
                'description': ws.cell(row=row, column=5).value or ''
            })
            row += 1
        return struct_define

    def create_model_attr(self, object_id: str, attr_data: dict) -> dict:
        """创建模型属性"""
        port = 8079
        path = f'object/{object_id}/attr'
        data = {
            'id': attr_data.get('id'),
            'name': attr_data.get('name'),
            'description': attr_data.get('description', ''),
            'required': attr_data.get('required', 'false'),
            'unique': attr_data.get('unique', 'false'),
            'readonly': attr_data.get('readonly', 'false'),
            'tag': attr_data.get('tag', []),
            'value': attr_data.get('value', {'type': 'str'})
        }
        response = self._request('POST', path, port=port, data=data)
        return response.json()

    def update_model_attr(self, object_id: str, attr_id: str, attr_data: dict) -> dict:
        """更新模型属性"""
        port = 8079
        path = f'object/{object_id}/attr/{attr_id}'
        data = {
            'name': attr_data.get('name'),
            'description': attr_data.get('description', ''),
            'required': attr_data.get('required', 'false'),
            'unique': attr_data.get('unique', 'false'),
            'readonly': attr_data.get('readonly', 'false'),
            'tag': attr_data.get('tag', []),
        }
        response = self._request('PUT', path, port=port, data=data)
        return response.json()

    def delete_model_attr(self, object_id: str, attr_id: str) -> dict:
        """删除模型属性"""
        port = 8079
        path = f'object/{object_id}/attr/{attr_id}'
        response = self._request('DELETE', path, port=port)
        return response.json()

    def create_relation(self, rel_data: dict) -> dict:
        """创建模型关系"""
        port = 8079
        path = 'object_relation'
        data = {
            'relation_id': rel_data.get('relation_id'),
            'name': rel_data.get('name', ''),
            'left_object_id': rel_data.get('left_object_id'),
            'left_id': rel_data.get('left_id'),
            'left_name': rel_data.get('left_name', ''),
            'left_max': rel_data.get('left_max', 1),
            'left_min': 0,
            'left_groups': [],
            'left_tags': [],
            'right_object_id': rel_data.get('right_object_id'),
            'right_id': rel_data.get('right_id'),
            'right_name': rel_data.get('right_name', ''),
            'right_max': rel_data.get('right_max', 1),
            'right_min': 0,
            'right_groups': [],
            'right_tags': [],
            'left_description': rel_data.get('left_description', ''),
            'right_description': rel_data.get('right_description', ''),
        }
        response = self._request('POST', path, port=port, data=data)
        return response.json()

    def update_relation(self, relation_id: str, rel_data: dict) -> dict:
        """更新模型关系"""
        port = 8079
        path = f'object_relation/{relation_id}'
        data = {
            'name': rel_data.get('name', ''),
            'left_name': rel_data.get('left_name', ''),
            'right_name': rel_data.get('right_name', ''),
            'left_description': rel_data.get('left_description', ''),
            'right_description': rel_data.get('right_description', ''),
        }
        response = self._request('PUT', path, port=port, data=data)
        return response.json()

    def delete_relation(self, relation_id: str) -> dict:
        """
        删除模型关系定义

        EasyOps API: DeleteRelation
        服务: logic.cmdb.service

        :param relation_id: 关系ID
        :return: 删除结果
        :rtype: dict
        """
        port = 8079
        response = self._request('DELETE', f'object_relation/{relation_id}', port=port)
        result = response.json()
        if result.get('code') == 0:
            logger.info(f"Deleted relation: {relation_id}")
        else:
            logger.error(f"Failed to delete relation {relation_id}: {result.get('message')}")
        return result

    def batch_archive_instance(self, object_id: str, instance_ids: list, batch_size: int = 1000) -> dict:
        """
        批量归档(删除)实例

        EasyOps API: BatchArchiveInstance
        服务: logic.cmdb.service

        :param object_id: 实例所属的模型ID
        :param instance_ids: 实例Id列表
        :param batch_size: 每批次处理的数量,默认1000
        :return: 归档结果统计
        :rtype: dict
        """
        port = 8079
        path = f'object/{object_id}/instance_archive_instances'
        total_archived = 0
        total_failed = 0

        for batch in self._batch_process(instance_ids, batch_size):
            data = {'instanceIds': batch}
            response = self._request('POST', path, port=port, data=data)
            result = response.json()
            if result.get('code') == 0:
                total_archived += len(batch)
            else:
                total_failed += len(batch)
                logger.error(f"Failed to archive batch: {result.get('message')}")

        logger.info(f"Archived {total_archived} instances of {object_id}, failed {total_failed}.")
        return {'archived': total_archived, 'failed': total_failed}

    def delete_instance_batch(self, object_id: str, instance_ids: list, batch_size: int = 100) -> dict:
        """
        批量删除实例

        EasyOps API: DeleteInstanceBatch
        服务: logic.cmdb.service

        :param object_id: 实例所属的模型ID
        :param instance_ids: 实例Id列表
        :param batch_size: 每批次处理的数量,默认1000
        :return: 删除结果统计
        :rtype: dict
        """
        port = 8079
        path = f'object/{object_id}/instance/_batch'
        total_deleted = 0
        total_failed = 0

        for batch in self._batch_process(instance_ids, batch_size):
            # instanceIds 用分号隔开
            params = {'instanceIds': ';'.join(batch)}
            response = self._request('DELETE', path, port=port, params=params)
            result = response.json()
            if result.get('code') == 0:
                failed_list = result.get('data', {}).get('deleteFailedInstances', [])
                total_deleted += len(batch) - len(failed_list)
                total_failed += len(failed_list)
            else:
                total_failed += len(batch)
                logger.error(f"Failed to delete batch: {result.get('message')}")

        logger.info(f"Deleted {total_deleted} instances of {object_id}, failed {total_failed}.")
        return {'deleted': total_deleted, 'failed': total_failed}

    def delete_object(self, object_id: str) -> dict:
        """
        删除模型

        EasyOps API: DeleteObject
        服务: logic.cmdb.service

        :param object_id: 模型ID
        :return: 删除结果
        :rtype: dict
        """
        port = 8079
        response = self._request('DELETE', f'object/{object_id}', port=port)
        result = response.json()
        if result.get('code') == 0:
            logger.info(f"Deleted object: {object_id}")
        else:
            logger.error(f"Failed to delete object {object_id}: {result.get('message')}")
        return result

    def cleanup_models_by_keyword(self, keyword: str, dry_run: bool = True) -> dict:
        """
        根据模型ID包含关键字清理数据及模型

        流程:
        1. ListObjectBasic 根据关键字查询模型
        2. get_model_desc 获取模型详情
        3. DeleteRelation 删除模型关系
        4. search_instance 获取所有模型实例的instanceId
        5. DeleteInstanceBatch 删除查询到的实例
        6. DeleteObject 删除模型

        :param keyword: 模型ID包含的关键字
        :param dry_run: 是否为试运行模式,True只打印不执行删除,默认True
        :return: 清理结果统计
        :rtype: dict
        """
        result = {
            'models_found': [],
            'relations_deleted': 0,
            'instances_deleted': 0,
            'models_deleted': 0,
            'errors': []
        }

        # 1. 查询包含关键字的模型
        objects = self.list_object_basic(q=keyword)
        result['models_found'] = [obj['objectId'] for obj in objects]
        logger.info(f"Found {len(objects)} models matching keyword '{keyword}': {result['models_found']}")

        if dry_run:
            logger.info("[DRY RUN] Would delete the following models and their data:")
            for obj in objects:
                logger.info(f"  - {obj['objectId']}: {obj.get('name', '')}")
            return result

        for obj in objects:
            object_id = obj['objectId']
            try:
                # 2. 获取模型详情
                model_desc = self.get_model_desc(object_id)
                relation_list = model_desc.get('relation_list', [])

                # 3. 删除模型关系
                for relation in relation_list:
                    relation_id = relation.get('relation_id')
                    if relation_id:
                        try:
                            self.delete_relation(relation_id)
                            result['relations_deleted'] += 1
                        except Exception as e:
                            result['errors'].append(f"Delete relation {relation_id}: {e}")

                # 4. 获取所有实例
                if not obj['isAbstract']:
                    instances = self.search_instance(object_id, fields=['instanceId'])
                    instance_ids = [inst['instanceId'] for inst in instances]

                    # 5. 批量删除实例
                    if instance_ids:
                        delete_result = self.delete_instance_batch(object_id, instance_ids)
                        result['instances_deleted'] += delete_result['deleted']

                # 6. 删除模型
                self.delete_object(object_id)
                result['models_deleted'] += 1

            except Exception as e:
                result['errors'].append(f"Cleanup model {object_id}: {e}")
                logger.error(f"Error cleaning up model {object_id}: {e}")

        logger.info(f"Cleanup completed: {result['models_deleted']} models, "
                    f"{result['relations_deleted']} relations, {result['instances_deleted']} instances deleted.")
        return result

    def import_model(self, model_info: Union[str, List[Dict[str, Any]]]) -> None:
        """
        将模型信息导入系统。

        EasyOps API: PostObjectImportApi
        服务: logic.cmdb.service

        :param model_info: 模型信息,可以是 JSON 字符串或字典列表。
        :return: 无返回值。
        :rtype: None
        """
        port = 8079
        # 如果是文件路径，先读取文件内容
        if isinstance(model_info, str) and os.path.isfile(model_info):
            with open(model_info, 'r', encoding='utf-8') as f:
                model_info = json.load(f)
        elif isinstance(model_info, str):
            try:
                model_info = json.loads(model_info)
            except json.JSONDecodeError:
                pass

        res = self._request('POST', '/v2/object_import', port=port, data={'object_list': model_info, 'ignore_dst_relation': True})

        if res.status_code == 200 and res.json()['code'] == 0:
            logger.info('Init model {} success'.format([i['objectId'] for i in model_info]))
        else:
            logger.error('Init model {} failed: {}'.format([i['objectId'] for i in model_info], res.json()))

    def list_tools(self, page: int = 1, page_size: int = 300, category: str = None,
                   name: str = None, plugin: bool = False, **kwargs) -> dict:
        """
        批量获取工具信息

        EasyOps API: ListTool
        服务: logic.tool_service

        :param page: 页码，默认1
        :param page_size: 页大小，默认300
        :param category: 工具分类筛选
        :param name: 工具名称筛选
        :param plugin: 是否显示插件，默认False
        :return: 工具列表
        """
        port = 8181
        params = {
            'page': page,
            'pageSize': page_size,
            'plugin': str(plugin).lower(),
        }
        if category:
            params['category'] = category
        if name:
            params['name'] = name
        params.update(kwargs)

        res = self._request('GET', '/tools', port=port, params=params)
        data = res.json()
        if data.get('code') == 0:
            tools = data.get('data', {}).get('list', [])
            total = data.get('data', {}).get('total', 0)
            logger.info(f'Found {len(tools)} tools (total: {total})')
            return data.get('data', {})
        else:
            logger.error(f'List tools failed: {data}')
            return {}

    def export_tools(self, tool_ids: list, output_path: str = None, export_libs: bool = True,
                     version_type: str = '$latest_version', timeout: int = 300) -> str:
        """
        批量导出工具

        EasyOps API: ExportToolBatch
        服务: logic.tool_service

        :param tool_ids: 工具ID列表，每项可以是 toolId 字符串或 {"toolId": "xxx", "versionId": "xxx"} 字典
        :param output_path: 输出文件路径，默认为当前目录下的 tools-{timestamp}.tar.gz
        :param export_libs: 是否导出工具引用的libs库，默认True
        :param version_type: 版本类型，可选值: $latest_version(最新版本), $latest_production(生产版本), $latest_development(开发版本)
        :param timeout: 请求超时时间（秒），默认300秒
        :return: 导出文件路径
        """
        port = 8181

        # 构造导出列表
        export_list = []
        for tool in tool_ids:
            if isinstance(tool, str):
                export_list.append({
                    'toolId': tool,
                    'versionId': version_type,
                    'exportLibs': export_libs
                })
            else:
                export_list.append({
                    'toolId': tool.get('toolId'),
                    'versionId': tool.get('versionId', version_type),
                    'exportLibs': tool.get('exportLibs', export_libs)
                })

        # 发送请求
        res = self._request('POST', '/api/tool_service/v1/batch/export',
                            port=port, data={'exportList': export_list}, timeout=timeout)

        if res is None:
            logger.error('Export tools failed: request returned None')
            return ''

        if res.status_code == 200 and res.headers.get('Content-Type', '').startswith('application/'):
            # 生成输出文件名
            if not output_path:
                from datetime import datetime
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                output_path = f'tools-{timestamp}.tar.gz'

            with open(output_path, 'wb') as f:
                f.write(res.content)
            logger.info(f'Tools exported to: {output_path}')
            return output_path
        else:
            logger.error(f'Export tools failed: {res.text}')
            return ''

    def get_tool_versions(self, tool_id: str, limit: int = 100,
                          env_type: str = None) -> list:
        """
        获取工具的所有版本列表

        EasyOps API: GetToolVersions
        服务: logic.tool_service

        :param tool_id: 工具ID
        :param limit: 每页数量，默认100
        :param env_type: 版本类型过滤（production/development），默认不过滤
        :return: 版本信息列表，每项包含 toolId, vId, vName, vCreateTime, envType 等
        :rtype: list
        """
        port = 8181
        all_versions = []
        paging_time = "9999-12-31"

        while True:
            params = {
                "orderBy": "vCreateTime",
                "orderType": "DESC",
                "startTime": "1970-01-01",
                "endingTime": "9999-12-31",
                "pagingTime": paging_time,
                "limit": limit,
                "fields": "toolId,vId,vName,vCreateTime,envType",
            }
            if env_type:
                params["envType"] = env_type

            res = self._request("GET", f"/tools/{tool_id}/versions",
                                port=port, params=params)
            data = res.json().get("data", {})
            items = data.get("list", [])
            all_versions.extend(items)

            total = data.get("total", 0)
            new_paging = data.get("pagingTime", "")
            logger.info(f"工具 {tool_id} 已获取 {len(all_versions)}/{total} 个版本")

            if len(all_versions) >= total or not items or not new_paging:
                break
            paging_time = new_paging

        return all_versions

    def import_tool_check(self, file_path: str) -> dict:
        """
        导入工具前检查，验证工具包是否存在冲突

        EasyOps API: ImportToolCheck
        服务: logic.tool_service

        :param file_path: 工具包文件路径（.tar.gz 格式）
        :return: 检查结果，包含冲突信息
        :rtype: dict
        """
        import tarfile
        port = 8181

        # 解析工具包获取工具信息
        tools_info = []
        with tarfile.open(file_path, 'r:gz') as tar:
            for member in tar.getmembers():
                if member.name.endswith('/config') or member.name == 'config':
                    f = tar.extractfile(member)
                    if f:
                        config = json.loads(f.read().decode('utf-8'))
                        tools_info.append({
                            'toolId': config.get('toolId', ''),
                            'name': config.get('name', ''),
                            'versionId': config.get('versionId', ''),
                            'versionName': config.get('versionName', '1.0.0')
                        })

        if not tools_info:
            # 尝试直接读取 config 文件
            with tarfile.open(file_path, 'r:gz') as tar:
                names = tar.getnames()
                for name in names:
                    if 'config' in name and not name.endswith('/'):
                        f = tar.extractfile(name)
                        if f:
                            config = json.loads(f.read().decode('utf-8'))
                            tools_info.append({
                                'toolId': config.get('toolId', ''),
                                'name': config.get('name', ''),
                                'versionId': config.get('versionId', ''),
                                'versionName': config.get('versionName', '1.0.0')
                            })
                            break

        data = {'systemImport': False, 'tools': tools_info}
        res = self._request('POST', '/api/tool_service/v1/batch/import/check', port=port, data=data, timeout=30)

        if res.status_code == 200:
            result = res.json()
            if result.get('code') == 0:
                logger.info(f"工具检查完成: {result.get('data', {})}")
                return result.get('data', {})
            else:
                logger.error(f"工具检查失败: {result.get('message')}")
                return result
        else:
            logger.error(f"工具检查请求失败: {res.status_code} - {res.text}")
            return {'error': res.text}

    def import_tool(self, file_path: str, new_name: str = None, new_version_name: str = None) -> dict:
        """
        导入工具包到 EasyOps

        EasyOps API: ImportTool
        服务: logic.tool_service

        :param file_path: 工具包文件路径（.tar.gz 格式）
        :param new_name: 新工具名称（可选，用于解决名称冲突）
        :param new_version_name: 新版本名称（可选，用于解决版本冲突）
        :return: 导入结果
        :rtype: dict
        """
        port = 8181
        path = 'tools/import'

        files = {'file': (os.path.basename(file_path), open(file_path, 'rb'), 'application/gzip')}
        form_data = {}
        if new_name:
            form_data['newName'] = new_name
        if new_version_name:
            form_data['newVersionName'] = new_version_name

        try:
            response = self._request('POST', path, port=port, timeout=60, files=files, form_data=form_data if form_data else None)
            result = response.json()
            if result.get('toolId') or result.get('tool'):
                tool_info = result.get('tool', result)
                logger.info(f"工具导入成功: {tool_info.get('name', 'unknown')} (ID: {tool_info.get('toolId', 'unknown')})")
                return result
            elif result.get('conflictList'):
                logger.warning(f"工具存在冲突: {result.get('conflictList')}")
                return result
            else:
                logger.info(f"工具导入结果: {result}")
                return result
        finally:
            files['file'][1].close()

    def execute_tool(self, tool_id: str, v_id: str, inputs: dict,
                     exec_user: str = "root",
                     verification_code: str = None,
                     windows_only_active_session: bool = False,
                     batch_strategy: dict = None,
                     timeout: int = 30) -> str:
        """
        执行工具，返回执行任务ID (execId)

        EasyOps API: ExecuteTool
        服务: logic.tool_service

        :param tool_id: 工具ID
        :param v_id: 工具版本ID (vId)
        :param inputs: 工具输入参数，结构需对齐工具版本的 inputs 定义，
                       例如指定执行目标主机的 cmdbInstanceFilterForm
        :param exec_user: 执行用户，默认 root
        :param verification_code: 执行前校验验证码（按业务需要填写）
        :param windows_only_active_session: Windows 是否仅活动会话执行，默认 False
        :param batch_strategy: 批量执行策略，默认 {"batchNum":0,"batchInterval":0,"failedStop":true,"enabled":false}
        :param timeout: 请求超时时间（秒），默认30
        :return: 执行任务ID (execId)，失败返回空串
        :rtype: str
        """
        port = 8181  # serviceName: logic.tool_service
        payload = {
            "toolId": tool_id,
            "vId": v_id,
            "execUser": exec_user,
            "inputs": inputs,
            "windowsOnlyActiveSession": windows_only_active_session,
            "batchStrategy": batch_strategy if batch_strategy else {
                "batchNum": 0,
                "batchInterval": 0,
                "failedStop": True,
                "enabled": False,
            },
        }
        if verification_code:
            payload["verificationCode"] = verification_code

        resp = self._request("POST", "/tools/execution", port=port, data=payload, timeout=timeout)
        data = resp.json()
        if data.get("code") == 0:
            exec_id = data.get("data", {}).get("execId", "")
            logger.info(f"工具已触发执行, execId: {exec_id}")
            return exec_id
        else:
            logger.error(f"执行工具失败: {data}")
            return ""

    def get_execute_result(self, exec_id: str, brief: bool = False,
                           target_ids: str = None, task_id: str = None,
                           step_id: str = None, use_target_id_as_key: bool = False,
                           timeout: int = 10) -> dict:
        """
        获取工具执行结果（执行状态、输出、各 agent 执行明细）

        EasyOps API: GetExecuteResult
        服务: logic.tool_service

        :param exec_id: 工具执行任务ID
        :param brief: 精简模式，不展示日志/输出/表格输出，默认 False
        :param target_ids: taskId 列表，用 ";" 分隔
        :param task_id: 所属流程的 taskId
        :param step_id: 在所属流程中的步骤Id
        :param use_target_id_as_key: agentData/agents 的 key 使用 targetId 而非 targetName，默认 False
        :param timeout: 请求超时时间（秒），默认10
        :return: 执行结果，包含 totalStatus（run/success/failed）、agentData、outputs 等
        :rtype: dict
        """
        port = 8181  # serviceName: logic.tool_service
        params = {
            "brief": str(brief).lower(),
            "useTargetIdAsKey": str(use_target_id_as_key).lower(),
        }
        if target_ids:
            params["targetIds"] = target_ids
        if task_id:
            params["taskId"] = task_id
        if step_id:
            params["stepId"] = step_id

        resp = self._request("GET", f"/tools/execution/{exec_id}", port=port, params=params, timeout=timeout)
        data = resp.json()
        if data.get("code") == 0:
            return data.get("data", {})
        else:
            logger.error(f"获取执行结果失败: {data}")
            return {}

    def wait_for_execution(self, exec_id: str, interval: int = 3,
                           max_wait: int = 3600, **kwargs) -> dict:
        """
        轮询等待工具执行完成（totalStatus 非 run 即视为结束）

        基于 get_execute_result 封装，适合在自动化脚本中阻塞等待执行结果。

        :param exec_id: 工具执行任务ID
        :param interval: 轮询间隔（秒），默认3
        :param max_wait: 最大等待时间（秒），默认3600
        :param kwargs: 透传给 get_execute_result 的其他参数（brief、target_ids 等）
        :return: 最终执行结果（totalStatus 为 success/failed/...）
        :rtype: dict
        """
        result = {}
        elapsed = 0
        while elapsed <= max_wait:
            result = self.get_execute_result(exec_id, **kwargs)
            status = result.get("totalStatus", "")
            logger.info(f"执行 {exec_id} 状态: {status} (已等待 {elapsed}s)")
            if status and status != "run":
                return result
            time.sleep(interval)
            elapsed += interval
        logger.warning(f"执行 {exec_id} 等待超时（{max_wait}s）")
        return result

    def create_agent_management_task(self, operation: str, targets: list,
                                     samplers: list = None, add_owner: bool = None,
                                     proxy_ip: str = None, keep_conf: bool = None,
                                     install_path: str = None, https: bool = None,
                                     insecure: bool = None, ca_cert: str = None,
                                     timeout: int = 30) -> str:
        """
        批量新增 agent 管理任务（安装/升级/卸载远端 agent）

        EasyOps API: CreateAgentManagementTask
        服务: logic.agent_admin

        :param operation: 任务动作，如 install/upgrade/uninstall
        :param targets: 目标 agent 列表，每项形如
                        {"ip":"172.0.0.1","port":22,"username":"root","password":"123"}
        :param samplers: 需要安装的 sampler 列表，如 ["easy_metric_sampler"]
        :param add_owner: 是否添加运维负责人
        :param proxy_ip: 代理服务器IP
        :param keep_conf: 升级时是否保留配置
        :param install_path: agent 安装路径
        :param https: 是否使用 HTTPS
        :param insecure: 是否跳过验证证书
        :param ca_cert: CA 证书文件路径
        :param timeout: 请求超时时间（秒），默认30
        :return: 任务ID (data.id)，失败返回空串
        :rtype: str
        """
        port = 8179  # serviceName: logic.agent_admin
        payload = {
            "operation": operation,
            "targets": targets,
        }
        if samplers:
            payload["samplers"] = samplers
        if add_owner is not None:
            payload["add_owner"] = add_owner
        if proxy_ip:
            payload["proxy_ip"] = proxy_ip
        if keep_conf is not None:
            payload["keepConf"] = keep_conf
        if install_path:
            payload["installPath"] = install_path
        if https is not None:
            payload["https"] = https
        if insecure is not None:
            payload["insecure"] = insecure
        if ca_cert:
            payload["ca_cert"] = ca_cert

        resp = self._request("POST", "/api/v1/agent_management/tasks", port=port, data=payload, timeout=timeout)
        data = resp.json()
        if data.get("code") == 0:
            task_id = data.get("data", {}).get("id", "")
            logger.info(f"agent 管理任务已创建, 任务ID: {task_id}")
            return task_id
        else:
            logger.error(f"创建 agent 管理任务失败: {data}")
            return ""

    def get_agent_management_task(self, task_id: str, page: int = 1,
                                  page_size: int = 100, status: str = None,
                                  timeout: int = 10) -> dict:
        """
        批量查询 agent 管理任务执行情况

        EasyOps API: GetAgentTasks
        服务: logic.agent_admin

        注：该接口为 GET /api/v1/agent_management/tasks/:id（接口文档中 curl 示例
            误写为 -X POST，实际以 GET 为准）。

        :param task_id: agent 管理任务ID（新建任务返回的 data.id）
        :param page: 页码，默认1
        :param page_size: 分页大小，默认100
        :param status: 任务状态筛选
        :param timeout: 请求超时时间（秒），默认10
        :return: 任务执行情况
        :rtype: dict
        """
        port = 8179  # serviceName: logic.agent_admin
        params = {
            "page": page,
            "pageSize": page_size,
        }
        if status:
            params["status"] = status

        resp = self._request("GET", f"/api/v1/agent_management/tasks/{task_id}", port=port, params=params, timeout=timeout)
        data = resp.json()
        if data.get("code") == 0:
            return data.get("data", {})
        else:
            logger.error(f"查询 agent 管理任务失败: {data}")
            return {}

    def extract_field_values(self, path: str, query: Optional[dict] = None) -> tuple[str, dict[str, str]]:
        """
        从指定路径提取字段值到实例ID的映射。

        Args:
            path: 模型路径，格式为 "对象ID.字段路径"，如 "HOST.eth.ip"
            query: 额外的查询条件

        Returns:
            (对象ID, {字段值: 实例ID})
        """
        import jsonpath

        obj_id, field_id = path.split('.', 1)
        field_jpath = '[*].'.join(field_id.split('.'))
        root_field = field_id.split('.')[0]

        merged_query = {root_field: {'$exists': True}}
        if query:
            merged_query.update(query)

        insts = self.search_instance(obj_id, fields=[field_id], query=merged_query)
        if not insts:
            return obj_id, {}

        return obj_id, {
            val: inst['instanceId']
            for inst in insts
            for val in (jsonpath.jsonpath(inst, field_jpath) or [])
        }

    def create_instance_relation(
        self,
        source_path: str,
        target_path: str,
        relation_id: str,
        source_query: Optional[dict] = None,
        target_query: Optional[dict] = None
    ) -> None:
        """
        根据字段值匹配创建模型实例间的关系。

        通过比较源模型和目标模型中指定字段的值，自动建立匹配实例之间的关系。

        Args:
            source_path: 源模型路径，格式为 "对象ID.字段路径"，如 "HOST.eth.ip"
            target_path: 目标模型路径，格式同上，如 "BASE_IP@ONEMODEL.address"
            relation_id: 关系ID，用于标识关系类型
            source_query: 源模型的额外查询条件
            target_query: 目标模型的额外查询条件

        Example:
            >>> api.create_relation('HOST.ip', 'BASE_IP@ONEMODEL.address', 'usedIP')
        """

        source_obj_id, source_val2instid = self.extract_field_values(source_path, source_query)
        if not source_val2instid:
            logger.warning(f"No instances found for {source_path} with query: {source_query}")
            return

        _, target_val2instid = self.extract_field_values(target_path, target_query)
        if not target_val2instid:
            logger.warning(f"No instances found for {target_path} with query: {target_query}")
            return

        # 匹配源和目标，按源实例ID分组
        relations = {}
        for val, source_instid in source_val2instid.items():
            if target_instid := target_val2instid.get(val):
                relations.setdefault(source_instid, []).append(target_instid)

        if not relations:
            logger.warning(f"No matching values found between {source_path} and {target_path}")
            return

        insts = [
            {'instanceId': src_id, relation_id: tgt_ids}
            for src_id, tgt_ids in relations.items()
        ]
        self.import_instance(source_obj_id, insts, 'instanceId')

    def create_property(
        self,
        object_id: str,
        attr_id: str,
        attr_name: str,
        value_type: str = "str",
        enum_values: List[str] = None,
        required: str = "false",
        unique: str = "false",
        readonly: str = "false",
        description: str = "",
        tag: List[str] = None
    ) -> dict:
        """
        为模型创建属性

        EasyOps API: easyops.api.cmdb.object_attribute@Create
        服务: logic.cmdb.service

        Args:
            object_id: 模型ID
            attr_id: 属性ID
            attr_name: 属性名称
            value_type: 属性值类型 (str/int/float/bool/enum/enums/arr/date/datetime/json/struct/structs/ip)
            enum_values: 枚举值列表，当 value_type 为 enum 或 enums 时使用
            required: 是否必填 ("true"/"false")
            unique: 是否唯一 ("true"/"false")
            readonly: 是否只读 ("true"/"false")
            description: 属性描述
            tag: 属性分类标签列表

        Returns:
            dict: API响应结果
        """
        port = 8079

        value = {
            "type": value_type,
            "default_type": "value",
            "default": "",
            "mode": "default"
        }

        # 如果是枚举类型，设置枚举值
        if enum_values and value_type in ("enum", "enums"):
            value["regex"] = enum_values

        data = {
            "id": attr_id,
            "name": attr_name,
            "readonly": readonly,
            "required": required,
            "unique": unique,
            "tag": tag or [],
            "description": description,
            "value": value
        }

        response = self._request("POST", f"/object/{object_id}/attr", port=port, data=data)
        return response.json()

    def import_plugin(self, file_path: str, name: str = None, version: str = "1.0.0", category: str = "自定义") -> dict:
        """
        导入监控采集插件

        EasyOps API: ImportPlugin
        服务: logic.collector_plugin_service

        :param file_path: 监控套件zip文件路径
        :param name: 插件名称（可选，默认从zip文件名提取）
        :param version: 插件版本（可选，默认1.0.0）
        :param category: 插件分类（可选，默认"自定义"）
        :return: 导入结果
        :rtype: dict
        """
        port = 8151
        path = 'api/v1/plugin/import'

        # 如果未指定名称，从文件名提取
        if not name:
            name = os.path.splitext(os.path.basename(file_path))[0]

        files = {'attachment': (os.path.basename(file_path), open(file_path, 'rb'), 'application/zip')}
        form_data = {
            'name': name,
            'version': version,
            'category': category
        }

        try:
            response = self._request('POST', path, port=port, timeout=120, files=files, form_data=form_data)
            result = response.json()
            data = result.get('data', result)
            logger.info(f"监控插件导入完成: {data}")
            return data
        finally:
            files['attachment'][1].close()

    def update_plugin(self, plugin_instance_id: str, file_path: str, version: str = "1.0.1") -> dict:
        """
        更新监控采集插件

        EasyOps API: UpdatePluginWithFile
        服务: logic.collector_plugin_service

        :param plugin_instance_id: 插件ID（必填）
        :param file_path: 监控套件zip文件路径
        :param version: 插件版本（可选，默认1.0.1）
        :return: 更新结果
        :rtype: dict
        """
        port = 8151
        path = f'api/v1/plugin/import_update/{plugin_instance_id}'

        files = {'attachment': (os.path.basename(file_path), open(file_path, 'rb'), 'application/zip')}
        form_data = {
            'version': version
        }

        try:
            response = self._request('PUT', path, port=port, timeout=120, files=files, form_data=form_data)
            result = response.json()
            data = result.get('data', result)
            logger.info(f"监控插件更新完成: {data}")
            return data
        finally:
            files['attachment'][1].close()

    def activate_collector_kit(self, plugin_instance_id: str, relate_object_id: str = None,
                                param: List[dict] = None, query: dict = None,
                                centralized_enable: bool = False, host_ids: List[str] = None,
                                collect_agent: str = None, not_require_job: bool = False) -> dict:
        """
        启用/激活采集套件

        EasyOps API: ActivateCollectorKit
        服务: easyops.api.collector_service.job.ActivateCollectorKit

        :param plugin_instance_id: 插件ID（必填）
        :param relate_object_id: 激活模型ID（可选）
        :param param: 启用参数列表，格式 [{"key": "参数名", "value": "参数值", "paramType": "类型"}]
        :param query: 查询条件，格式 {"filter": "过滤条件"}
        :param centralized_enable: 是否集中采集（默认False）
        :param host_ids: 集中采集的实例ID列表
        :param collect_agent: 执行采集的机器IP
        :param not_require_job: 是否不创建采集任务（默认False）
        :return: 激活结果，包含 resources, totalStatus, relateObjectId
        :rtype: dict
        """
        port = 12000
        path = 'api/v1/collector/kit/activate'

        data = {
            'instanceId': plugin_instance_id,
            'centralizedEnable': centralized_enable,
            'notRequireJob': not_require_job
        }

        if relate_object_id:
            data['relateObjectId'] = relate_object_id
        if param:
            data['param'] = param
        if query:
            data['query'] = query
        if host_ids:
            data['hostIds'] = host_ids
        if collect_agent:
            data['collectAgent'] = collect_agent

        # 临时添加 giraffe-contract-name header
        original_headers = self.headers.copy()
        self.headers['giraffe-contract-name'] = 'easyops.api.collector_service.job.ActivateCollectorKit'
        try:
            response = self._request('POST', path, port=port, timeout=60, data=data)
        finally:
            self.headers = original_headers
        result = response.json()

        data = result.get('data', result)
        resources = data.get('resources', [])
        if result.get('code') == 0:
            logger.info('导入成功')
        for res in resources:
            logger.info(f"  - {res.get('type', '')}: {res.get('name', '')} -> {res.get('result', '')} (数量: {res.get('count', 0)})")

        return data

    def list_operation_log(self, start_time: str, end_time: str,
                           system: str = "", topic: str = "",
                           ctime_order: str = "desc", status: str = "",
                           operator: str = "", app_name: str = "",
                           business: str = "", target_name: str = "",
                           query: str = "", user: str = "",
                           page_size: int = 100) -> list:
        """
        查询操作日志（自动翻页返回所有数据）

        EasyOps API: ListOperationLog
        服务: logic.notify

        :param start_time: 开始时间，格式 2006-01-02 15:04:05
        :param end_time: 截止时间
        :param system: 系统名称，如 deploy/job/scheduler
        :param topic: 事件主题，多个逗号分隔
        :param ctime_order: 排序方式 desc/asc
        :param status: 事件状态
        :param operator: 操作人
        :param app_name: 应用名称
        :param business: 系统名称
        :param target_name: 目标名称
        :param query: 关键字搜索
        :param user: 通知相关人
        :param page_size: 每页条数，默认100，最大3000
        :return: 所有符合条件的操作日志列表
        :rtype: list
        """
        port = 8069
        all_data = []
        page = 1

        while True:
            params = {
                "page": page,
                "pageSize": page_size,
                "start_time": start_time,
                "end_time": end_time,
            }
            if ctime_order:
                params["ctime_order"] = ctime_order
            if system:
                params["system"] = system
            if topic:
                params["topic"] = topic
            if status:
                params["status"] = status
            if operator:
                params["operator"] = operator
            if app_name:
                params["app_name"] = app_name
            if business:
                params["business"] = business
            if target_name:
                params["target_name"] = target_name
            if query:
                params["query"] = query
            if user:
                params["user"] = user

            resp = self._request("GET", "/operation/log", port=port, params=params)
            data = resp.json().get("data", {})
            items = data.get("list", [])
            all_data.extend(items)
            total = data.get("total", 0)

            logger.info(f"已获取 {len(all_data)}/{total} 条操作日志")

            if len(all_data) >= total or not items:
                break
            page += 1

        return all_data

    def list_service_catalog(self, parent_id: str = "",
                              category: str = "",
                              ignore_empty: bool = False,
                              hide: bool = False,
                              ignore_permission_check: bool = False,
                              only_need_root: bool = False,
                              page_size: int = 100) -> list:
        """
        获取服务目录列表（自动翻页返回所有数据）

        EasyOps API: ListCatalog
        服务: logic.flowable_service

        :param parent_id: 父目录的 instanceId，为空则返回根目录
        :param category: 服务实例类型，为空返回所有目录
        :param ignore_empty: 忽略没有服务实例的目录
        :param hide: 隐藏或显示
        :param ignore_permission_check: 是否忽略权限点校验
        :param only_need_root: 是否只要根目录
        :param page_size: 每页条数，默认100
        :return: 所有服务目录列表
        :rtype: list
        """
        port = 8134
        all_data = []
        page = 1

        while True:
            params = {
                "page": page,
                "page_size": page_size,
            }
            if parent_id:
                params["parentID"] = parent_id
            if category:
                params["category"] = category
            if ignore_empty:
                params["ignoreEmpty"] = ignore_empty
            if hide:
                params["hide"] = hide
            if ignore_permission_check:
                params["ignorePermissionCheck"] = ignore_permission_check
            if only_need_root:
                params["onlyNeedRoot"] = only_need_root

            resp = self._request("GET", "/api/flowable_service/v1/service_catalog",
                                 port=port, params=params)
            data = resp.json().get("data", {})
            items = data.get("list", [])
            all_data.extend(items)
            total = data.get("total", 0)

            logger.info(f"已获取 {len(all_data)}/{total} 条服务目录")

            if len(all_data) >= total or not items:
                break
            page += 1

        return all_data

    def create_service_catalog(self, name: str, parent_id: str = "",
                                parent_name: str = "",
                                description: str = "",
                                url: str = "",
                                is_target_blank: bool = False,
                                built_in: bool = False) -> dict:
        """
        创建服务目录

        EasyOps API: CreateCatalog
        服务: logic.flowable_service

        :param name: 目录名称（必填）
        :param parent_id: 父目录 instanceId，为空则创建在根目录下
        :param parent_name: 父目录名称（仅展示用）
        :param description: 目录描述
        :param url: 服务链接
        :param is_target_blank: 跳转方式，True 为外链
        :param built_in: 是否为内置，默认 False
        :return: 创建结果，包含 instanceId
        :rtype: dict
        """
        port = 8134
        payload = {"name": name}
        if parent_id:
            payload["parentID"] = parent_id
        if parent_name:
            payload["parentName"] = parent_name
        if description:
            payload["description"] = description
        if url:
            payload["url"] = url
        if is_target_blank:
            payload["isTargetBlank"] = is_target_blank
        if built_in:
            payload["builtIn"] = built_in

        resp = self._request("POST", "/api/flowable_service/v1/service_catalog",
                             port=port, json=payload)
        data = resp.json().get("data", {})
        logger.info(f"服务目录创建成功: {name}, instanceId={data.get('instanceId', '')}")
        return data

    def delete_service_catalog(self, catalog_id: str) -> dict:
        """
        删除服务目录

        EasyOps API: DeleteServiceCatalog
        服务: logic.flowable_service

        :param catalog_id: 服务目录 instanceId（必填）
        :return: 删除结果
        :rtype: dict
        """
        port = 8134
        resp = self._request("DELETE",
                             f"/api/flowable_service/v1/service_catalog/{catalog_id}",
                             port=port)
        data = resp.json().get("data", {})
        logger.info(f"服务目录删除成功: catalog_id={catalog_id}")
        return data

    def tsdb_column_query(self, object_ids: List[str], start_time: int,
                          end_time: int, fields: List[str] = None,
                          filter: dict = None, limit: int = None,
                          next_token: str = None) -> list:
        """
        TsdbColumn数据查询（自动翻页返回所有数据）

        EasyOps API: TsdbColumnQuery
        服务: logic.data_exchange

        :param object_ids: 列存储模型名列表
        :param start_time: 查询开始时间，单位毫秒（包含）
        :param end_time: 查询结束时间，单位毫秒（不包含）
        :param fields: 返回属性列表，需要全部字段时传 ["*"]
        :param filter: 过滤条件，与easy_core查询filter一致
        :param limit: 数据数量大小限制
        :param next_token: 分页token
        :return: 所有查询数据列表
        :rtype: list
        """
        port = 8152
        all_data = []
        token = next_token

        while True:
            body = {
                "database": self.org,
                "object_ids": object_ids,
                "start_time": start_time,
                "end_time": end_time,
            }
            if fields:
                body["fields"] = fields
            if filter:
                body["filter"] = filter
            if limit:
                body["limit"] = limit
            if token:
                body["next_token"] = token

            resp = self._request("POST",
                                 "/api/v1/data_exchange/tsdb_column/query",
                                 port=port, data=body)
            result = resp.json()
            items = result.get("data", [])
            if isinstance(items, dict):
                items = items.get("data", [])
            all_data.extend(items)

            token = result.get("data", {}).get("next_token") if isinstance(result.get("data"), dict) else None
            logger.info(f"已获取 {len(all_data)} 条 TsdbColumn 数据")

            if not token or not items:
                break

        return all_data

    def tsdb_column_object_get_base_list(self, object_ids: List[str] = None,
                                          is_local: bool = False) -> list:
        """
        TsdbColumnObject批量获取定义

        EasyOps API: TsdbColumnObjectGetBaseList
        服务: logic.data_exchange

        :param object_ids: 模型名列表，为空时获取全部
        :param is_local: 是否查询本地
        :return: 列模型基础信息列表
        :rtype: list
        """
        port = 8152
        body = {"database": self.org}
        if object_ids:
            body["object_ids"] = object_ids
        if is_local:
            body["is_local"] = is_local

        resp = self._request("POST",
                             "/api/v1/data_exchange/tsdb_column_object/get_base_list",
                             port=port, data=body)
        result = resp.json()
        return result.get("base_list", [])

    def tsdb_column_update_by_filter(self, object_id: str, filter: dict,
                                      update_data: dict) -> int:
        """
        TsdbColumn数据批量更新

        EasyOps API: TsdbColumnUpdateByFilter
        服务: logic.data_exchange

        :param object_id: 模型名
        :param filter: 查询条件
        :param update_data: 要更新的数据
        :return: 修改的数量
        :rtype: int
        """
        port = 8152
        body = {
            "database": self.org,
            "object_id": object_id,
            "filter": filter,
            "update_data": update_data,
        }
        resp = self._request("PUT",
                             "/api/v1/data_exchange/tsdb_column/update_by_filter",
                             port=port, data=body)
        result = resp.json()
        count = result.get("data", {}).get("update_count", 0)
        logger.info(f"TsdbColumn更新完成: {count} 条")
        return count

    def olap_query(self, model: str, measures: list, dims: list,
                   filters: list, query: dict = None,
                   order: list = None, limit: dict = None,
                   epoch: str = "s", **kwargs) -> dict:
        """
        OLAP 数据查询与聚合统计

        EasyOps API: QueryV3
        服务: logic.data_exchange

        :param model: 数据源, cmdb 时为 "cmdb.model", clickhouse 时为 "库名.表名"
        :param measures: 指标聚合规则列表, 如 [{"name":"cpu","function":{"expression":"avg","args":["cpu"]}}]
        :param dims: 聚合维度列表, 如 ["time(auto)", "objectId"]
        :param filters: 过滤条件列表, 如 [{"name":"time","operator":">=","value":"now-1h"}]
        :param query: CMDB 过滤条件
        :param order: 排序规则, 如 [{"name":"time","order":"desc"}]
        :param limit: 分页限制, 如 {"limit": 100, "offset": 0}
        :param epoch: 返回时间单位, "ms" 或 "s", 默认 "s"
        :return: 查询结果, 包含 total, list, from, to, step 等字段
        :rtype: dict
        """
        port = 8152
        body = {
            "model": model,
            "measures": measures,
            "dims": dims,
            "filters": filters,
            "epoch": epoch,
        }
        if query:
            body["query"] = query
        if order:
            body["order"] = order
        if limit:
            body["limit"] = limit
        for k, v in kwargs.items():
            body[k] = v
        resp = self._request("POST", "/api/v3/data_exchange/olap", port=port, data=body)
        result = resp.json()
        total = result.get("data", {}).get("total", 0)
        logger.info(f"OLAP查询完成: {total} 条数据")
        return result.get("data", {})

    # ==================== ITSM 流程管理 ====================

    def list_process_definition(self, name: str = "", category: str = "",
                                page: int = 1, page_size: int = 3000) -> list:
        """
        查询ITSM流程定义列表（自动翻页返回所有数据）

        EasyOps API: ListProcessDefinition
        服务: logic.flowable_service

        :param name: 流程名称（模糊匹配）
        :param category: 流程分类
        :param page: 起始页码，默认1
        :param page_size: 每页条数，默认3000
        :return: 所有符合条件的流程定义列表
        :rtype: list
        """
        port = 8134
        all_data = []
        while True:
            params = {"page": page, "pageSize": page_size}
            if name:
                params["name"] = name
            if category:
                params["category"] = category
            resp = self._request("GET", "/api/flowable_service/v1/process_definition",
                                 port=port, params=params)
            data = resp.json().get("data", {})
            items = data.get("list", [])
            all_data.extend(items)
            total = data.get("total", 0)
            logger.info(f"已获取 {len(all_data)}/{total} 条流程定义")
            if len(all_data) >= total or not items:
                break
            page += 1
        return all_data

    def get_process_definition_versions(self, definition_id: str,
                                        page: int = 1,
                                        page_size: int = 3000) -> list:
        """
        获取流程定义的版本列表（自动翻页返回所有数据）

        EasyOps API: GetProcessDefinitionVersions
        服务: logic.flowable_service

        :param definition_id: 流程定义实例Id
        :param page: 起始页码，默认1
        :param page_size: 每页条数，默认3000
        :return: 所有版本列表
        :rtype: list
        """
        port = 8134
        all_data = []
        while True:
            params = {"page": page, "pageSize": page_size}
            resp = self._request(
                "GET",
                f"/api/flowable_service/v1/definition/{definition_id}/version",
                port=port, params=params)
            data = resp.json().get("data", {})
            items = data.get("list", [])
            all_data.extend(items)
            total = data.get("total", 0)
            logger.info(f"已获取 {len(all_data)}/{total} 条版本")
            if len(all_data) >= total or not items:
                break
            page += 1
        return all_data

    def get_process_definition_version_v2(self, definition_id: str,
                                          version_id: str) -> dict:
        """
        获取流程版本V2详情（含节点信息、表单绑定、bpmnXML等）

        EasyOps API: GetProcessDefinitionVersionV2
        服务: logic.flowable_service

        :param definition_id: 流程定义实例Id
        :param version_id: 流程版本实例Id
        :return: 流程版本详情（含taskInfo、bpmnXML等）
        :rtype: dict
        """
        port = 8134
        resp = self._request(
            "GET",
            f"/api/flowable_service/v2/definition/{definition_id}/version/{version_id}",
            port=port)
        return resp.json().get("data", {})

    def get_form_schema_version(self, form_id: str,
                                version_id: str) -> dict:
        """
        获取表单版本详情（含字段定义、标准字段等）

        EasyOps API: GetFormSchemaVersion
        服务: logic.flowable_service

        :param form_id: 表单实例Id
        :param version_id: 表单版本Id
        :return: 表单版本详情
        :rtype: dict
        """
        port = 8134
        resp = self._request(
            "GET",
            f"/api/flowable_service/v1/form/{form_id}/version/{version_id}",
            port=port)
        return resp.json().get("data", {})

    def register_user(self, name: str, password: str, email: str,
                      nickname: str = '', is_admin: bool = False) -> dict:
        """
        注册 EasyOps 用户

        EasyOps API: UserRegister
        服务: logic.user_service

        :param name: 用户名
        :param password: 密码
        :param email: 邮箱
        :param nickname: 昵称（可选）
        :param is_admin: 是否管理员，默认 False
        :return: 注册结果
        :rtype: dict
        """
        port = 8111
        payload = {
            "name": name,
            "password": password,
            "email": email,
            "org": int(self.org),
        }
        if nickname:
            payload["nickname"] = nickname
        if is_admin:
            payload["isAdmin"] = True
        resp = self._request("POST", "/api/v1/users/register", port=port,
                             json=payload)
        return resp.json()

    def import_and_register_users(self, file_path: str,
                                  email_tpl: str = '{name}@rhhn.com',
                                  password_tpl: str = '{name}@easyops2026',
                                  keys: list = None) -> dict:
        """
        从 Excel 导入用户到 CMDB 并批量注册 EasyOps 账号

        流程：
        1. 读取 Excel 获取用户名和昵称列表
        2. 调用 import_instance_excel 导入用户实例到 CMDB
        3. 逐个调用 register_user 注册用户账号

        :param file_path: 用户登记表 Excel 路径
        :param email_tpl: 邮箱模板，{name} 会被替换为用户名，默认 '{name}@rhhn.com'
        :param password_tpl: 密码模板，{name} 会被替换为用户名，默认 '{name}@easyops2026'
        :param keys: CMDB 导入唯一键，默认 ['name']（用户名）
        :return: 导入和注册结果汇总
        :rtype: dict
        """
        import openpyxl

        if not keys:
            keys = ['name']

        # 步骤1：读取 Excel 获取用户列表
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # 第1行表头（中文列名），第2行子表头，第3行说明，数据从第4行开始
        users = []
        for row in ws.iter_rows(min_row=4, values_only=True):
            name = row[0]  # 第一列：用户名
            if not name or not str(name).strip():
                continue
            nickname = row[2] if len(row) > 2 and row[2] else ''  # 第三列：用户昵称
            users.append({
                'name': str(name).strip(),
                'nickname': str(nickname).strip() if nickname else ''
            })
        wb.close()
        logger.info(f"从 Excel 读取到 {len(users)} 个用户")

        if not users:
            logger.warning("Excel 中没有有效用户数据")
            return {'import_result': None, 'register_results': []}

        # 步骤2：导入用户实例到 CMDB
        import_result = None
        logger.info("开始导入用户到 CMDB...")
        try:
            import_result = self.import_instance_excel(
                object_id='USER@EASYOPS',
                file_path=file_path,
                keys=keys
            )
        except Exception as e:
            logger.error(f"CMDB 导入失败（不影响后续注册）: {e}")
            import_result = {'error': str(e)}

        # 步骤3：批量注册用户账号
        logger.info("开始注册用户账号...")
        register_results = {'success': [], 'failed': []}
        for user in users:
            uname = user['name']
            email = email_tpl.format(name=uname)
            password = password_tpl.format(name=uname)
            try:
                result = self.register_user(
                    name=uname,
                    password=password,
                    email=email,
                    nickname=user.get('nickname', '')
                )
                code = result.get('code', -1)
                if code == 0:
                    register_results['success'].append(uname)
                    logger.info(f"注册成功: {uname}")
                else:
                    register_results['failed'].append({
                        'name': uname,
                        'error': result.get('error', result.get('message', str(result)))
                    })
                    logger.warning(f"注册失败: {uname} - {result.get('error', result.get('message', ''))}")
            except Exception as e:
                register_results['failed'].append({'name': uname, 'error': str(e)})
                logger.error(f"注册异常: {uname} - {e}")

        logger.info(f"注册完成: 成功 {len(register_results['success'])}, "
                     f"失败 {len(register_results['failed'])}")
        return {
            'import_result': import_result,
            'register_results': register_results
        }

    def alter_password(self, name: str, password: str) -> Dict:
        """
        修改用户密码

        API: AlterPassword
        服务: logic.user_service
        端口: 8111

        :param name: 用户名（必填）
        :param password: 新密码（必填）
        :return: API 响应
        :rtype: dict
        """
        port = 8111
        path = "/api/v1/users/alter_password"
        import base64
        data = {
            "name": name,
            "password": base64.b64encode(password.encode("utf-8")).decode("utf-8")
        }
        result = self._request("POST", path, port=port, data=data).json()
        if result.get("code") != 0:
            raise Exception(f"修改密码失败: {result.get('error', result)}")
        logger.info(f"修改密码成功: user={name}")
        return result

    def list_inspection_suite(self, keyword: str = "", page: int = 1,
                             page_size: int = 20) -> list:
        """
        获取巡检套件列表（支持关键字搜索和分页）

        EasyOps API: ListInspectionInfo
        服务: logic.inspection

        :param keyword: 模糊过滤关键字，默认空字符串表示不过滤
        :param page: 页码，默认1
        :param page_size: 每页大小，默认20
        :return: 巡检套件列表
        :rtype: list
        """
        port = 8103
        path = "/api/v1/inspection"
        params = {
            "page": page,
            "pageSize": page_size,
        }
        if keyword:
            params["keyword"] = keyword
        result = self._request("GET", path, port=port, params=params).json()
        data = result.get("data", {})
        items = data.get("list", [])
        total = data.get("total", 0)
        logger.info(f"获取巡检套件列表: {len(items)}/{total} 条")
        return items

    def list_all_inspection_suites(self, keyword: str = "",
                                   page_size: int = 100) -> list:
        """
        获取所有巡检套件（自动翻页）

        EasyOps API: ListInspectionInfo
        服务: logic.inspection

        :param keyword: 模糊过滤关键字
        :param page_size: 每页大小，默认100
        :return: 所有巡检套件列表
        :rtype: list
        """
        port = 8103
        path = "/api/v1/inspection"
        all_data = []
        page = 1
        while True:
            params = {
                "page": page,
                "pageSize": page_size,
            }
            if keyword:
                params["keyword"] = keyword
            result = self._request("GET", path, port=port, params=params).json()
            data = result.get("data", {})
            items = data.get("list", [])
            all_data.extend(items)
            total = data.get("total", 0)
            logger.info(f"已获取 {len(all_data)}/{total} 条巡检套件")
            if len(all_data) >= total or not items:
                break
            page += 1
        return all_data

    def import_inspection_suite(self, file_path: str) -> dict:
        """
        导入巡检套件

        EasyOps API: ImportSuite
        服务: logic.inspection

        :param file_path: 套件 tar 包文件路径
        :return: API 响应
        :rtype: dict
        """
        port = 8103
        path = "/api/v1/inspection-import"
        with open(file_path, "rb") as f:
            files = {"file": (os.path.basename(file_path), f, "application/x-tar")}
            result = self._request("POST", path, port=port, timeout=60, files=files).json()
        logger.info(f"导入巡检套件: {file_path}")
        return result

    def delete_inspection_suite(self, plugin_id: str) -> dict:
        """
        卸载巡检套件

        EasyOps API: DeleteInspectionInfo
        服务: logic.inspection

        :param plugin_id: 套件ID（pluginId），可从 list_inspection_suite 返回的 id 字段获取
        :return: API 响应
        :rtype: dict
        """
        port = 8103
        path = f"/api/v1/inspection/{plugin_id}"
        result = self._request("DELETE", path, port=port).json()
        logger.info(f"卸载巡检套件: pluginId={plugin_id}")
        return result


def list_functions():
    """列出类中所有公开方法及其文档摘要"""
    import inspect
    methods = inspect.getmembers(EasyopsAPI, predicate=inspect.isfunction)
    for name, method in sorted(methods):
        if name.startswith('_'):
            continue
        sig = inspect.signature(method)
        # 去掉 self 参数
        params = [str(p) for p in sig.parameters.values() if p.name != 'self']
        doc = method.__doc__
        # 取 docstring 第一行非空行作为摘要
        summary = ''
        if doc:
            for line in doc.strip().splitlines():
                line = line.strip()
                if line:
                    summary = line
                    break
        print(f"  {name}({', '.join(params)})")
        if summary:
            print(f"      {summary}")
        print()

def main():
    # 命令行执行
    import argparse
    parser = argparse.ArgumentParser(description='EasyOps API CLI')

    # 环境管理命令
    env_group = parser.add_argument_group('环境管理')
    env_group.add_argument('--env-list', action='store_true', help='列出所有环境')
    env_group.add_argument('--env-add', type=str, metavar='ID', help='添加新环境')
    env_group.add_argument('--env-use', type=str, metavar='ID', help='切换当前环境')
    env_group.add_argument('--env-del', type=str, metavar='ID', help='删除环境')
    env_group.add_argument('--env-show', type=str, nargs='?', const='', metavar='ID', help='显示环境详情')
    env_group.add_argument('--env-edit', type=str, metavar='ID', help='编辑环境配置')

    # API 调用参数
    api_group = parser.add_argument_group('API调用')
    api_group.add_argument('--host', type=str, help='EasyOps host (覆盖环境配置)')
    api_group.add_argument('--org', type=str, help='EasyOps org (覆盖环境配置)')
    api_group.add_argument('--user', type=str, help='EasyOps user (覆盖环境配置)')
    api_group.add_argument('--ak', type=str, help='EasyOps access key')
    api_group.add_argument('--sk', type=str, help='EasyOps secret key')
    api_group.add_argument('-f', '--func', type=str, help='Function name')
    api_group.add_argument('-a', '--args', type=str, help='JSON string of arguments')
    api_group.add_argument('--list-func', action='store_true', help='List all available functions')
    api_group.add_argument('--dry-run', action='store_true', help='打印curl命令，不发送请求')
    api_group.add_argument('--local-data', type=str, metavar='DIR',
                           help='离线数据目录路径，包含 {model_id}.json 文件；'
                                '指定后优先从本地查询，无本地数据时回退在线')
    api_group.add_argument('--debug', action='store_true', help='Enable debug logging')

    args = parser.parse_args()

    # 处理环境管理命令
    env_mgr = EnvManager()
    if args.env_list:
        env_mgr.list_envs()
        return
    if args.env_add:
        env_mgr.add_env(args.env_add)
        return
    if args.env_use:
        env_mgr.use_env(args.env_use)
        return
    if args.env_del:
        env_mgr.del_env(args.env_del)
        return
    if args.env_show is not None:
        env_mgr.show_env(args.env_show or None)
        return
    if args.env_edit:
        env_mgr.edit_env(args.env_edit)
        return

    # 处理 API 调用
    if args.list_func:
        list_functions()
        return
    if not args.func:
        parser.error('-f/--func is required when not using --list-func or env commands')
    if args.debug:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.INFO)

    try:
        api = EasyopsAPI(
            host=args.host,
            org=args.org,
            user=args.user,
            ak=args.ak,
            sk=args.sk,
            dry_run=args.dry_run,
            local_data=args.local_data
        )
        func = getattr(api, args.func)
        args_json = args.args
        if args_json:
            args_dict = json.loads(args_json)
            result = func(**args_dict)
        else:
            result = func()
        # dry_run 模式下不打印结果（curl命令已在方法内打印）
        if result and not args.dry_run:
            logger.info(f'Result: {result}')
    except requests.exceptions.ConnectionError as e:
        logger.error(f"连接失败: {e}")
        print_env_help(env_mgr.get_current_env_id(), args.host or env_mgr.get_current_env().get('host') if env_mgr.get_current_env() else None)
        sys.exit(1)
    except Exception as e:
        logger.error(f"执行失败: {e}")
        raise

if __name__ == "__main__":
    main()


