"""CSV ↔ Excel 互转"""

import csv
from pathlib import Path
from .base import BaseConverter, ConvertResult, register


@register
class CsvToExcel(BaseConverter):
    name = "csv-excel"
    source_formats = ["csv", "tsv"]
    target_formats = ["xlsx"]
    description = "CSV/TSV 转 Excel"
    dependencies = ["openpyxl"]

    def convert(self, input_path: Path, output_path: Path, **options) -> ConvertResult:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        delimiter = "\t" if input_path.suffix == ".tsv" else options.get("delimiter", ",")
        encoding = options.get("encoding", "utf-8")

        wb = Workbook()
        ws = wb.active
        ws.title = options.get("sheet_name", input_path.stem)

        with open(input_path, "r", encoding=encoding, newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    # 首行加表头样式
                    if row_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="E8E8E8", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")

        # 自动列宽
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        wb.save(str(output_path))
        return ConvertResult(True, output_path=output_path, message="CSV 已转为 Excel")


@register
class ExcelToCsv(BaseConverter):
    name = "excel-csv"
    source_formats = ["xlsx", "xls"]
    target_formats = ["csv"]
    description = "Excel 转 CSV"
    dependencies = ["openpyxl"]

    def convert(self, input_path: Path, output_path: Path, **options) -> ConvertResult:
        from openpyxl import load_workbook

        wb = load_workbook(str(input_path), read_only=True)
        sheet_name = options.get("sheet_name")
        ws = wb[sheet_name] if sheet_name else wb.active

        encoding = options.get("encoding", "utf-8")
        delimiter = options.get("delimiter", ",")

        with open(output_path, "w", encoding=encoding, newline="") as f:
            writer = csv.writer(f, delimiter=delimiter)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([str(v) if v is not None else "" for v in row])

        wb.close()
        return ConvertResult(True, output_path=output_path, message="Excel 已转为 CSV")
