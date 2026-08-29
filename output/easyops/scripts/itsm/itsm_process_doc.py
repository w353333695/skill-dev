#!/usr/bin/env python3
"""
ITSM 流程说明文档生成脚本

根据流程名称/分类查询 ITSM 流程，获取流程详情（节点、表单、脚本等），
生成 Excel 文档，每个流程一个 Sheet。

支持两种认证方式:
    1. 内网调用（默认）: 通过 agent 配置自动获取 host/org
    2. OpenAPI 调用: 使用 AK/SK 签名认证

使用方法:
    python itsm_process_doc.py --name "主机申请"
    python itsm_process_doc.py --category "主机管理"
    python itsm_process_doc.py --name "主机申请" --host 172.30.0.149 --org 1888
"""

import requests
import json
import logging
import time
import platform
import hashlib
import hmac
import yaml
import io
import re
import xml.etree.ElementTree as ET
from functools import wraps
from typing import List, Dict, Any, Optional, Union
from urllib.parse import urlencode
from pathlib import Path

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def retry(times=3, delay=1, backoff=2):
    """重试装饰器"""
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            attempt = 0
            current_delay = delay
            while attempt < times:
                try:
                    return func(*args, **kwargs)
                except Exception as e:
                    attempt += 1
                    if attempt >= times:
                        logger.error(f"重试{times}次后仍然失败: {e}")
                        raise
                    logger.warning(f"第{attempt}次尝试失败,{current_delay}秒后重试: {e}")
                    time.sleep(current_delay)
                    current_delay *= backoff
        return wrapper
    return decorator


class EasyOpsClient:
    """EasyOps API 客户端"""

    PORT_APP_MAP = {}

    def __init__(self, host: Optional[str] = None, org: Optional[str] = None,
                 user: str = "defaultUser", ak: str = "", sk: str = ""):
        if not host:
            host, org = self.__get_host_and_org()
        self.host = host
        self.org = org
        self.headers = {
            "user": user,
            "org": org,
            "Content-Type": "application/json"
        }
        if ak and sk:
            self.is_openapi = True
            self.ak = ak
            self.sk = sk
            self.headers["Host"] = "openapi.easyops-only.com"
        else:
            self.is_openapi = False

    def __get_host_and_org(self) -> tuple:
        if platform.system().lower() == "windows":
            conf_path = "C:\\easyOps\\agent\\conf\\conf.yaml"
        else:
            conf_path = "/usr/local/easyops/agent/conf/conf.yaml"
        with open(conf_path, 'r') as f:
            dic = yaml.load(f, Loader=yaml.FullLoader)
        org = dic['base']['client_id']
        host = dic['command']['server_groups'][0]['hosts'][0]['ip'].split(',')[0]
        return host, str(org)

    def __signature(self, method: str, uri: str, params: Dict = None,
                    data: str = "{}") -> Dict:
        params = dict(params) if params else {}
        request_time = str(int(time.time()))
        method = method.upper()
        content_type = "application/json" if method in ("POST", "PUT") else ""
        url_param = "".join(f"{k}{params[k]}" for k in sorted(params.keys()))
        content_md5 = ""
        if method in ("POST", "PUT") and data:
            md5 = hashlib.md5()
            md5.update(data.encode("utf-8") if isinstance(data, str) else data)
            content_md5 = md5.hexdigest()
        string_to_sign = "\n".join([
            method, uri, url_param, content_type,
            content_md5, request_time, self.ak
        ]).encode()
        signature = hmac.new(
            self.sk.encode(), string_to_sign, hashlib.sha1
        ).hexdigest()
        params.update({
            "accesskey": self.ak,
            "signature": signature,
            "expires": request_time
        })
        return params

    @retry()
    def _request(self, method: str, path: str, port: int,
                 **kwargs) -> requests.Response:
        data = kwargs.get('data')
        params = kwargs.get('params')
        if data:
            request_body = json.dumps(data)
            del kwargs['data']
        else:
            request_body = None
        method = method.upper()
        headers = self.headers.copy()

        if self.is_openapi:
            app_name = self.PORT_APP_MAP.get(port)
            if not app_name:
                raise ValueError(f"端口 {port} 未在 PORT_APP_MAP 中配置")
            uri = f"/{app_name}/{path.lstrip('/')}"
            url = f"http://{self.host}{uri}"
            sign_params = self.__signature(
                method, uri, params=params, data=request_body or "{}"
            )
            url = url + "?" + urlencode(sign_params)
            params = None
            if method in ("GET", "DELETE"):
                headers.pop("Content-Type", None)
            headers.pop('org', None)
        else:
            url = f"http://{self.host}:{port}/{path.lstrip('/')}"

        response = requests.request(
            method=method, url=url, headers=headers,
            data=request_body, timeout=20, **kwargs
        )
        response.raise_for_status()
        return response

    # ==================== ITSM API ====================

    def list_process_definition(self, name: str = "", category: str = "",
                                page_size: int = 3000) -> list:
        """查询ITSM流程定义列表"""
        port = 8134
        all_data, page = [], 1
        while True:
            params = {"page": page, "pageSize": page_size}
            if name:
                params["name"] = name
            if category:
                params["category"] = category
            resp = self._request("GET", "/api/flowable_service/v1/process_definition",
                                 port=port, params=params)
            data = resp.json().get("data", {})
            items = data.get("list", [])
            all_data.extend(items)
            total = data.get("total", 0)
            logger.info(f"已获取 {len(all_data)}/{total} 条流程定义")
            if len(all_data) >= total or not items:
                break
            page += 1
        return all_data

    def get_process_versions(self, definition_id: str) -> list:
        """获取流程版本列表"""
        port = 8134
        all_data, page = [], 1
        while True:
            params = {"page": page, "pageSize": 3000}
            resp = self._request(
                "GET",
                f"/api/flowable_service/v1/definition/{definition_id}/version",
                port=port, params=params)
            data = resp.json().get("data", {})
            items = data.get("list", [])
            all_data.extend(items)
            if len(all_data) >= data.get("total", 0) or not items:
                break
            page += 1
        return all_data

    def get_version_detail(self, definition_id: str, version_id: str) -> dict:
        """获取流程版本V2详情"""
        port = 8134
        resp = self._request(
            "GET",
            f"/api/flowable_service/v2/definition/{definition_id}/version/{version_id}",
            port=port)
        return resp.json().get("data", {})

    def get_form_version(self, form_id: str, version_id: str) -> dict:
        """获取表单版本详情"""
        port = 8134
        resp = self._request(
            "GET",
            f"/api/flowable_service/v1/form/{form_id}/version/{version_id}",
            port=port)
        return resp.json().get("data", {})

    # ==================== CMDB / 工具 API ====================

    def search_instance(self, model_id: str, fields: list = None,
                        query: dict = None, page_size: int = 3000) -> list:
        """搜索CMDB实例"""
        port = 8079
        all_data, page = [], 1
        body = {
            "fields": fields or ["*"],
            "query": query or {},
            "page": page,
            "page_size": page_size,
        }
        while True:
            body["page"] = page
            resp = self._request(
                "POST", f"/v3/object/{model_id}/instance/_search",
                port=port, data=body)
            data = resp.json().get("data", {})
            items = data.get("list", [])
            all_data.extend(items)
            if len(all_data) >= data.get("total", 0) or not items:
                break
            page += 1
        return all_data

    def list_service_instances(self, page_size: int = 3000) -> list:
        """获取服务目录列表（用于匹配流程与服务ID）"""
        port = 8134
        all_data, page = [], 1
        while True:
            params = {"page": page, "pageSize": page_size}
            resp = self._request(
                "GET", "/api/flowable_service/v1/service_instance",
                port=port, params=params)
            data = resp.json().get("data", {})
            items = data.get("list", [])
            all_data.extend(items)
            total = data.get("total", 0)
            logger.info(f"已获取 {len(all_data)}/{total} 条服务实例")
            if len(all_data) >= total or not items:
                break
            page += 1
        return all_data

    def list_tools(self, tool_ids: list = None) -> list:
        """获取工具信息"""
        port = 8181
        if tool_ids:
            # 按 ID 批量查
            all_data = []
            for tid in tool_ids:
                try:
                    resp = self._request("GET", f"/tools/{tid}", port=port)
                    data = resp.json().get("data", {})
                    if data:
                        all_data.append(data)
                except Exception:
                    pass
            return all_data
        resp = self._request("GET", "/tools", port=port,
                             params={"page": 1, "pageSize": 3000})
        return resp.json().get("data", {}).get("list", [])


# ==================== 文档生成 ====================

# 审批人类型映射
USER_TYPE_MAP = {
    "loginUser": "发起人",
    "specifyUser": "指定用户",
    "lastExec": "上一执行人",
    "specifyGroup": "指定用户组",
    "formField": "表单字段",
    "superior": "上级",
    "deptHead": "部门负责人",
    "lastExecLeader": "上一步执行人领导",
}

# 审批方式映射
APPROVE_TYPE_MAP = {
    "single": "单人审批",
    "countersign": "会签",
    "or": "或签",
}

# 处理方式映射
HANDLING_MAP = {
    "directly": "直接处理",
    "claim_directly": "先认领后处理",
    "auto": "轮流处理",
    "transfer": "先派单，后认领，再处理",
    "delegate": "先派单后处理",
}

# 表单字段类型映射
FIELD_TYPE_MAP = {
    "INPUT": "文本输入",
    "TEXTAREA": "多行文本",
    "SELECT": "下拉选择",
    "MULTISELECT": "多选下拉",
    "RADIO": "单选",
    "CHECKBOX": "复选",
    "DATE": "日期",
    "DATETIME": "日期时间",
    "NUMBERINPUT": "数字输入",
    "MODALSELECT": "弹窗选择",
    "UPLOAD": "文件上传",
    "SWITCH": "开关",
    "CASCADER": "级联选择",
    "TREESELECT": "树选择",
    "RICHTEXT": "富文本",
    "LABEL": "标签",
    "DIVIDER": "分割线",
    "CMDBINSTANCESELECT": "CMDB实例选择",
    "MULTIPLESELECT": "多选",
    "BUTTON": "按钮",
    "ARRATINPUT": "数组输入",
    "DEPARTMENT_SELECTOR": "部门选择",
    "USER_SELECTOR": "用户选择",
    "USER_GROUP_SELECTOR": "用户组选择",
    "COMMONDATE": "日期时间",
    "TIMERANGE": "时间范围",
    "DATERANGE": "日期范围",
    "SLIDER": "滑块",
    "IFRAME": "内嵌页面",
    "CMDBCASCADER": "CMDB级联选择",
    "TIPS": "提示信息",
    "DATAINHERIT": "数据继承",
}


def bpmn_to_png(bpmn_xml: str) -> Optional[bytes]:
    """
    将 BPMN XML 转换为 PNG 流程图（使用 Pillow，无需系统级依赖）。
    解析 BPMNDiagram 中的 Shape/Edge 坐标，直接绘制 PNG。
    """
    if not bpmn_xml:
        return None

    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        logger.warning("需安装 Pillow: pip install Pillow")
        return None

    ns = {
        'bpmn2': 'http://www.omg.org/spec/BPMN/20100524/MODEL',
        'bpmndi': 'http://www.omg.org/spec/BPMN/20100524/DI',
        'dc': 'http://www.omg.org/spec/DD/20100524/DC',
        'di': 'http://www.omg.org/spec/DD/20100524/DI',
    }

    try:
        root = ET.fromstring(bpmn_xml)
    except ET.ParseError:
        return None

    # 收集节点名称
    node_names = {}
    process = root.find('.//bpmn2:process', ns)
    if process is None:
        return None
    for elem in process:
        tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        eid = elem.get('id', '')
        name = elem.get('name', '')
        if eid:
            node_names[eid] = (name, tag)

    # 收集图形信息
    shapes = []
    edges = []
    min_x, min_y = float('inf'), float('inf')
    max_x, max_y = 0, 0

    for shape in root.findall('.//bpmndi:BPMNShape', ns):
        bpmn_elem = shape.get('bpmnElement', '')
        bounds = shape.find('dc:Bounds', ns)
        if bounds is None:
            continue
        x = float(bounds.get('x', 0))
        y = float(bounds.get('y', 0))
        w = float(bounds.get('width', 100))
        h = float(bounds.get('height', 40))
        name, tag = node_names.get(bpmn_elem, ('', ''))
        shapes.append({'id': bpmn_elem, 'x': x, 'y': y, 'w': w, 'h': h,
                        'name': name, 'tag': tag})
        min_x, min_y = min(min_x, x), min(min_y, y)
        max_x, max_y = max(max_x, x + w), max(max_y, y + h)

    for edge in root.findall('.//bpmndi:BPMNEdge', ns):
        waypoints = []
        for wp in edge.findall('di:waypoint', ns):
            waypoints.append((float(wp.get('x', 0)), float(wp.get('y', 0))))
        if waypoints:
            edges.append(waypoints)
            for px, py in waypoints:
                min_x, min_y = min(min_x, px), min(min_y, py)
                max_x, max_y = max(max_x, px), max(max_y, py)

    if not shapes:
        return None

    # 缩放和偏移
    scale = 1.5  # 适中清晰度，避免图片过大
    padding = 50
    img_w = int((max_x - min_x + padding * 2) * scale)
    img_h = int((max_y - min_y + padding * 2) * scale)
    ox = (-min_x + padding) * scale
    oy = (-min_y + padding) * scale

    img = Image.new('RGB', (img_w, img_h), '#FFFFFF')
    draw = ImageDraw.Draw(img)

    # 尝试加载字体
    font = _get_font(int(12 * scale))
    small_font = _get_font(int(10 * scale))

    # 颜色定义
    TASK_FILL = '#E3F2FD'
    TASK_BORDER = '#1976D2'
    START_FILL = '#C8E6C9'
    START_BORDER = '#388E3C'
    END_FILL = '#FFCDD2'
    END_BORDER = '#D32F2F'
    GW_FILL = '#FFF9C4'
    GW_BORDER = '#F9A825'
    LINE_COLOR = '#666666'
    TEXT_COLOR = '#333333'

    # 画连线
    for waypoints in edges:
        scaled = [(x * scale + ox, y * scale + oy) for x, y in waypoints]
        for i in range(len(scaled) - 1):
            draw.line([scaled[i], scaled[i + 1]], fill=LINE_COLOR, width=int(1.5 * scale))
        # 箭头
        if len(scaled) >= 2:
            _draw_arrow(draw, scaled[-2], scaled[-1], LINE_COLOR, int(4 * scale))

    # 画节点
    for s in shapes:
        sx = s['x'] * scale + ox
        sy = s['y'] * scale + oy
        sw = s['w'] * scale
        sh = s['h'] * scale
        tag = s['tag']
        name = s['name']

        if 'startEvent' in tag.lower() or tag == 'startEvent':
            cx, cy = sx + sw / 2, sy + sh / 2
            r = min(sw, sh) / 2
            draw.ellipse([cx - r, cy - r, cx + r, cy + r],
                         fill=START_FILL, outline=START_BORDER, width=int(2 * scale))
            if name:
                _draw_centered_text(draw, name, cx, cy + r + 10 * scale, font, TEXT_COLOR)
        elif 'endEvent' in tag.lower() or tag == 'endEvent':
            cx, cy = sx + sw / 2, sy + sh / 2
            r = min(sw, sh) / 2
            draw.ellipse([cx - r, cy - r, cx + r, cy + r],
                         fill=END_FILL, outline=END_BORDER, width=int(3 * scale))
            if name:
                _draw_centered_text(draw, name, cx, cy + r + 10 * scale, font, TEXT_COLOR)
        elif 'gateway' in tag.lower() or tag.endswith('Gateway'):
            cx, cy = sx + sw / 2, sy + sh / 2
            hw, hh = sw / 2, sh / 2
            diamond = [(cx, sy), (sx + sw, cy), (cx, sy + sh), (sx, cy)]
            draw.polygon(diamond, fill=GW_FILL, outline=GW_BORDER, width=int(1.5 * scale))
            # 网关内部符号区分
            sym_size = min(hw, hh) * 0.4
            lw = max(int(2 * scale), 2)
            if 'exclusive' in tag.lower():
                # 排他网关: ✕ 叉号
                draw.line([(cx - sym_size, cy - sym_size), (cx + sym_size, cy + sym_size)],
                          fill=GW_BORDER, width=lw)
                draw.line([(cx + sym_size, cy - sym_size), (cx - sym_size, cy + sym_size)],
                          fill=GW_BORDER, width=lw)
            elif 'parallel' in tag.lower():
                # 并行网关: ＋ 加号
                draw.line([(cx - sym_size, cy), (cx + sym_size, cy)],
                          fill=GW_BORDER, width=lw)
                draw.line([(cx, cy - sym_size), (cx, cy + sym_size)],
                          fill=GW_BORDER, width=lw)
            elif 'inclusive' in tag.lower():
                # 包容网关: ○ 圆圈
                draw.ellipse([cx - sym_size, cy - sym_size, cx + sym_size, cy + sym_size],
                             outline=GW_BORDER, width=lw)
        else:
            # 圆角矩形
            r = 8 * scale
            draw.rounded_rectangle([sx, sy, sx + sw, sy + sh], radius=r,
                                   fill=TASK_FILL, outline=TASK_BORDER,
                                   width=int(1.5 * scale))
            if name:
                cx, cy = sx + sw / 2, sy + sh / 2
                _draw_centered_text(draw, name, cx, cy, font, TEXT_COLOR, max_width=sw - 10 * scale)

    # 输出 PNG
    buf = io.BytesIO()
    img.save(buf, format='PNG', optimize=True)
    return buf.getvalue()


def _get_font(size: int):
    """尝试加载中文字体"""
    from PIL import ImageFont
    font_paths = [
        '/System/Library/Fonts/PingFang.ttc',
        '/System/Library/Fonts/STHeiti Light.ttc',
        '/System/Library/Fonts/Hiragino Sans GB.ttc',
        '/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc',
        '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',
        'C:\\Windows\\Fonts\\msyh.ttc',
        'C:\\Windows\\Fonts\\simhei.ttf',
    ]
    for fp in font_paths:
        try:
            return ImageFont.truetype(fp, size)
        except (OSError, IOError):
            continue
    return ImageFont.load_default()


def _draw_centered_text(draw, text: str, cx: float, cy: float, font,
                        color: str, max_width: float = 0):
    """居中绘制文本，支持自动换行"""
    if not text:
        return
    if max_width > 0 and font.getlength(text) > max_width:
        mid = len(text) // 2
        lines = [text[:mid], text[mid:]]
    else:
        lines = [text]

    line_height = font.size if hasattr(font, 'size') else 14
    total_h = line_height * len(lines)
    start_y = cy - total_h / 2

    for i, line in enumerate(lines):
        bbox = font.getbbox(line)
        tw = bbox[2] - bbox[0]
        draw.text((cx - tw / 2, start_y + i * line_height), line, fill=color, font=font)


def _draw_arrow(draw, p1, p2, color, size):
    """在 p2 处画箭头"""
    import math
    dx = p2[0] - p1[0]
    dy = p2[1] - p1[1]
    length = math.sqrt(dx * dx + dy * dy)
    if length == 0:
        return
    dx, dy = dx / length, dy / length
    # 箭头两翼
    ax = p2[0] - dx * size
    ay = p2[1] - dy * size
    lx = ax - dy * size * 0.5
    ly = ay + dx * size * 0.5
    rx = ax + dy * size * 0.5
    ry = ay - dx * size * 0.5
    draw.polygon([(p2[0], p2[1]), (lx, ly), (rx, ry)], fill=color)


def _flatten_cascader_options(options: list, prefix: str = "") -> list:
    """递归展平级联选项为可读字符串列表"""
    result = []
    for opt in options:
        label = opt.get('label', '')
        value = opt.get('value', '')
        path = f"{prefix}/{label}" if prefix else label
        children = opt.get('children', [])
        if children:
            result.extend(_flatten_cascader_options(children, path))
        else:
            result.append(f"{path}({value})")
    return result


def parse_form_fields(form_data: dict) -> List[Dict]:
    """
    解析表单定义，提取字段信息。
    返回字段列表，每个字段包含: section, label, type, modelField, required,
    defaultValue, options, regex, placeholder, desc,
    unique, readonly, highLight, desensitization, showFullInfo, hidden
    """
    fields = []
    fd = form_data.get('formDefinition', '')
    if not fd:
        return fields
    if isinstance(fd, str):
        try:
            fd = json.loads(fd)
        except json.JSONDecodeError:
            return fields

    if not isinstance(fd, list):
        return fields

    for section in fd:
        section_name = section.get('name', '')
        section_type = section.get('type', '')
        for prop in section.get('propertys', []):
            opts = prop.get('options', {})
            extra = opts.get('extraProps', {})
            field_attr = extra.get('fieldAttr', [])

            field = {
                'section': section_name,
                'section_type': section_type,
                'label': prop.get('label', ''),
                'type': prop.get('type', ''),
                'type_cn': FIELD_TYPE_MAP.get(prop.get('type', ''), prop.get('type', '')),
                'modelField': prop.get('modelField', ''),
                'required': opts.get('required', extra.get('required', False)),
                'defaultValue': '',
                'options': '',
                'regex': '',
                'placeholder': str(opts.get('placeholder', extra.get('placeholder', ''))),
                'desc': str(opts.get('desc', extra.get('desc', ''))),
                # 扩展字段属性
                'unique': 'only' in field_attr or opts.get('only', False),
                'readonly': 'disabled' in field_attr or opts.get('disabled', False),
                'highLight': opts.get('highLight', False),
                'desensitization': extra.get('isPasswordInput', False) or extra.get('desensitization', False),
                'showFullInfo': opts.get('isMore', False),
                'hidden': not opts.get('enabled', True),
            }

            # defaultValue - 处理各种类型
            dv = opts.get('defaultValue', extra.get('defaultValue', ''))
            if isinstance(dv, bool):
                field['defaultValue'] = "是" if dv else "否"
            elif isinstance(dv, list):
                field['defaultValue'] = ', '.join(str(v) for v in dv) if dv else ''
            elif dv is not None and dv != '':
                field['defaultValue'] = str(dv)

            # 枚举选项 - 同时检查 extraProps.items 和 extraProps.options
            items = extra.get('items', [])
            if not items:
                items = extra.get('options', [])

            if items:
                ptype = prop.get('type', '')
                if ptype == 'CASCADER':
                    # 级联选项展平显示
                    opt_strs = _flatten_cascader_options(items)
                else:
                    opt_strs = []
                    for i in items:
                        label_str = i.get('label', '')
                        value_str = i.get('value', '')
                        is_default = i.get('isDefault', False)
                        entry = f"{label_str}({value_str})"
                        if is_default:
                            entry += " [默认]"
                        opt_strs.append(entry)
                field['options'] = '\n'.join(opt_strs)

            # 正则
            pattern = opts.get('pattern', extra.get('pattern', ''))
            if pattern and opts.get('isEnablePattern', False):
                hint = opts.get('patternErrorHint', '')
                field['regex'] = f"{pattern}" + (f" ({hint})" if hint else '')
            elif pattern:
                field['regex'] = str(pattern)

            fields.append(field)

    return fields


# 网关类型映射
GATEWAY_TYPE_MAP = {
    'exclusiveGateway': '排他网关',
    'parallelGateway': '并行网关',
    'inclusiveGateway': '包容网关',
}


def parse_bpmn_gateways(bpmn_xml: str) -> Dict[str, Dict]:
    """
    从 BPMN XML 中解析网关信息。
    返回: {gateway_bpmn_id: {type, type_cn, conditions: [{name, expression, target_ref}]}}
    同时返回 taskInfo 节点 ID 到网关的映射关系。
    """
    if not bpmn_xml:
        return {}, {}

    ns = {
        'bpmn2': 'http://www.omg.org/spec/BPMN/20100524/MODEL',
    }

    try:
        root = ET.fromstring(bpmn_xml)
    except ET.ParseError:
        return {}, {}

    process = root.find('.//bpmn2:process', ns)
    if process is None:
        return {}, {}

    # 收集网关信息
    gateways = {}
    for elem in process:
        tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if tag.endswith('Gateway') or 'gateway' in tag.lower():
            gw_id = elem.get('id', '')
            gateways[gw_id] = {
                'type': tag,
                'type_cn': GATEWAY_TYPE_MAP.get(tag, tag),
                'conditions': [],
            }

    # 收集 sequenceFlow 中的条件（从网关出发的连线）
    for elem in process:
        tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if tag == 'sequenceFlow':
            src = elem.get('sourceRef', '')
            tgt = elem.get('targetRef', '')
            name = elem.get('name', '')
            if src in gateways and name:
                cond_text = ''
                for child in elem:
                    ctag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
                    if 'condition' in ctag.lower():
                        cond_text = child.text or ''
                gateways[src]['conditions'].append({
                    'name': name,
                    'expression': cond_text,
                    'target_ref': tgt,
                })

    # 建立 taskInfo 节点 ID -> 所属网关的映射
    # taskInfo 中的网关节点（如"排他网关1"）的 node.links.incoming/outgoing 连线
    # 可以关联到 BPMN 中的 sequenceFlow，从而找到对应的网关
    # 简化方案：通过 sequenceFlow 的 source/target 关系，找到网关后面紧跟的 task 节点
    task_gateway_map = {}  # task_bpmn_id -> gateway_info
    for elem in process:
        tag = elem.tag.split('}')[-1] if '}' in elem.tag else elem.tag
        if tag == 'sequenceFlow':
            src = elem.get('sourceRef', '')
            tgt = elem.get('targetRef', '')
            if src in gateways:
                # 网关出发到某个 task 节点
                task_gateway_map[tgt] = gateways[src]

    return gateways, task_gateway_map


def _resolve_display_names(client: EasyOpsClient, tasks: list):
    """
    解析所有节点中的审批人和脚本ID为可读名称。
    - USER name -> nickname（CMDB USER模型）
    - USER_GROUP instanceId -> name（CMDB USER_GROUP模型）
    - 脚本 toolId -> 脚本名称（tool_service）
    查不到的保留原始值。
    """
    # 收集所有需要解析的 ID
    user_names = set()
    group_ids = set()
    script_ids = set()

    for task in tasks:
        node = task.get('node', {})
        setting = task.get('setting') or {}
        for u in node.get('assigneeListUser', []):
            if u:
                user_names.add(u)
        for g in node.get('assigneeGroups', []):
            gid = g.lstrip(':') if isinstance(g, str) else ''
            if gid:
                group_ids.add(gid)
        scripts = setting.get('scriptSettings') or {}
        for phase in ('preScript', 'postScript'):
            s = scripts.get(phase) or {}
            for sid in s.get('scriptIdList', []):
                if sid:
                    script_ids.add(sid)

    # 批量查询 USER nickname
    user_map = {}  # name -> nickname
    if user_names:
        try:
            users = client.search_instance(
                "USER",
                fields=["name", "nickname"],
                query={"name": {"$in": list(user_names)}}
            )
            for u in users:
                nick = u.get('nickname') or u.get('name', '')
                user_map[u.get('name', '')] = nick
            logger.info(f"  解析用户: {len(user_map)}/{len(user_names)} 个")
        except Exception as e:
            logger.warning(f"  查询USER失败: {e}")

    # 批量查询 USER_GROUP name
    group_map = {}  # instanceId -> name
    if group_ids:
        try:
            groups = client.search_instance(
                "USER_GROUP",
                fields=["instanceId", "name"],
                query={"instanceId": {"$in": list(group_ids)}}
            )
            for g in groups:
                group_map[g.get('instanceId', '')] = g.get('name', '')
            logger.info(f"  解析用户组: {len(group_map)}/{len(group_ids)} 个")
        except Exception as e:
            logger.warning(f"  查询USER_GROUP失败: {e}")

    # 批量查询脚本名称
    tool_map = {}  # toolId -> name
    if script_ids:
        try:
            tools = client.list_tools(tool_ids=list(script_ids))
            for t in tools:
                tool_map[t.get('toolId', t.get('id', ''))] = t.get('name', '')
            logger.info(f"  解析脚本: {len(tool_map)}/{len(script_ids)} 个")
        except Exception as e:
            logger.warning(f"  查询工具失败: {e}")

    # 写入解析后的显示名到 task 中
    for task in tasks:
        node = task.get('node', {})
        setting = task.get('setting') or {}

        # 审批人显示名
        parts = []
        for u in node.get('assigneeListUser', []):
            parts.append(user_map.get(u, u))
        for g in node.get('assigneeGroups', []):
            gid = g.lstrip(':') if isinstance(g, str) else ''
            gname = group_map.get(gid, g)
            parts.append(f"[组]{gname}")
        task['_assignee_display'] = ', '.join(parts)

        # 脚本显示名
        scripts = setting.get('scriptSettings') or {}
        script_parts = []
        for phase, label in [('preScript', '前置'), ('postScript', '后置')]:
            s = scripts.get(phase) or {}
            ids = s.get('scriptIdList', [])
            if ids:
                names = [tool_map.get(sid, sid) for sid in ids]
                ops = s.get('operations', [])
                ops_str = f"({','.join(ops)})" if ops else ''
                script_parts.append(f"{label}{ops_str}: {', '.join(names)}")
        task['_script_display'] = '\n'.join(script_parts)


def generate_excel(processes_data: List[Dict], output_path: str,
                   host: str = "", process_service_map: Dict[str, str] = None):
    """
    生成 Excel 文档。
    每个流程一个 Sheet，包含:
    1. 流程SVG图
    2. 流程基本信息
    3. 节点信息表（审批人、审批方式、表单绑定、脚本绑定）
    4. 表单信息表（字段名称、类型、说明、枚举/正则等）
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XlImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # 删除默认sheet
    wb.remove(wb.active)

    # 样式定义
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    section_font = Font(bold=True, size=12)
    section_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    wrap_align = Alignment(wrap_text=True, vertical='top')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for proc in processes_data:
        proc_name = proc['name']
        # Sheet名最长31字符
        sheet_name = proc_name[:31] if len(proc_name) > 31 else proc_name
        ws = wb.create_sheet(title=sheet_name)

        row = 1

        # ---- 流程基本信息 ----
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        cell = ws.cell(row=row, column=1, value=f"流程信息: {proc_name}")
        cell.font = Font(bold=True, size=14)
        cell.fill = section_fill
        row += 1

        # 构建发起链接
        service_id = ""
        definition_id = proc.get('definition_id', '')
        if process_service_map and definition_id:
            service_id = process_service_map.get(definition_id, '')
        launch_url = ""
        if host and service_id:
            launch_url = f"https://{host}/next/itsc-ticket-center/ticket-apply?serviceId={service_id}&isKnowledge=false"

        info_items = [
            ("流程名称", proc_name),
            ("分类", proc.get('category', '')),
            ("创建人", proc.get('creator', '')),
            ("创建时间", proc.get('ctime', '')),
            ("当前版本", proc.get('version_name', '')),
            ("版本说明", proc.get('version_memo', '')),
            ("说明", proc.get('memo', '')),
        ]
        for label, value in info_items:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
            row += 1

        # 发起链接（超链接）
        ws.cell(row=row, column=1, value="发起链接").font = Font(bold=True)
        if launch_url:
            link_cell = ws.cell(row=row, column=2, value=launch_url)
            link_cell.hyperlink = launch_url
            link_cell.font = Font(color="0563C1", underline="single")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        row += 1

        row += 1

        # ---- 流程图 ----
        png_data = proc.get('png_data')
        if png_data:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
            cell = ws.cell(row=row, column=1, value="流程图")
            cell.font = Font(bold=True, size=14)
            cell.fill = section_fill
            row += 1

            try:
                img = XlImage(io.BytesIO(png_data))
                # 缩放到合适大小
                max_w, max_h = 900, 500
                ratio = min(max_w / img.width, max_h / img.height, 1)
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)
                ws.add_image(img, f"A{row}")
                img_rows = max(int(img.height / 15), 5)
                row += img_rows + 1
            except Exception as e:
                ws.cell(row=row, column=1, value=f"[流程图渲染失败: {e}]")
                row += 1

        row += 1

        # ---- 节点信息表 ----
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        cell = ws.cell(row=row, column=1, value="节点信息")
        cell.font = Font(bold=True, size=14)
        cell.fill = section_fill
        row += 1

        node_headers = ["序号", "节点名称", "网关类型", "网关条件",
                        "处理方式", "审批方式",
                        "审批人类型", "审批人", "绑定表单", "绑定脚本"]
        for col, h in enumerate(node_headers, 1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center_align
            c.border = thin_border
        row += 1

        task_gw_map = proc.get('task_gateway_map', {})

        for idx, task in enumerate(proc.get('tasks', []), 1):
            node = task.get('node', {})
            setting = task.get('setting', {})
            form_info = task.get('formInfo', {})

            # 审批人信息（已解析为显示名）
            user_type = USER_TYPE_MAP.get(node.get('userType', ''), node.get('userType', ''))
            assignee_str = task.get('_assignee_display', '')

            # 网关信息 - 通过节点 BPMN ID 查找前置网关
            node_id = node.get('id', '')
            gw_info = task_gw_map.get(node_id, {})
            gw_type_str = gw_info.get('type_cn', '')
            gw_cond_parts = []
            for cond in gw_info.get('conditions', []):
                expr = cond.get('expression', '')
                name = cond.get('name', '')
                if name:
                    entry = name
                    if expr:
                        entry += f": {expr}"
                    gw_cond_parts.append(entry)
            gw_cond_str = '\n'.join(gw_cond_parts)

            # 表单绑定
            form_str = ''
            if form_info.get('formId'):
                form_str = f"{form_info.get('formName', '')} (v{form_info.get('formVersionName', '')})"

            # 脚本绑定（已解析为脚本名称）
            script_str = task.get('_script_display', '')

            values = [
                idx,
                node.get('name', ''),
                gw_type_str,
                gw_cond_str,
                HANDLING_MAP.get(node.get('handling', ''), node.get('handling', '')),
                APPROVE_TYPE_MAP.get(node.get('approveType', ''), node.get('approveType', '')),
                user_type,
                assignee_str,
                form_str,
                script_str,
            ]
            for col, v in enumerate(values, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.alignment = wrap_align
                c.border = thin_border
            row += 1

        row += 2

        # ---- 表单信息 ----
        forms = proc.get('forms', {})
        if forms:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
            cell = ws.cell(row=row, column=1, value="表单信息")
            cell.font = Font(bold=True, size=14)
            cell.fill = section_fill
            row += 1

            for form_key, form_data in forms.items():
                form_schema = form_data.get('formSchema', {})
                form_name = form_schema.get('name', form_key)

                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=17)
                cell = ws.cell(row=row, column=1, value=f"表单: {form_name}")
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
                row += 1

                # 表单基本信息
                form_info_items = [
                    ("表单名称", form_name),
                    ("分类", form_schema.get('category', '')),
                    ("创建人", form_schema.get('creator', '')),
                    ("说明", form_schema.get('memo', '')),
                ]
                for label, value in form_info_items:
                    ws.cell(row=row, column=1, value=label).font = Font(bold=True)
                    ws.cell(row=row, column=2, value=str(value))
                    row += 1
                row += 1

                # 字段表
                field_headers = ["分组", "字段名称", "字段类型", "类型说明",
                                 "字段标识", "必填",
                                 "默认值", "枚举选项", "正则", "占位提示", "说明"]
                for col, h in enumerate(field_headers, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.alignment = center_align
                    c.border = thin_border
                row += 1

                fields = parse_form_fields(form_data)
                for f in fields:
                    values = [
                        f['section'],
                        f['label'],
                        f['type'],
                        f['type_cn'],
                        f['modelField'],
                        "是" if f['required'] else "否",
                        f['defaultValue'],
                        f['options'],
                        f['regex'],
                        f['placeholder'],
                        f['desc'],
                    ]
                    for col, v in enumerate(values, 1):
                        c = ws.cell(row=row, column=col, value=v)
                        c.alignment = wrap_align
                        c.border = thin_border
                    row += 1

                row += 2

        # 调整列宽
        col_widths = [6, 18, 10, 20, 12, 12, 15, 25, 15, 15, 20]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    logger.info(f"文档已保存: {output_path}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description='ITSM 流程说明文档生成')
    parser.add_argument('--name', type=str, default='', help='流程名称（模糊匹配）')
    parser.add_argument('--category', type=str, default='', help='流程分类')
    parser.add_argument('--host', type=str, help='EasyOps host')
    parser.add_argument('--org', type=str, help='EasyOps org')
    parser.add_argument('--user', type=str, default='defaultUser', help='用户名')
    parser.add_argument('--ak', type=str, default='', help='Access Key')
    parser.add_argument('--sk', type=str, default='', help='Secret Key')
    parser.add_argument('-o', '--output', type=str, help='输出文件路径')
    args = parser.parse_args()

    client = EasyOpsClient(args.host, args.org, args.user, args.ak, args.sk)
    api_host = client.host

    # 0. 获取服务列表，构建流程定义ID -> 服务ID的映射
    process_service_map = {}
    try:
        services = client.list_service_instances()
        for svc in services:
            svc_id = svc.get('instanceId', '')
            associated = svc.get('associatedProcess', {})
            # associatedProcess 的 instanceId 即为流程定义ID
            if isinstance(associated, list):
                for ap in associated:
                    proc_def_id = ap.get('instanceId', '') if isinstance(ap, dict) else ''
                    if proc_def_id:
                        process_service_map[proc_def_id] = svc_id
            elif isinstance(associated, dict):
                proc_def_id = associated.get('instanceId', '')
                if proc_def_id:
                    process_service_map[proc_def_id] = svc_id
        logger.info(f"已构建 {len(process_service_map)} 条流程-服务映射")
    except Exception as e:
        logger.warning(f"获取服务列表失败，发起链接将为空: {e}")

    # 1. 查询流程
    logger.info(f"查询流程: name={args.name}, category={args.category}")
    definitions = client.list_process_definition(name=args.name, category=args.category)
    if not definitions:
        logger.warning("未找到匹配的流程")
        return

    logger.info(f"找到 {len(definitions)} 个流程")

    processes_data = []

    for defn in definitions:
        proc_name = defn.get('name', '')
        definition_id = defn.get('instanceId', '')
        logger.info(f"处理流程: {proc_name} ({definition_id})")

        try:
            # 2. 获取版本列表，找最新生产版本（isMain=True 且 state=done）
            versions = client.get_process_versions(definition_id)
        except Exception as e:
            logger.warning(f"流程 {proc_name} 获取版本失败，跳过: {e}")
            continue
        main_version = None
        for v in versions:
            if v.get('isMain') and v.get('state') == 'done':
                main_version = v
                break
        if not main_version:
            # 没有主版本，取第一个 done 的
            for v in versions:
                if v.get('state') == 'done':
                    main_version = v
                    break
        if not main_version:
            logger.warning(f"流程 {proc_name} 没有可用版本，跳过")
            continue

        version_id = main_version.get('vInstanceId', '')
        logger.info(f"  使用版本: {main_version.get('versionName')} ({version_id})")

        # 3. 获取版本详情
        try:
            detail = client.get_version_detail(definition_id, version_id)
        except Exception as e:
            logger.warning(f"流程 {proc_name} 获取版本详情失败，跳过: {e}")
            continue

        # 4. 生成流程图 PNG
        png_data = bpmn_to_png(detail.get('bpmnXML', ''))

        # 4.5 解析 BPMN 网关信息
        _, task_gateway_map = parse_bpmn_gateways(detail.get('bpmnXML', ''))

        # 5. 获取表单详情
        forms = {}
        seen_forms = set()
        for task in detail.get('taskInfo', []):
            fi = task.get('formInfo', {})
            form_id = fi.get('formId', '')
            form_ver_id = fi.get('formVersionId', '')
            if form_id and form_ver_id and form_id not in seen_forms:
                seen_forms.add(form_id)
                logger.info(f"  获取表单: {fi.get('formName', '')} ({form_id})")
                try:
                    form_detail = client.get_form_version(form_id, form_ver_id)
                    forms[form_id] = form_detail
                except Exception as e:
                    logger.error(f"  获取表单失败: {e}")

        # 6. 解析审批人和脚本名称
        _resolve_display_names(client, detail.get('taskInfo', []))

        # 组装数据
        proc_data = {
            'name': proc_name,
            'definition_id': definition_id,
            'category': defn.get('category', ''),
            'creator': defn.get('creator', ''),
            'ctime': defn.get('ctime', ''),
            'memo': detail.get('definition', {}).get('memo', ''),
            'version_name': main_version.get('versionName', ''),
            'version_memo': main_version.get('vMemo', ''),
            'tasks': detail.get('taskInfo', []),
            'png_data': png_data,
            'forms': forms,
            'task_gateway_map': task_gateway_map,
        }
        processes_data.append(proc_data)

    if not processes_data:
        logger.warning("没有可生成文档的流程")
        return

    # 生成输出文件名
    if args.output:
        output_path = args.output
    else:
        name_part = args.name or args.category or "全部流程"
        output_path = f"ITSM流程文档_{name_part}_{time.strftime('%Y%m%d%H%M%S')}.xlsx"

    generate_excel(processes_data, output_path,
                   host=api_host, process_service_map=process_service_map)
    print(f"\n文档已生成: {output_path}")
    print(f"包含 {len(processes_data)} 个流程")


if __name__ == "__main__":
    main()
